"""Excel idxal sehrbazı — ekran 12 «Excel import» (spec §6.1/3).

Üç addım (dizayn `state.step` 1→2→3):

  1. **Fayl yüklə** — `.xlsx/.xlsm`, ≤10 MB, ≤1000 sətir.
  2. **Uyğunlaşdırma** — fayldakı fənn/qrup adları kataloqla tutuşdurulur.
     Tapılmayan ad SİLİNMİR: sətir `*_text` fallback-ı ilə qalır və sarı
     işarələnir («Mətn kimi qalacaq»).
  3. **Nəticə** — idxal hesabatı; təsdiqdən sonra sətirlər QARALAMA sənədə
     əlavə olunur.

Fayl BÜTÖVLÜKDƏ yaddaşa alınmır (``openpyxl`` read-only) və heç bir addım
mövcud sətri ƏZMİR — idxal yalnız ƏLAVƏ edir (``intake/parsing.py`` naxışı).

Başlıqlar rəsmi TAPŞIRIQ şablonunun sütunlarıdır; sütun sırası dəyişsə də
fayl oxunur (başlıq adına görə xəritələnir).
"""

from __future__ import annotations

import unicodedata

from django.apps import apps as django_apps
from django.db import transaction

from core.audit import log_action
from core.constants import AuditAction, OrgUnitType

from ..constants import DegreeLevel, EducationForm, RowKind, Season
from ..models import TeachingTaskRow
from .scoping import WorkloadDenied, ensure_can_manage
from .tasks import resolve_specialty_and_faculty

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_ROWS = 1000
ALLOWED_SUFFIXES = (".xlsx", ".xlsm")

#: Başlıq (normallaşdırılmış) → sətir sahəsi. Rəsmi TAPŞIRIQ şablonu.
HEADER_MAP = {
    "semestr": "season",
    "qruplar": "groups_text",
    "qrup": "groups_text",
    "fenn": "subject_text",
    "fennin adi": "subject_text",
    "ixtisas": "specialty_text",
    "forma": "education_form",
    "tedris formasi": "education_form",
    "seviyye": "degree_level",
    "pille": "degree_level",
    "telebe": "student_count",
    "telebe sayi": "student_count",
    "birl": "union_count",
    "birlesme": "union_count",
    "yarimq": "subgroup_count",
    "yarimqrup": "subgroup_count",
    "muhazire plan": "lecture_plan",
    "muhazire cemi": "lecture_total",
    "seminar plan": "seminar_plan",
    "seminar cemi": "seminar_total",
    "laboratoriya plan": "lab_plan",
    "laboratoriya cemi": "lab_total",
    "meslehet": "consult_hours",
    "imtahan": "exam_hours",
    "buraxilis": "thesis_hours",
    "doktorant": "postgrad_hours",
    "tecrube": "practice_production_hours",
    "cemi": "total_hours",
    "kredit": "credits",
}

_INT_FIELDS = {
    "student_count",
    "union_count",
    "subgroup_count",
    "lecture_plan",
    "lecture_total",
    "seminar_plan",
    "seminar_total",
    "lab_plan",
    "lab_total",
    "consult_hours",
    "exam_hours",
    "thesis_hours",
    "postgrad_hours",
    "practice_production_hours",
    "total_hours",
}

_SEASONS = {"payiz": Season.FALL, "yaz": Season.SPRING, "yay": Season.SUMMER}
_FORMS = {
    "eyani": EducationForm.EYANI,
    "qiyabi": EducationForm.QIYABI,
    "intensiv": EducationForm.INTENSIV,
    "distant": EducationForm.DISTANT,
}
_LEVELS = {
    "bakalavr": DegreeLevel.BACHELOR,
    "magistr": DegreeLevel.MASTER,
    "doktorantura": DegreeLevel.PHD,
}


class ImportFileError(Exception):
    """Faylın ÖZÜ oxunmadı — sətir xətasından fərqli (bütün fayl rədd olunur)."""

    def __init__(self, code: str, message: str):
        super().__init__(code, message)
        self.code = code
        self.message = message


def normalize(value) -> str:
    """«Mühazirə plan» → «muhazire plan» (diakritikasız, kiçik, tək boşluq)."""
    text = str(value or "").strip().lower()
    text = text.replace("ı", "i").replace("ə", "e").replace("ğ", "g").replace("ş", "s")
    text = text.replace("ç", "c").replace("ö", "o").replace("ü", "u")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace(".", " ").replace("/", " ").replace("(", " ").replace(")", " ")
    return " ".join(text.split())


def _to_int(value) -> int:
    text = str(value or "").strip().replace(" ", "").replace("—", "").replace("–", "")
    if not text or text == "-":
        return 0
    try:
        return int(float(text.replace(",", ".")))
    except (TypeError, ValueError):
        return 0


def parse_workbook(upload) -> list[dict]:
    """Yüklənmiş faylı sətir dict-lərinə çevirir (yazma YOXDUR)."""
    name = getattr(upload, "name", "") or ""
    if not any(name.lower().endswith(suffix) for suffix in ALLOWED_SUFFIXES):
        raise ImportFileError("workload.bad_suffix", "Yalnız .xlsx / .xlsm faylı qəbul olunur.")
    size = getattr(upload, "size", 0) or 0
    if size > MAX_UPLOAD_BYTES:
        raise ImportFileError("workload.file_too_big", "Fayl 10 MB-dan böyükdür.")
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover — openpyxl `requirements/base.txt`-dədir
        raise ImportFileError("workload.openpyxl_missing", "Excel oxuyucusu əlçatan deyil.") from exc

    try:
        workbook = load_workbook(upload, read_only=True, data_only=True)
    except Exception as exc:
        raise ImportFileError("workload.unreadable", "Fayl oxunmadı — Excel formatı gözlənilir.") from exc

    sheet = workbook[workbook.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    mapping: dict = {}
    for raw in rows_iter:
        candidate = {}
        for position, cell in enumerate(raw or ()):
            key = HEADER_MAP.get(normalize(cell))
            if key and key not in candidate.values():
                candidate[position] = key
        if "subject_text" in candidate.values():
            mapping = candidate
            break
    if not mapping:
        raise ImportFileError("workload.no_header", "Başlıq sətri tapılmadı — «Fənn» sütunu olmalıdır.")

    parsed: list[dict] = []
    for raw in rows_iter:
        if raw is None or not any(str(cell or "").strip() for cell in raw):
            continue
        record: dict = {}
        for position, key in mapping.items():
            value = raw[position] if position < len(raw) else None
            record[key] = _to_int(value) if key in _INT_FIELDS else str(value or "").strip()
        if not record.get("subject_text"):
            continue
        record["season"] = str(_SEASONS.get(normalize(record.get("season")), Season.FALL))
        record["education_form"] = str(_FORMS.get(normalize(record.get("education_form")), EducationForm.EYANI))
        record["degree_level"] = str(_LEVELS.get(normalize(record.get("degree_level")), DegreeLevel.BACHELOR))
        parsed.append(record)
        if len(parsed) >= MAX_ROWS:
            break
    workbook.close()
    if not parsed:
        raise ImportFileError("workload.empty_file", "Faylda sətir tapılmadı.")
    return parsed


def build_mapping(*, organization, records) -> dict:
    """Addım 2 — fənn və qrup adlarının kataloqla tutuşdurulması."""
    Subject = django_apps.get_model("registrar", "Subject")
    OrgUnit = django_apps.get_model("organizations", "OrgUnit")

    subject_index = {
        normalize(subject.name): subject
        for subject in Subject.objects.filter(organization=organization, is_active=True).only("id", "name", "code")
    }
    group_index = {
        normalize(unit.name): unit
        for unit in OrgUnit.objects.filter(organization=organization, unit_type=OrgUnitType.GROUP, is_active=True).only(
            "id", "name", "path"
        )
    }
    specialty_index = {
        normalize(unit.name): unit
        for unit in OrgUnit.objects.filter(
            organization=organization, unit_type=OrgUnitType.SPECIALTY, is_active=True
        ).only("id", "name", "path")
    }

    entries: list[dict] = []
    matched = 0
    for record in records:
        subject = subject_index.get(normalize(record.get("subject_text")))
        specialty = specialty_index.get(normalize(record.get("specialty_text")))
        group_names = [part.strip() for part in str(record.get("groups_text") or "").split(",") if part.strip()]
        groups = [group_index.get(normalize(name)) for name in group_names]
        record["_subject_id"] = str(subject.pk) if subject else ""
        record["_specialty_id"] = str(specialty.pk) if specialty else ""
        record["_group_ids"] = [str(unit.pk) for unit in groups if unit is not None]
        matched += int(bool(subject))
        entries.append(
            {
                "source": record.get("subject_text"),
                "kind": "subject",
                "target": subject.name if subject else "",
                "credit": record.get("credits") or "",
                "matched": bool(subject),
            }
        )
        for name, unit in zip(group_names, groups):
            entries.append(
                {
                    "source": name,
                    "kind": "group",
                    "target": unit.name if unit is not None else "",
                    "credit": "",
                    "matched": unit is not None,
                }
            )
    return {
        "records": records,
        "entries": entries,
        "matched": matched,
        "unmatched": sum(1 for entry in entries if not entry["matched"]),
        "row_count": len(records),
        "total_hours": sum(int(record.get("total_hours") or 0) for record in records),
    }


@transaction.atomic
def apply_import(*, task, actor, records, request=None) -> dict:
    """Addım 3 — sətirləri QARALAMA sənədə ƏLAVƏ edir (heç nə silinmir)."""
    ensure_can_manage(actor, task.chair_id)
    from .. import state_machine as sm

    if task.status not in sm.OFFICE_EDITABLE:
        raise WorkloadDenied("workload.task_not_editable", "Bu statusda idxal mümkün deyil.")

    created = 0
    text_only = 0
    for record in records:
        specialty_id = record.get("_specialty_id") or None
        specialty, faculty = resolve_specialty_and_faculty(task.organization, specialty_id)
        row = TeachingTaskRow(
            organization=task.organization,
            task=task,
            season=record.get("season") or Season.FALL,
            row_kind=RowKind.TEACHING,
            subject_id=record.get("_subject_id") or None,
            subject_text=record.get("subject_text") or "",
            specialty=specialty,
            specialty_text=record.get("specialty_text") or "",
            faculty=faculty,
            groups_text=record.get("groups_text") or "",
            education_form=record.get("education_form") or EducationForm.EYANI,
            degree_level=record.get("degree_level") or DegreeLevel.BACHELOR,
            student_count=int(record.get("student_count") or 0),
            union_count=max(int(record.get("union_count") or 1), 1),
            subgroup_count=max(int(record.get("subgroup_count") or 1), 1),
            lecture_plan=int(record.get("lecture_plan") or 0),
            lecture_total=int(record.get("lecture_total") or 0),
            seminar_plan=int(record.get("seminar_plan") or 0),
            seminar_total=int(record.get("seminar_total") or 0),
            lab_plan=int(record.get("lab_plan") or 0),
            lab_total=int(record.get("lab_total") or 0),
            consult_hours=int(record.get("consult_hours") or 0),
            exam_hours=int(record.get("exam_hours") or 0),
            thesis_hours=int(record.get("thesis_hours") or 0),
            postgrad_hours=int(record.get("postgrad_hours") or 0),
            practice_production_hours=int(record.get("practice_production_hours") or 0),
            credits=str(record.get("credits") or ""),
            credits_value=_to_int(record.get("credits")),
        )
        row.total_hours = int(record.get("total_hours") or 0) or row.computed_total_hours
        row.save()
        group_ids = record.get("_group_ids") or []
        if group_ids:
            OrgUnit = django_apps.get_model("organizations", "OrgUnit")
            row.groups.set(list(OrgUnit.objects.filter(organization=task.organization, pk__in=group_ids)))
        created += 1
        text_only += int(not record.get("_subject_id"))

    log_action(
        AuditAction.CREATE,
        user=getattr(actor, "user", None),
        organization=task.organization,
        obj=task,
        new_values={"imported": created, "text_only": text_only},
        reason="workload.rows_imported_from_excel",
        request=request,
        resource_type="workload.TeachingTask",
        resource_id=str(task.pk),
        resource_repr=f"{task.chair_id} · {task.academic_year}",
    )
    return {"created": created, "text_only": text_only}


__all__ = [
    "ALLOWED_SUFFIXES",
    "HEADER_MAP",
    "MAX_ROWS",
    "MAX_UPLOAD_BYTES",
    "ImportFileError",
    "apply_import",
    "build_mapping",
    "normalize",
    "parse_workbook",
]
