"""«Dərs yüküm» (müəllim) endpoint-ləri — yalnız-oxu + Excel ixracı."""

from __future__ import annotations

from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_GET

from ..constants import PERM_VIEW
from ..services import teacher_workload_rows, teacher_workload_summary, teacher_years
from ._base import active_organization, actor_for, error, no_org

_EXPORT_HEADERS = (
    "Tədris ili",
    "Semestr",
    "Fənn",
    "Qrup(lar)",
    "Fəaliyyət",
    "Saat",
    "Forma",
    "Səviyyə",
    "Saathesabı",
)


@never_cache
@login_required
@require_GET
def my_rows(request) -> JsonResponse:
    organization = active_organization(request)
    if organization is None:
        return no_org()
    actor = actor_for(request)
    if not actor.has(PERM_VIEW):
        return error("workload.view_denied", "Dərs yükü məlumatına baxış icazəniz yoxdur.", status=403)
    year = request.GET.get("year") or ""
    season = request.GET.get("season") or ""
    rows = teacher_workload_rows(organization=organization, teacher=request.user, academic_year=year, season=season)
    return JsonResponse(
        {
            "ok": True,
            "years": teacher_years(organization=organization, teacher=request.user),
            "academic_year": year,
            "rows": rows,
            "summary": teacher_workload_summary(organization=organization, teacher=request.user, academic_year=year),
        }
    )


@never_cache
@login_required
@require_GET
def my_export(request):
    """Fərdi yük cədvəli — XLSX (openpyxl requirements/base.txt-dədir)."""
    organization = active_organization(request)
    if organization is None:
        return no_org()
    actor = actor_for(request)
    if not actor.has(PERM_VIEW):
        return error("workload.view_denied", "İcazəniz yoxdur.", status=403)
    year = request.GET.get("year") or ""
    rows = teacher_workload_rows(organization=organization, teacher=request.user, academic_year=year)

    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dərs yükü"
    sheet.append(list(_EXPORT_HEADERS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    total = 0
    for row in rows:
        total += int(row["hours"] or 0)
        sheet.append(
            [
                row["academic_year"],
                row["season_label"],
                row["subject"],
                row["groups"],
                row["activity_label"],
                row["hours"],
                row["education_form"],
                row["degree_level"],
                "Bəli" if row["is_hourly_paid"] else "",
            ]
        )
    sheet.append(["", "", "", "", "CƏMİ", total, "", "", ""])
    sheet[sheet.max_row][4].font = Font(bold=True)
    sheet[sheet.max_row][5].font = Font(bold=True)
    for column, width in zip("ABCDEFGHI", (12, 10, 40, 26, 20, 8, 12, 12, 12)):
        sheet.column_dimensions[column].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"ders-yuku-{year or 'hamisi'}.xlsx".replace("/", "-")
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


__all__ = ["my_export", "my_rows"]
