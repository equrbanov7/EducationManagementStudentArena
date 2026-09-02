"""Tələbə idxalı — YÜKLƏNMİŞ FAYLIN oxunması (.xlsx / .csv).

Qapı fail-closed-dur: ölçü həddi, uzantı allow-list-i və sətir tavanı burada
tətbiq olunur. Fayl BÜTÖVLÜKDƏ yaddaşa alınmır — `openpyxl` read-only rejimdə,
CSV isə axınla oxunur; hər iki halda ``MAX_ROWS``-dan sonra oxu dayanır.

Sütunlar başlıq SATRINA görə xəritələnir (bax ``spec.header_index``), yəni
şablonun sütun sırası dəyişsə də fayl oxunur. Başlıq tapılmasa `IntakeFileError`.
"""

from __future__ import annotations

import csv
import io

from django.utils.translation import pgettext

from .spec import SHEET_NAME, columns, header_index, normalize_header

_CTX = "student_intake"

#: Yüklənən faylın yuxarı həddi (bir qəbul siyahısı üçün bol-bol kifayətdir).
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
#: Bir faylda emal olunan maksimum tələbə sətri.
MAX_ROWS = 2000
#: Qəbul olunan uzantılar.
ALLOWED_SUFFIXES = (".xlsx", ".xlsm", ".csv")


class IntakeFileError(Exception):
    """Faylın özü oxunmadı — sətir-səviyyəli xətadan FƏRQLİ (bütün fayl rədd olunur)."""

    def __init__(self, code: str, message: str):
        super().__init__(code, message)
        self.code = code
        self.message = message

    def __str__(self):
        return self.code


def _suffix(name: str) -> str:
    lowered = str(name or "").strip().lower()
    for suffix in ALLOWED_SUFFIXES:
        if lowered.endswith(suffix):
            return suffix
    return ""


def _blank(values) -> bool:
    return not any(str(value or "").strip() for value in values)


def _map_headers(raw_headers) -> dict:
    """Başlıq sətri → ``{sütun indeksi: açar}``. Tanınan açar yoxdursa xəta."""

    index = header_index()
    mapping: dict = {}
    for position, raw in enumerate(raw_headers):
        key = index.get(normalize_header(raw))
        if key and key not in mapping.values():
            mapping[position] = key
    required = {column.key for column in columns() if column.required}
    if not required.issubset(set(mapping.values())):
        missing = sorted(required - set(mapping.values()))
        raise IntakeFileError(
            "intake_headers_missing",
            pgettext(_CTX, "Faylın başlıq sətrində məcburi sütunlar tapılmadı: %s") % ", ".join(missing),
        )
    return mapping


def _rows_from_values(row_iterable, *, mapping) -> list:
    rows: list = []
    for number, raw in row_iterable:
        values = list(raw)
        if _blank(values):
            continue
        record = {column.key: "" for column in columns()}
        for position, key in mapping.items():
            if position < len(values):
                record[key] = str(values[position] if values[position] is not None else "").strip()
        record["_row"] = number
        rows.append(record)
        if len(rows) >= MAX_ROWS:
            break
    return rows


def _read_xlsx(payload: bytes) -> list:
    try:
        from openpyxl import load_workbook
    except Exception:  # pragma: no cover — paket olmayan mühit
        raise IntakeFileError(
            "intake_xlsx_unsupported",
            pgettext(_CTX, "Bu serverdə .xlsx oxunmur — faylı CSV kimi yadda saxlayıb yenidən yükləyin."),
        ) from None
    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception:
        raise IntakeFileError(
            "intake_file_unreadable",
            pgettext(_CTX, "Fayl oxunmadı — zədəli və ya dəstəklənməyən Excel faylıdır."),
        ) from None
    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    mapping = None
    rows_source: list = []
    for number, raw in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = list(raw)
        if mapping is None:
            if _blank(values):
                continue
            mapping = _map_headers(values)
            continue
        rows_source.append((number, values))
        if len(rows_source) > MAX_ROWS + 8:
            break
    if mapping is None:
        raise IntakeFileError("intake_file_empty", pgettext(_CTX, "Fayl boşdur — başlıq sətri tapılmadı."))
    return _rows_from_values(_strip_hint_row(rows_source, mapping), mapping=mapping)


def _read_csv(payload: bytes) -> list:
    for encoding in ("utf-8-sig", "utf-8", "cp1254", "latin-1"):
        try:
            text = payload.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover — latin-1 həmişə açır
        raise IntakeFileError("intake_file_unreadable", pgettext(_CTX, "Fayl oxunmadı — kodlaşdırma tanınmadı."))

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)

    mapping = None
    rows_source: list = []
    for number, values in enumerate(reader, start=1):
        if mapping is None:
            if _blank(values):
                continue
            mapping = _map_headers(values)
            continue
        rows_source.append((number, values))
        if len(rows_source) > MAX_ROWS + 8:
            break
    if mapping is None:
        raise IntakeFileError("intake_file_empty", pgettext(_CTX, "Fayl boşdur — başlıq sətri tapılmadı."))
    return _rows_from_values(_strip_hint_row(rows_source, mapping), mapping=mapping)


def _strip_hint_row(rows_source, mapping):
    """Şablonun İZAH sətrini (başlıqdan sonrakı kursiv sətir) atır.

    İzah sətrində FİN sütununda «7 simvol, A-Z0-9 (məcburi)» kimi mətn olur —
    onu tələbə sətri kimi oxusaydıq hər faylda bir saxta «xəta» görünərdi.
    """

    fin_position = next((position for position, key in mapping.items() if key == "fin"), None)
    result = []
    for offset, (number, values) in enumerate(rows_source):
        if offset == 0 and fin_position is not None and fin_position < len(values):
            candidate = str(values[fin_position] or "").strip()
            if len(candidate) > 7:
                continue
        result.append((number, values))
    return result


def read_rows(uploaded_file) -> list:
    """Yüklənmiş faylı sətir siyahısına çevirir (`_row` = fayldakı sətir nömrəsi)."""

    if uploaded_file is None:
        raise IntakeFileError("intake_file_required", pgettext(_CTX, "Fayl seçilməyib."))
    size = int(getattr(uploaded_file, "size", 0) or 0)
    if size > MAX_UPLOAD_BYTES:
        raise IntakeFileError(
            "intake_file_too_large",
            pgettext(_CTX, "Fayl çox böyükdür (maksimum 5 MB)."),
        )
    suffix = _suffix(getattr(uploaded_file, "name", ""))
    if not suffix:
        raise IntakeFileError(
            "intake_file_type_unsupported",
            pgettext(_CTX, "Yalnız .xlsx və ya .csv faylı qəbul olunur."),
        )
    payload = uploaded_file.read()
    if not payload:
        raise IntakeFileError("intake_file_empty", pgettext(_CTX, "Fayl boşdur."))
    if len(payload) > MAX_UPLOAD_BYTES:
        raise IntakeFileError("intake_file_too_large", pgettext(_CTX, "Fayl çox böyükdür (maksimum 5 MB)."))
    rows = _read_csv(payload) if suffix == ".csv" else _read_xlsx(payload)
    if not rows:
        raise IntakeFileError("intake_no_rows", pgettext(_CTX, "Faylda tələbə sətri tapılmadı."))
    return rows


__all__ = [
    "ALLOWED_SUFFIXES",
    "MAX_ROWS",
    "MAX_UPLOAD_BYTES",
    "IntakeFileError",
    "read_rows",
]
