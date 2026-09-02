"""
Excel (.xlsx) faylından istifadəçiləri toplu yaradır (bütün rollar üçün).

Şablon: `python manage.py make_user_import_template --output x.xlsx`.
Sütunlar: username, ad_soyad, rol, teskilat_slug, vahid, ilkin_parol, qeyd.

Hər sətir üçün:
  * User yaradılır (username + ilkin parol; boşdursa təhlükəsiz parol generasiya olunur);
  * Profile ilk-giriş axınına salınır (``password_change_required=True``,
    ``email_verified=False``) → istifadəçi ilk girişdə email + OTP + yeni parol tələb edir;
  * Membership yaradılır (org rolu + ``scope_unit`` = "vahid" OrgUnit-i).

⚠️ Bu komanda PROD-da QƏSDƏN BAĞLIDIR (``core/management/command_safety.py``) və
bağlı qalır. Universitetin gündəlik TƏLƏBƏ idxalı üçün nəzarətli səth profil
kabinetindəki «Tələbə idxalı» bölməsidir (`user.import` icazəsi, hər sətir audit
olunur) — bax ``apps/accounts/services/intake/``. Parol siyasəti həmin paketlə
paylaşılır (``generate_initial_password``).

Mövcud username-lər ÖTÜRÜLÜR (üzərinə yazılmır). ``--dry-run`` heç nə yazmır.
Yaradılan parolları paylamaq üçün ``--csv`` ilə fayla yazın.

İstifadə:
    python manage.py import_users_from_excel --file /path/users.xlsx --csv /tmp/creds.csv
    python manage.py import_users_from_excel --file /path/users.xlsx --dry-run
"""

import csv

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.accounts.services.intake import generate_initial_password
from core.management.command_safety import ProductionCommandSafetyMixin
from core.roles import ProfileRole

User = get_user_model()

# Membership rol adı → denormalizasiya olunmuş Profile.role (has_role üçün).
# Tanınmayan rollar üçün MEMBER (Membership əsas mənbədir).
_ROLE_TO_PROFILE = {
    "student": ProfileRole.STUDENT,
    "lead_student": ProfileRole.LEAD_STUDENT,
    "teacher": ProfileRole.TEACHER,
    "instructor": ProfileRole.TEACHER,
    "assistant": ProfileRole.ASSISTANT_TEACHER,
    "hr": ProfileRole.HR,
    "owner": ProfileRole.ORG_OWNER,
    "rector": ProfileRole.ORG_ADMIN,
    "vice_rector": ProfileRole.ORG_ADMIN,
    "dean": ProfileRole.ORG_ADMIN,
    "director": ProfileRole.ORG_ADMIN,
    "chair_head": ProfileRole.ORG_ADMIN,
}

_COLUMNS = ("username", "ad_soyad", "rol", "teskilat_slug", "vahid", "ilkin_parol", "qeyd")


class Command(ProductionCommandSafetyMixin, BaseCommand):
    safety_command_name = "import_users_from_excel"
    help = "Excel faylından istifadəçiləri toplu yaradır (rol + vahid + ilk-giriş axını)."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Doldurulmuş .xlsx faylının yolu")
        parser.add_argument("--csv", dest="csv_path", help="Yaradılan username,parol siyahısını bu fayla yaz")
        parser.add_argument("--sheet", default="İstifadəçilər", help="Vərəqin adı (default: İstifadəçilər)")
        parser.add_argument("--dry-run", action="store_true", help="Heç nə yazma — yalnız nəyin olacağını göstər")

    def handle(self, *args, **options):
        # Request-xarici (management) DB işi RLS transaction-pooling təhlükəsiz
        # kontekstdə + bypass_rls (superadmin aləti, bütün təşkilatlar).
        from core.rls import bypass_rls
        from core.rls_pooling import rls_worker_atomic

        with rls_worker_atomic(), bypass_rls():
            self._import(*args, **options)

    def _read_rows(self, path, sheet_name):
        from openpyxl import load_workbook

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise CommandError(f"Fayl oxunmadı: {exc}")
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Vərəq tapılmadı: {sheet_name!r} (mövcud: {wb.sheetnames})")
        ws = wb[sheet_name]
        rows = []
        for i, raw in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:
                continue  # başlıq
            values = [("" if c is None else str(c).strip()) for c in (list(raw) + [""] * len(_COLUMNS))]
            record = dict(zip(_COLUMNS, values))
            if not record["username"]:
                continue  # boş sətir
            record["_row"] = i
            rows.append(record)
        return rows

    def _import(self, *args, **options):
        from apps.organizations.models import Membership, Organization, OrgUnit

        rows = self._read_rows(options["file"], options["sheet"])
        if not rows:
            self.stdout.write(self.style.WARNING("Faylda istifadəçi sətri tapılmadı."))
            return

        org_cache = {}
        created, skipped, errors = [], [], []

        def resolve_org(slug):
            if slug not in org_cache:
                org_cache[slug] = Organization.objects.filter(slug=slug).first()
            return org_cache[slug]

        for rec in rows:
            username = rec["username"]
            org = resolve_org(rec["teskilat_slug"])
            if org is None:
                errors.append((rec["_row"], username, f"təşkilat tapılmadı: {rec['teskilat_slug']!r}"))
                continue
            role = org.roles.filter(name=rec["rol"], is_active=True).first()
            if role is None:
                errors.append((rec["_row"], username, f"rol tapılmadı: {rec['rol']!r}"))
                continue
            unit = None
            if rec["vahid"]:
                unit = OrgUnit.objects.filter(organization=org, name=rec["vahid"]).first()
                if unit is None:
                    errors.append((rec["_row"], username, f"vahid tapılmadı: {rec['vahid']!r}"))
                    continue
            if User.objects.filter(username=username).exists():
                skipped.append((username, "artıq mövcuddur"))
                continue
            rec["_org"], rec["_role"], rec["_unit"] = org, role, unit
            created.append(rec)

        if options["dry_run"]:
            for rec in created:
                self.stdout.write(
                    f"  + {rec['username']} → {rec['rol']} @ {rec['teskilat_slug']}"
                    + (f" ({rec['vahid']})" if rec["vahid"] else "")
                )
            self._summary(created, skipped, errors, dry=True)
            return

        rows_out = []
        with transaction.atomic():
            for rec in created:
                password = rec["ilkin_parol"] or generate_initial_password()
                first, _, last = rec["ad_soyad"].partition(" ")
                user = User.objects.create_user(username=rec["username"], password=password)
                if first:
                    user.first_name = first
                if last:
                    user.last_name = last
                if first or last:
                    user.save(update_fields=["first_name", "last_name"])

                profile = user.profile
                profile.organization = rec["_org"]
                profile.organization_type = rec["_org"].org_type
                profile.role = _ROLE_TO_PROFILE.get(rec["_role"].name, ProfileRole.MEMBER)
                profile.password_change_required = True
                profile.email_verified = False
                profile.save(
                    update_fields=[
                        "organization",
                        "organization_type",
                        "role",
                        "password_change_required",
                        "email_verified",
                        "updated_at",
                    ]
                )
                Membership.objects.update_or_create(
                    user=user,
                    organization=rec["_org"],
                    role=rec["_role"],
                    scope_unit=rec["_unit"],
                    defaults={"is_primary": True, "is_active": True},
                )
                rows_out.append((rec["username"], password, rec["rol"]))

        if options.get("csv_path") and rows_out:
            with open(options["csv_path"], "w", newline="", encoding="utf-8") as fh:
                writer = csv.writer(fh)
                writer.writerow(["username", "password", "role"])
                writer.writerows(rows_out)
            self.stdout.write(f"CSV yazıldı: {options['csv_path']}")

        self._summary(created, skipped, errors, dry=False)

    def _summary(self, created, skipped, errors, *, dry):
        for username, reason in skipped:
            self.stdout.write(self.style.WARNING(f"  ⤼ {username}: {reason}"))
        for row, username, reason in errors:
            self.stdout.write(self.style.ERROR(f"  ✗ sətir {row} ({username}): {reason}"))
        verb = "yaradılacaq" if dry else "yaradıldı"
        style = self.style.WARNING if dry else self.style.SUCCESS
        self.stdout.write(
            style(
                f"{'DRY-RUN: ' if dry else '✅ '}{len(created)} istifadəçi {verb}; "
                f"{len(skipped)} ötürüldü; {len(errors)} xəta. "
                "İlk girişdə: email + OTP + yeni parol tələb olunacaq."
            )
        )
