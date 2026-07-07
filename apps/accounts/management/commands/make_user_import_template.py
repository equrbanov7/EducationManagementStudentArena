"""
İstifadəçi toplu-yaratma üçün Excel (.xlsx) şablonu yaradır.

Superadmin bu şablonu doldurub istifadəçiləri toplu yaradır (bütün rollar üçün).
Şablonda:
  * "İstifadəçilər" vərəqi — başlıq + hər əsas rol üçün nümunə sətir;
  * "Təlimat" vərəqi — sütunların izahı, etibarlı rol adları və vahid tipləri.

İstifadə:
    python manage.py make_user_import_template --output /path/istifadeci_import.xlsx

Qeydlər:
  * Email ŞABLONDA YOXDUR — istifadəçi ilk girişdə özü təyin edir (email_verified=False).
  * "ilkin_parol" boşdursa sistem avtomatik təhlükəsiz parol yaradır (istifadəçiyə verilməlidir).
  * "vahid" dekan/kafedra müdiri üçün scope (fakültə/kafedra), tələbə üçün qrupdur.
"""

from django.core.management.base import BaseCommand

# Nümunə sətirlər — hər əsas rol yerləşdirməsi (username, ad_soyad, rol, təşkilat_slug, vahid, ilkin_parol, qeyd)
_EXAMPLE_ROWS = [
    ("rektor.aliyev", "Rəşad Əliyev", "rector", "qerbi-kaspi-universiteti", "", "", "Rektor — bütün təşkilat"),
    ("prorektor.ceferova", "Nərmin Cəfərova", "vice_rector", "qerbi-kaspi-universiteti", "", "", "Prorektor"),
    (
        "dekan.mammadov",
        "Elçin Məmmədov",
        "dean",
        "qerbi-kaspi-universiteti",
        "Mühəndislik fakültəsi",
        "",
        "Dekan — fakültə scope",
    ),
    (
        "kafedra.hesenli",
        "Aygün Həsənli",
        "chair_head",
        "qerbi-kaspi-universiteti",
        "Kompüter elmləri kafedrası",
        "",
        "Kafedra müdiri — kafedra scope",
    ),
    ("muellim.quliyev", "Kənan Quliyev", "teacher", "qerbi-kaspi-universiteti", "", "", "Müəllim"),
    (
        "assistent.veliyeva",
        "Səbinə Vəliyeva",
        "assistant",
        "qerbi-kaspi-universiteti",
        "",
        "",
        "Müəllim köməkçisi / assistent",
    ),
    ("merkez.rehber", "İmtahan Mərkəzi", "exam_center", "qerbi-kaspi-universiteti", "", "", "İmtahan mərkəzi rəhbəri"),
    ("hr.ismayilova", "Günay İsmayılova", "hr", "qerbi-kaspi-universiteti", "", "", "İnsan resursları"),
    ("bas.telebe", "Tural Nəbiyev", "lead_student", "qerbi-kaspi-universiteti", "875i qrupu", "", "Baş tələbə"),
    ("telebe01", "Leyla Hüseynova", "student", "qerbi-kaspi-universiteti", "875i qrupu", "", "Tələbə — nümunə"),
    ("telebe02", "Rauf Səfərov", "student", "qerbi-kaspi-universiteti", "842A1 qrupu", "", "Tələbə — nümunə"),
]

_HEADERS = ["username", "ad_soyad", "rol", "teskilat_slug", "vahid", "ilkin_parol", "qeyd"]

_VALID_ROLES = [
    ("rector", "Rektor — bütün təşkilat (org-geniş)"),
    ("vice_rector", "Prorektor — org-geniş"),
    ("owner", "Təşkilat sahibi — org-geniş"),
    ("dean", "Dekan — öz fakültəsi (vahid tələb olunur)"),
    ("chair_head", "Kafedra müdiri — öz kafedrası (vahid tələb olunur)"),
    ("director", "Direktor / departament rəhbəri"),
    ("teacher", "Müəllim"),
    ("assistant", "Müəllim köməkçisi / assistent"),
    ("instructor", "Təlimatçı / müəllim"),
    ("exam_center", "İmtahan mərkəzi rəhbəri"),
    ("hr", "İnsan resursları"),
    ("lead_student", "Baş tələbə"),
    ("student", "Tələbə"),
]

_INSTRUCTIONS = [
    ("Sütun", "İzah"),
    ("username", "Məcburi. Unikal istifadəçi adı (giriş üçün). Boşluqsuz."),
    ("ad_soyad", "Ad və soyad."),
    ("rol", "Məcburi. Aşağıdakı etibarlı rol adlarından biri."),
    ("teskilat_slug", "Təşkilatın slug-ı (məs. qerbi-kaspi-universiteti)."),
    ("vahid", "Akademik vahidin adı. Dekan→fakültə, kafedra müdiri→kafedra, tələbə→qrup. Org-geniş rollar üçün boş."),
    ("ilkin_parol", "Boş → sistem avtomatik yaradır. Doldurulubsa həmin parol qoyulur (istifadəçiyə verilir)."),
    ("qeyd", "Sərbəst qeyd (import-a təsir etmir)."),
    ("", ""),
    ("VACİB", "Email şablonda YOXDUR — istifadəçi ilk girişdə özü təyin edir (email + OTP + yeni parol)."),
    ("VACİB", "Hər istifadəçi ilk girişdə parolu dəyişməli, email təsdiqləməlidir (təhlükəsiz ilkin axın)."),
]


class Command(BaseCommand):
    help = "İstifadəçi toplu-yaratma üçün Excel (.xlsx) şablonu yaradır (bütün rollar üçün nümunə ilə)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            required=True,
            help="Yaradılacaq .xlsx faylının yolu",
        )

    def handle(self, *args, **options):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ── Vərəq 1: İstifadəçilər ──────────────────────────────────────────
        ws = wb.active
        ws.title = "İstifadəçilər"
        header_fill = PatternFill("solid", fgColor="1D4ED8")
        header_font = Font(bold=True, color="FFFFFF")
        for col, name in enumerate(_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for r, row in enumerate(_EXAMPLE_ROWS, start=2):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)
        widths = [18, 22, 16, 26, 30, 14, 34]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"

        # ── Vərəq 2: Təlimat + etibarlı dəyərlər ───────────────────────────
        ref = wb.create_sheet("Təlimat")
        for r, (a, b) in enumerate(_INSTRUCTIONS, start=1):
            ca = ref.cell(row=r, column=1, value=a)
            ref.cell(row=r, column=2, value=b)
            if r == 1 or a == "VACİB":
                ca.font = Font(bold=True)
        start = len(_INSTRUCTIONS) + 3
        ref.cell(row=start, column=1, value="Etibarlı rollar").font = Font(bold=True)
        ref.cell(row=start, column=2, value="İzah").font = Font(bold=True)
        for i, (role, desc) in enumerate(_VALID_ROLES, start=start + 1):
            ref.cell(row=i, column=1, value=role)
            ref.cell(row=i, column=2, value=desc)
        ref.column_dimensions["A"].width = 22
        ref.column_dimensions["B"].width = 70

        wb.save(options["output"])
        self.stdout.write(self.style.SUCCESS(f"Şablon yaradıldı: {options['output']}"))
