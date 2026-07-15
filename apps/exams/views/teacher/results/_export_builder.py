"""İmtahan nəticələri XLSX export builder-i (C2, 2026-07-03).

`export_exam_results_xlsx` view-undan çıxarılıb ki, həm sinxron yol (kiçik
datasetlər), həm də export job worker-i (böyük datasetlər) EYNİ kodu çağırsın.
Davranış birə-birdir. request-dən asılı deyil — filterlənmiş attempts siyahısı
parametr kimi gəlir.
"""

from datetime import datetime
from io import BytesIO

from django.utils import timezone
from django.utils.text import slugify

from apps.exams.public import calculate_test_attempt_result

from ._helpers import (
    _appeal_bonus_map_for,
    _apply_appeal_bonus,
    _attempt_effective_duration,
    _attempt_effective_finish,
    _available_groups_for_exam,
)

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def build_exam_results_xlsx_export(exam, attempts_list):
    """(filename, content_type, bytes) qaytarır — attachment üçün hazır."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - deps quraşdırılıb
        raise RuntimeError("openpyxl is not installed. Add openpyxl to requirements.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = (exam.title[:28] + "…") if len(exam.title) > 30 else exam.title

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    is_test = exam.exam_type == "test"
    is_written_or_coding = exam.exam_type in ("written", "coding")

    headers = [
        "#",
        "Qruplar",
        "Ad Soyad",
        "İstifadəçi adı",
        "E-poçt",
        "Status",
        "Başlama",
        "Bitmə",
        "Müddət (s:dq:sn)",
    ]
    if is_written_or_coding:
        headers += ["Müəllim balı", "Yoxlanıb"]
    if is_test:
        headers += ["Düzgün", "Səhv", "Cavabsız", "Verilmiş sual", "Bal", "Maks. bal", "Faiz", "Apel. bonus"]
    else:
        headers += ["Düzgün", "Səhv", "Verilmiş sual"]

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws.row_dimensions[1].height = 28

    now = timezone.now()
    # Loop-invariant precompute (əvvəl export loop-u içində N+1 idi):
    #  • iştirakçı qruplarının id-ləri — bir dəfə (əvvəl hər attempt-də təkrar).
    #  • hər user üçün üzv olduğu qrup adları — TƏK sorğu ilə dict (əvvəl hər
    #    attempt üçün ayrıca sorğu = N+1).
    available_group_ids = list(_available_groups_for_exam(exam).values_list("id", flat=True))
    _attempt_user_ids = {att.user_id for att in attempts_list}
    # Apellyasiya bonusları (tək sorğu) — export-dakı Bal/Faiz effektiv olsun.
    appeal_bonus_by_attempt = _appeal_bonus_map_for(attempts_list) if is_test else {}
    groups_by_user: dict[int, list[str]] = {}
    if available_group_ids and _attempt_user_ids:
        from apps.exams.models import StudentGroup

        for _uid, _gname in (
            StudentGroup.objects.filter(id__in=available_group_ids, students__id__in=_attempt_user_ids)
            .values_list("students__id", "name")
            .order_by("name")
        ):
            groups_by_user.setdefault(_uid, []).append(_gname)
    for row_idx, att in enumerate(attempts_list, start=2):
        effective_finish, _ = _attempt_effective_finish(att, now=now)
        effective_duration = _attempt_effective_duration(att, effective_finish)

        # İştirakçı qruplardan user-in üzv olduqları — loop-dan əvvəl tək sorğu
        # ilə qurulmuş dict-dən (per-attempt sorğu yox).
        user_groups = ", ".join(groups_by_user.get(att.user_id, []))

        # Müddəti hh:mm:ss formatına çevir
        if effective_duration is not None:
            d_total = max(int(effective_duration), 0)
            d_h, d_m, d_s = d_total // 3600, (d_total % 3600) // 60, d_total % 60
            duration_str = f"{d_h:02d}:{d_m:02d}:{d_s:02d}"
        else:
            duration_str = ""

        row = [
            row_idx - 1,
            user_groups,
            att.user.get_full_name() or att.user.username,
            att.user.username,
            att.user.email or "",
            att.get_status_display(),
            # UTC saxlanan tarixi əvvəlcə yerli zonaya (Asia/Baku) çevir, sonra tz-i
            # at — əks halda Excel UTC divar-saatını göstərir (UI-dən 4 saat geri).
            timezone.localtime(att.started_at).replace(tzinfo=None) if att.started_at else "",
            timezone.localtime(effective_finish).replace(tzinfo=None) if effective_finish else "",
            duration_str,
        ]

        if is_written_or_coding:
            row += [
                att.teacher_score if att.teacher_score is not None else "",
                "Bəli" if att.checked_by_teacher else "Xeyr",
            ]

        if is_test:
            test_result = calculate_test_attempt_result(att, answers=att.answers.all())
            appeal_bonus = appeal_bonus_by_attempt.get(att.id) or 0
            if appeal_bonus:
                test_result = _apply_appeal_bonus(test_result, appeal_bonus)
            row += [
                test_result.correct_count,
                test_result.wrong_count,
                test_result.unanswered_count,
                test_result.delivered_count,
                float(test_result.score_display) if test_result.score_display else 0,
                float(test_result.max_score_display) if test_result.max_score_display else 0,
                float(test_result.percentage_display) if test_result.percentage_display else 0,
                float(appeal_bonus) if appeal_bonus else 0,
            ]
        else:
            delivered = att.correct_count + att.wrong_count
            row += [att.correct_count, att.wrong_count, delivered]

        # Sola düzlənən sütunlar: Qruplar (2), Ad Soyad (3), İstifadəçi adı (4), E-poçt (5), Başlama (7), Bitmə (8)
        left_columns = {2, 3, 4, 5, 7, 8}
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = left if col_idx in left_columns else center
            if isinstance(value, datetime):
                cell.number_format = "DD.MM.YYYY HH:MM"

    # Sütun genişlikləri (header sırasına uyğun)
    # #, Qruplar, Ad Soyad, İstifadəçi adı, E-poçt, Status, Başlama, Bitmə, Müddət
    widths = [5, 26, 26, 20, 28, 16, 20, 20, 16]
    if is_written_or_coding:
        widths += [14, 12]  # Müəllim balı, Yoxlanıb
    if is_test:
        widths += [10, 10, 12, 16, 8, 12, 8, 12]  # Düzgün, Səhv, Cavabsız, Verilmiş, Bal, Maks, Faiz, Apel.
    else:
        widths += [10, 10, 14]  # Düzgün, Səhv, Verilmiş sual
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{slugify(exam.title) or 'exam'}-results-{timezone.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return filename, XLSX_CONTENT_TYPE, buffer.getvalue()
