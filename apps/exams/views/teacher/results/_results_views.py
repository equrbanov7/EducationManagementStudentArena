"""results paketi — müəllim nəticə view funksiyaları (qrup)."""

from datetime import datetime
from io import BytesIO
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from django.utils.translation import pgettext, pgettext_lazy

from apps.exams.models import ExamAnswer, ExamAttempt
from apps.exams.services.access_policy import _ensure_teacher
from apps.exams.services.result_calculation import calculate_test_attempt_result
from apps.exams.services.review_visibility import (
    resolve_exam_attempt_name_visibility as _resolve_attempt_name_visibility,
)
from apps.exams.services.review_visibility import (
    resolve_exam_attempt_review_window_seconds as _resolve_attempt_review_window_seconds,
)
from apps.exams.views.shared.tenant import get_teacher_exam_or_404
from core.permissions import request_has_permission

from ._helpers import (
    _appeal_bonus_map_for,
    _append_query_params,
    _apply_appeal_bonus,
    _apply_results_filters,
    _attempt_effective_duration,
    _attempt_effective_finish,
    _available_groups_for_exam,
    _build_anonymous_name,
    _resolve_attempt_action_state,
    _resolve_profile_navigation,
)


@login_required
def teacher_exam_results(request, slug):
    """
    Müəllim üçün imtahan nəticələri:
    - solda bütün cəhdlər cədvəli
    - aşağıda/sağda seçilmiş cəhdin cavabları + qiymətləndirmə formu
    """
    _ensure_teacher(request.user)
    exam = get_teacher_exam_or_404(request, slug=slug)
    profile_return_url, navigation_params = _resolve_profile_navigation(request, default_section="my-exams")
    exam_navigation_query = urlencode(navigation_params)
    exam_detail_url = _append_query_params(
        reverse("exams:teacher_exam_detail", kwargs={"slug": exam.slug}),
        **navigation_params,
    )

    selected_attempt = None
    selected_answers = None

    # ---------- POST: müəllim bal + feedback saxlayır ----------
    if request.method == "POST":
        if not request_has_permission(request, "grade.input"):
            messages.error(
                request,
                pgettext_lazy("exams.view.results.message", "grading_permission_required"),
            )
            return redirect(request.path)

        attempt_id = request.POST.get("attempt_id")
        score_raw = request.POST.get("teacher_score", "").strip()
        feedback = request.POST.get("teacher_feedback", "").strip()
        from apps.notifications.public import notify_student_about_feedback

        selected_attempt = get_object_or_404(ExamAttempt, id=attempt_id, exam=exam)

        if score_raw:
            try:
                score_val = int(score_raw)
            except ValueError:
                messages.error(request, pgettext_lazy("exams.view.results.message", "score_must_be_integer"))
            else:
                if 0 <= score_val <= 100:
                    selected_attempt.teacher_score = score_val
                    selected_attempt.teacher_feedback = feedback
                    selected_attempt.mark_checked()
                    notify_student_about_feedback(
                        task=exam,
                        student=selected_attempt.user,
                        task_kind="exam",
                        extra_metadata={"attempt_id": selected_attempt.id},
                    )
                    messages.success(request, pgettext_lazy("exams.view.results.message", "score_feedback_saved"))
                    return redirect(
                        _append_query_params(
                            request.path,
                            attempt=selected_attempt.id,
                            **navigation_params,
                        )
                    )
                else:
                    messages.error(request, pgettext_lazy("exams.view.results.message", "score_range_0_100"))
        else:
            # yalnız feedback saxlanılır
            selected_attempt.teacher_score = None
            selected_attempt.teacher_feedback = feedback
            selected_attempt.checked_by_teacher = False
            selected_attempt.save(
                update_fields=[
                    "teacher_score",
                    "teacher_feedback",
                    "checked_by_teacher",
                ]
            )
            notify_student_about_feedback(
                task=exam,
                student=selected_attempt.user,
                task_kind="exam",
                extra_metadata={"attempt_id": selected_attempt.id},
            )
            messages.success(request, pgettext_lazy("exams.view.results.message", "feedback_saved"))
            return redirect(
                _append_query_params(
                    request.path,
                    attempt=selected_attempt.id,
                    **navigation_params,
                )
            )

    attempts, filter_state = _apply_results_filters(exam, request)
    search_query = filter_state["search_query"]
    status_filter = filter_state["status_filter"]
    checked_filter = filter_state["checked_filter"]
    date_from_raw = filter_state["date_from_raw"]
    date_to_raw = filter_state["date_to_raw"]
    group_id = filter_state["group_id"]
    sort_by = filter_state["sort_by"]
    sort_dir = filter_state["sort_dir"]

    max_score = 100 if exam.exam_type == "test" else exam.questions.aggregate(total=Sum("points")).get("total") or 0
    pending_count = 0 if exam.exam_type == "test" else attempts.filter(checked_by_teacher=False).count()
    graded_count = attempts.count() if exam.exam_type == "test" else attempts.filter(checked_by_teacher=True).count()
    review_stats = {
        "total": attempts.count(),
        "pending": pending_count,
        "graded": graded_count,
        "max_score": max_score,
    }
    paginator = Paginator(attempts, 12)
    page_obj = paginator.get_page(request.GET.get("page"))
    attempts_page = list(page_obj.object_list)

    # ---------- GET: hansı attempt seçilib? ----------
    if selected_attempt is None:
        attempt_param = request.GET.get("attempt")
        if attempt_param:
            selected_attempt = exam.attempts.filter(id=attempt_param).select_related("user").first()

    if selected_attempt:
        selected_answers = (
            ExamAnswer.objects.filter(attempt=selected_attempt)
            .select_related("question")
            .prefetch_related("question__options", "selected_options")
            .order_by("question__order", "question__id")
        )
        if exam.exam_type == "test":
            selected_attempt.test_result = calculate_test_attempt_result(
                selected_attempt,
                answers=list(selected_answers),
            )

    now = timezone.now()
    attempts_data = []

    # Apellyasiya bonusları (səhifədəki cəhdlər üçün tək sorğu) — qəbul olunmuş
    # apellyasiyalar müəllimin gördüyü Bal/Faiz sütunlarında da əks olunsun.
    appeal_bonus_by_attempt = _appeal_bonus_map_for(attempts_page) if exam.exam_type == "test" else {}

    for att in attempts_page:
        anonymous_name = _build_anonymous_name(attempt_id=att.id, user_id=att.user_id, exam_id=exam.id)

        can_view_name, identity_window_seconds = _resolve_attempt_name_visibility(att, current_time=now)
        review_window_seconds = _resolve_attempt_review_window_seconds(att, current_time=now)
        action_state = _resolve_attempt_action_state(
            att,
            can_view_name=can_view_name,
            review_window_seconds=review_window_seconds,
            identity_window_seconds=identity_window_seconds,
        )
        real_name = att.user.get_full_name() or att.user.username
        effective_finish, finish_inferred = _attempt_effective_finish(att, now=now)
        effective_duration = _attempt_effective_duration(att, effective_finish)

        test_result = (
            calculate_test_attempt_result(att, answers=att.answers.all()) if exam.exam_type == "test" else None
        )
        appeal_bonus = appeal_bonus_by_attempt.get(att.id) or 0
        if test_result is not None:
            delivered_count = test_result.delivered_count
            if appeal_bonus:
                test_result = _apply_appeal_bonus(test_result, appeal_bonus)
        else:
            delivered_count = att.correct_count + att.wrong_count
        attempts_data.append(
            {
                "attempt": att,
                "test_result": test_result,
                "appeal_bonus": appeal_bonus,
                "delivered_count": delivered_count,
                "anonymous_name": anonymous_name,
                "real_name": real_name,
                "can_view_name": can_view_name,
                "seconds_remaining": review_window_seconds or identity_window_seconds or 0,
                "action_label": action_state["label"],
                "action_url": _append_query_params(
                    reverse(action_state["url_name"], kwargs={"slug": exam.slug, "attempt_id": att.id}),
                    **navigation_params,
                ),
                "countdown_seconds": action_state["countdown_seconds"],
                "countdown_mode": action_state["countdown_mode"],
                "effective_finish": effective_finish,
                "finish_inferred": finish_inferred,
                "effective_duration_seconds": effective_duration,
            }
        )

    # Perf: əvvəl burada `list(attempts)` bütün filtrlənmiş cəhdləri (bütün
    # cavab+option prefetch-ləri ilə) yaddaşa yükləyirdi və fastest_attempts/
    # hardest_questions hesablanırdı — template-də heç istifadə olunmurdu.
    # Silindi: səhifə yalnız 12 cəhdlik paginated dataset yükləyir.
    group_filter_value = str(group_id) if group_id else ""
    pagination_query = urlencode(
        {
            key: value
            for key, value in {
                "q": search_query,
                "status": status_filter,
                "checked": checked_filter,
                "date_from": date_from_raw,
                "date_to": date_to_raw,
                "group": group_filter_value,
                "sort_by": sort_by,
                "sort_dir": sort_dir,
                "attempt": (request.GET.get("attempt") or "").strip(),
                "from_section": (request.GET.get("from_section") or "").strip(),
                "return_to": (request.GET.get("return_to") or "").strip(),
            }.items()
            if value not in ("", None)
        }
    )

    sort_base_query = urlencode(
        {
            key: value
            for key, value in {
                "q": search_query,
                "status": status_filter,
                "checked": checked_filter,
                "date_from": date_from_raw,
                "date_to": date_to_raw,
                "group": group_filter_value,
                "attempt": (request.GET.get("attempt") or "").strip(),
                "from_section": (request.GET.get("from_section") or "").strip(),
                "return_to": (request.GET.get("return_to") or "").strip(),
            }.items()
            if value not in ("", None)
        }
    )

    # Export URL — bütün filtrlər saxlanılır, pagination çıxarılır
    export_query = urlencode(
        {
            key: value
            for key, value in {
                "q": search_query,
                "status": status_filter,
                "checked": checked_filter,
                "date_from": date_from_raw,
                "date_to": date_to_raw,
                "group": group_filter_value,
                "sort_by": sort_by,
                "sort_dir": sort_dir,
            }.items()
            if value not in ("", None)
        }
    )
    export_xlsx_url = reverse("exams:export_exam_results_xlsx", kwargs={"slug": exam.slug})
    if export_query:
        export_xlsx_url = f"{export_xlsx_url}?{export_query}"

    available_groups = _available_groups_for_exam(exam)
    teacher_display = exam.author.get_full_name() or exam.author.username
    teacher_username = exam.author.username

    return render(
        request,
        "exams/teacher/teacher_exam_results.html",
        {
            "exam": exam,
            "attempts": page_obj.object_list,
            "attempts_data": attempts_data,
            "page_obj": page_obj,
            "selected_attempt": selected_attempt,
            "selected_answers": selected_answers,
            "profile_return_url": profile_return_url,
            "exam_detail_url": exam_detail_url,
            "exam_navigation_query": exam_navigation_query,
            "source_back_label": pgettext("exams.template.teacher_exam_detail", "action_back"),
            "review_stats": review_stats,
            "search_query": search_query,
            "status_filter": status_filter,
            "checked_filter": checked_filter,
            "date_from": date_from_raw,
            "date_to": date_to_raw,
            "sort_by": sort_by,
            "sort_dir": sort_dir,
            "sort_base_query": sort_base_query,
            "pagination_query": pagination_query,
            "can_delete_attempts": request_has_permission(request, "exam.delete"),
            "available_groups": available_groups,
            "group_filter": group_filter_value,
            "export_xlsx_url": export_xlsx_url,
            "teacher_display": teacher_display,
            "teacher_username": teacher_username,
        },
    )


@login_required
def export_exam_results_xlsx(request, slug):
    """İmtahan nəticələrini xlsx olaraq export et.

    - Eyni filtrlər tətbiq olunur (group, status, checked, date, q).
    - Pagination YOX — bütün uyğun cəhdlər.
    - Tenant isolation: get_teacher_exam_or_404 bunu təmin edir.
    """
    _ensure_teacher(request.user)
    exam = get_teacher_exam_or_404(request, slug=slug)

    # Eyni filtrləri tətbiq et
    attempts_qs, _ = _apply_results_filters(exam, request)
    attempts_list = list(attempts_qs)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            "openpyxl is not installed. Add openpyxl to requirements.",
            status=500,
            content_type="text/plain",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = (exam.title[:28] + "…") if len(exam.title) > 30 else exam.title

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    is_test = exam.exam_type == "test"
    is_written_or_coding = exam.exam_type in ("written", "coding")

    headers = [
        "#",
        "Qruplar",
        "Ad Soyad",
        "İstifadəçi adı",
        "E-poçt",
        "Status",
        "Başlama",
        "Bitmə",
        "Müddət (s:dq:sn)",
    ]
    if is_written_or_coding:
        headers += ["Müəllim balı", "Yoxlanıb"]
    if is_test:
        headers += ["Düzgün", "Səhv", "Cavabsız", "Verilmiş sual", "Bal", "Maks. bal", "Faiz", "Apel. bonus"]
    else:
        headers += ["Düzgün", "Səhv", "Verilmiş sual"]

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws.row_dimensions[1].height = 28

    now = timezone.now()
    # Loop-invariant precompute (əvvəl export loop-u içində N+1 idi):
    #  • iştirakçı qruplarının id-ləri — bir dəfə (əvvəl hər attempt-də təkrar).
    #  • hər user üçün üzv olduğu qrup adları — TƏK sorğu ilə dict (əvvəl hər
    #    attempt üçün ayrıca sorğu = N+1).
    available_group_ids = list(_available_groups_for_exam(exam).values_list("id", flat=True))
    _attempt_user_ids = {att.user_id for att in attempts_list}
    # Apellyasiya bonusları (tək sorğu) — export-dakı Bal/Faiz effektiv olsun.
    appeal_bonus_by_attempt = _appeal_bonus_map_for(attempts_list) if is_test else {}
    groups_by_user: dict[int, list[str]] = {}
    if available_group_ids and _attempt_user_ids:
        from apps.exams.models import StudentGroup

        for _uid, _gname in (
            StudentGroup.objects.filter(id__in=available_group_ids, students__id__in=_attempt_user_ids)
            .values_list("students__id", "name")
            .order_by("name")
        ):
            groups_by_user.setdefault(_uid, []).append(_gname)
    for row_idx, att in enumerate(attempts_list, start=2):
        effective_finish, _ = _attempt_effective_finish(att, now=now)
        effective_duration = _attempt_effective_duration(att, effective_finish)

        # İştirakçı qruplardan user-in üzv olduqları — loop-dan əvvəl tək sorğu
        # ilə qurulmuş dict-dən (per-attempt sorğu yox).
        user_groups = ", ".join(groups_by_user.get(att.user_id, []))

        # Müddəti hh:mm:ss formatına çevir
        if effective_duration is not None:
            d_total = max(int(effective_duration), 0)
            d_h, d_m, d_s = d_total // 3600, (d_total % 3600) // 60, d_total % 60
            duration_str = f"{d_h:02d}:{d_m:02d}:{d_s:02d}"
        else:
            duration_str = ""

        row = [
            row_idx - 1,
            user_groups,
            att.user.get_full_name() or att.user.username,
            att.user.username,
            att.user.email or "",
            att.get_status_display(),
            att.started_at.replace(tzinfo=None) if att.started_at else "",
            effective_finish.replace(tzinfo=None) if effective_finish else "",
            duration_str,
        ]

        if is_written_or_coding:
            row += [
                att.teacher_score if att.teacher_score is not None else "",
                "Bəli" if att.checked_by_teacher else "Xeyr",
            ]

        if is_test:
            test_result = calculate_test_attempt_result(att, answers=att.answers.all())
            appeal_bonus = appeal_bonus_by_attempt.get(att.id) or 0
            if appeal_bonus:
                test_result = _apply_appeal_bonus(test_result, appeal_bonus)
            row += [
                test_result.correct_count,
                test_result.wrong_count,
                test_result.unanswered_count,
                test_result.delivered_count,
                float(test_result.score_display) if test_result.score_display else 0,
                float(test_result.max_score_display) if test_result.max_score_display else 0,
                float(test_result.percentage_display) if test_result.percentage_display else 0,
                float(appeal_bonus) if appeal_bonus else 0,
            ]
        else:
            delivered = att.correct_count + att.wrong_count
            row += [att.correct_count, att.wrong_count, delivered]

        # Sola düzlənən sütunlar: Qruplar (2), Ad Soyad (3), İstifadəçi adı (4), E-poçt (5), Başlama (7), Bitmə (8)
        left_columns = {2, 3, 4, 5, 7, 8}
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = left if col_idx in left_columns else center
            if isinstance(value, datetime):
                cell.number_format = "DD.MM.YYYY HH:MM"

    # Sütun genişlikləri (header sırasına uyğun)
    # #, Qruplar, Ad Soyad, İstifadəçi adı, E-poçt, Status, Başlama, Bitmə, Müddət
    widths = [5, 26, 26, 20, 28, 16, 20, 20, 16]
    if is_written_or_coding:
        widths += [14, 12]  # Müəllim balı, Yoxlanıb
    if is_test:
        widths += [10, 10, 12, 16, 8, 12, 8, 12]  # Düzgün, Səhv, Cavabsız, Verilmiş, Bal, Maks, Faiz, Apel.
    else:
        widths += [10, 10, 14]  # Düzgün, Səhv, Verilmiş sual
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{slugify(exam.title) or 'exam'}-results-{timezone.now().strftime('%Y%m%d-%H%M')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
