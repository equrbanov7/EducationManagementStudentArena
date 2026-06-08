import json
from datetime import datetime, timedelta
from io import BytesIO
from urllib.parse import urlencode, urlsplit

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.crypto import salted_hmac
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils.text import slugify
from django.utils.translation import pgettext, pgettext_lazy
from django.views.decorators.http import require_http_methods

from apps.exams.models import ExamAnswer, ExamAttempt
from apps.exams.services.access_policy import _ensure_teacher
from apps.exams.services.ai_grading import has_ai_gradeable_answer_content, has_written_answer_content
from apps.exams.services.randomizer import generate_random_questions_for_attempt
from apps.exams.services.result_calculation import calculate_test_attempt_result
from apps.exams.services.review_visibility import attempt_review_window_locked as _attempt_review_window_locked
from apps.exams.services.review_visibility import (
    resolve_exam_attempt_name_visibility as _resolve_attempt_name_visibility,
)
from apps.exams.services.review_visibility import (
    resolve_exam_attempt_review_window_seconds as _resolve_attempt_review_window_seconds,
)
from apps.exams.views.shared.tenant import get_teacher_exam_or_404, tenant_scoped_exams
from core.permissions import request_has_permission

ANONYMOUS_NAME_TOKEN_SALT = "exams.teacher_results.anonymous_name"  # nosec B105
ANONYMOUS_NAME_CODE_LENGTH = 6


def _user_display_name(user):
    return user.get_full_name() or user.username


def _coding_submission_file_items(submission):
    if not submission:
        return []
    code_files = list(submission.code_files.all())
    if code_files:
        return [
            {
                "name": code_file.name,
                "content": code_file.content or "",
                "language": code_file.language or "",
                "is_main": bool(code_file.is_main),
            }
            for code_file in code_files
        ]
    return [
        {
            "name": item.get("name") or "file.txt",
            "content": item.get("content") or "",
            "language": item.get("language") or "",
            "is_main": bool(item.get("is_main")),
        }
        for item in (submission.files or [])
        if isinstance(item, dict)
    ]


def _sync_coding_answers_from_final_submissions(attempt):
    if getattr(attempt.exam, "exam_type", "") != "coding":
        return

    seen_question_ids = set()
    submissions = (
        attempt.coding_submissions.filter(is_final=True)
        .select_related("question", "question__question")
        .order_by("question__question__order", "question__question_id", "-submitted_at", "-id")
    )
    for submission in submissions:
        base_question = getattr(submission.question, "question", None)
        if not base_question or base_question.id in seen_question_ids:
            continue
        seen_question_ids.add(base_question.id)

        answer, created = ExamAnswer.objects.get_or_create(
            attempt=attempt,
            question=base_question,
            defaults={"text_answer": submission.submitted_code or ""},
        )
        if not created and not (answer.text_answer or "").strip() and submission.submitted_code:
            answer.text_answer = submission.submitted_code
            answer.save(update_fields=["text_answer", "updated_at"])


def _build_answer_review_item(answer):
    answer_files = list(answer.files.all())
    has_text_answer = bool((getattr(answer, "text_answer", "") or "").strip())
    has_paint_answer = bool(getattr(answer, "paint_image", None)) or bool(getattr(answer, "has_paint", False))
    has_file_answer = bool(answer_files)
    selected_options_count = len(list(answer.selected_options.all()))
    has_answer_content = has_written_answer_content(
        student_answer=answer.text_answer,
        answer_files=answer_files,
        paint_image=getattr(answer, "paint_image", None),
    )

    coding_submission = None
    if getattr(answer.question.exam, "exam_type", "") == "coding":
        coding_submission = (
            answer.attempt.coding_submissions.filter(question__question=answer.question, is_final=True)
            .prefetch_related("code_files")
            .order_by("-submitted_at")
            .first()
        )
    coding_files = _coding_submission_file_items(coding_submission)
    has_answer_content = has_answer_content or bool(coding_files)

    return {
        "question": answer.question,
        "answer": answer,
        "coding_submission": coding_submission,
        "coding_files": coding_files,
        "coding_file_count": len(coding_files),
        "answer_files": answer_files,
        "has_text_answer": has_text_answer,
        "has_paint_answer": has_paint_answer,
        "has_file_answer": has_file_answer,
        "selected_options_count": selected_options_count,
        "has_answer_content": has_answer_content,
        "has_ai_gradeable_content": has_ai_gradeable_answer_content(
            student_answer=answer.text_answer,
            answer_files=answer_files,
            paint_image=getattr(answer, "paint_image", None),
        ),
    }


def _append_query_params(url, **params):
    clean_params = {key: value for key, value in params.items() if value not in (None, "")}
    if not clean_params:
        return url
    separator = "&" if "?" in url else "?"
    return f"{url}{separator}{urlencode(clean_params)}"


def _safe_same_origin_redirect_path(request, candidate_url):
    raw_url = (candidate_url or "").strip()
    if not raw_url:
        return ""

    if not url_has_allowed_host_and_scheme(
        raw_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return ""

    parsed = urlsplit(raw_url)
    if parsed.netloc and parsed.netloc != request.get_host():
        return ""

    path = parsed.path or "/"
    query = f"?{parsed.query}" if parsed.query else ""
    fragment = f"#{parsed.fragment}" if parsed.fragment else ""
    return f"{path}{query}{fragment}"


def _resolve_profile_navigation(request, *, default_section="my-exams"):
    requested_profile_section = (request.GET.get("from_section") or "").strip()
    valid_profile_sections = {
        "my-exams",
        "assigned-exams",
        "profile-info",
        "my-courses",
        "assigned-courses",
        "courses",
        "pending-review",
        "review-results",
    }
    if requested_profile_section not in valid_profile_sections:
        requested_profile_section = default_section

    fallback_profile_return_url = f"{reverse('accounts:profile')}?section={requested_profile_section}"
    explicit_return_url = _safe_same_origin_redirect_path(
        request,
        request.GET.get("return_to") or request.GET.get("next") or request.META.get("HTTP_REFERER"),
    )
    profile_return_url = explicit_return_url or fallback_profile_return_url

    navigation_params = {
        "from_section": requested_profile_section,
        "return_to": profile_return_url,
    }
    return profile_return_url, navigation_params


def _build_attempt_timing_context(attempt):
    started_at = getattr(attempt, "started_at", None)
    finished_at = getattr(attempt, "finished_at", None)
    duration_seconds = getattr(attempt, "duration_seconds", None)

    if duration_seconds is None and started_at and finished_at:
        duration_seconds = max(int((finished_at - started_at).total_seconds()), 0)

    if finished_at is None and started_at and duration_seconds is not None:
        finished_at = started_at + timedelta(seconds=duration_seconds)

    return {
        "started_at": started_at,
        "finished_at": finished_at,
        "duration_seconds": duration_seconds,
        "has_duration": duration_seconds is not None,
    }


def _parse_filter_date(raw_value):
    raw_date = (raw_value or "").strip()
    if not raw_date:
        return "", None
    try:
        return raw_date, datetime.strptime(raw_date, "%Y-%m-%d").date()
    except ValueError:
        return "", None


def _resolve_attempt_action_state(attempt, *, can_view_name, review_window_seconds, identity_window_seconds):
    if attempt.exam.exam_type == "test":
        return {
            "label": "Bax",
            "url_name": "exams:teacher_view_attempt",
            "countdown_seconds": 0,
            "countdown_mode": "",
        }

    if attempt.checked_by_teacher:
        if review_window_seconds:
            return {
                "label": "Yenidən yoxla",
                "url_name": "exams:teacher_check_attempt",
                "countdown_seconds": review_window_seconds,
                "countdown_mode": "recheck",
            }
        return {
            "label": "Bax",
            "url_name": "exams:teacher_view_attempt",
            "countdown_seconds": 0,
            "countdown_mode": "",
        }

    return {
        "label": "Yoxla",
        "url_name": "exams:teacher_check_attempt",
        "countdown_seconds": identity_window_seconds if not can_view_name else 0,
        "countdown_mode": "identity" if not can_view_name and identity_window_seconds else "",
    }


def _build_anonymous_name(*, attempt_id: int, user_id: int, exam_id: int) -> str:
    token = salted_hmac(
        ANONYMOUS_NAME_TOKEN_SALT,
        f"{attempt_id}:{user_id}:{exam_id}",
    ).hexdigest()
    return pgettext("exams.view.results.label", "anonymous_student").format(
        code=token[:ANONYMOUS_NAME_CODE_LENGTH].upper()
    )


def _available_groups_for_exam(exam):
    """İmtahanın iştirakçılarının üzv olduğu qruplar + allowed_groups (tenant-scoped).

    - allowed_groups-i daxil et (təyin olunmuş qruplar).
    - Plus: imtahanın cəhdləri olan tələbələrin üzv olduqları qruplar.
    - Tenant: yalnız bu imtahanın organization-na aid qrupları göstər (əgər varsa).
    """
    from apps.exams.models import StudentGroup

    attempt_user_ids = exam.attempts.values_list("user_id", flat=True)
    qs = StudentGroup.objects.filter(Q(exams=exam) | Q(students__id__in=attempt_user_ids))
    org_id = getattr(exam, "organization_id", None)
    if org_id:
        qs = qs.filter(organization_id=org_id)
    return qs.distinct().order_by("name")


def _attempt_effective_finish(attempt, *, now=None):
    """İmtahanı bitirməyən tələbə üçün effektiv bitmə vaxtı.

    - Əgər finished_at varsa, onu qaytarır.
    - Yoxsa: started_at + exam.total_duration_minutes (əgər müddət bitibsə).
    - Müddət hələ bitməyibsə None.
    """
    finished_at = getattr(attempt, "finished_at", None)
    if finished_at:
        return finished_at, False
    started_at = getattr(attempt, "started_at", None)
    duration_minutes = getattr(attempt.exam, "total_duration_minutes", None)
    if not started_at or not duration_minutes:
        return None, False
    deadline = started_at + timedelta(minutes=int(duration_minutes))
    if (now or timezone.now()) >= deadline:
        return deadline, True
    return None, False


def _apply_results_filters(exam, request):
    """Şərt: teacher_exam_results və export_exam_results_xlsx ortaq filter logic-i.

    Qaytarır: (attempts_qs, filter_state_dict)
    """
    attempts = exam.attempts.select_related("user", "exam").prefetch_related(
        "answers__question__options",
        "answers__selected_options",
    )

    search_query = (request.GET.get("q") or "").strip()
    if search_query:
        attempts = attempts.filter(
            Q(user__username__icontains=search_query)
            | Q(user__first_name__icontains=search_query)
            | Q(user__last_name__icontains=search_query)
        )

    status_filter = (request.GET.get("status") or "all").strip().lower()
    allowed_status_filters = {"all", "draft", "in_progress", "submitted", "expired"}
    if status_filter not in allowed_status_filters:
        status_filter = "all"
    if status_filter != "all":
        attempts = attempts.filter(status=status_filter)

    checked_filter = (request.GET.get("checked") or "all").strip().lower()
    if checked_filter == "checked":
        attempts = attempts.filter(checked_by_teacher=True)
    elif checked_filter == "unchecked":
        attempts = attempts.filter(checked_by_teacher=False)
    else:
        checked_filter = "all"

    date_from_raw, date_from = _parse_filter_date(request.GET.get("date_from"))
    date_to_raw, date_to = _parse_filter_date(request.GET.get("date_to"))
    if date_from:
        attempts = attempts.filter(started_at__date__gte=date_from)
    if date_to:
        attempts = attempts.filter(started_at__date__lte=date_to)

    # Group filter — imtahanın iştirakçılarının üzvü olduğu istənilən qrupdan
    group_filter_raw = (request.GET.get("group") or "").strip().lower()
    group_id = None
    if group_filter_raw and group_filter_raw != "all":
        try:
            group_id = int(group_filter_raw)
        except ValueError:
            group_id = None
    if group_id is not None and _available_groups_for_exam(exam).filter(id=group_id).exists():
        attempts = attempts.filter(user__student_groups_as_student__id=group_id).distinct()
    else:
        group_id = None

    sort_by = (request.GET.get("sort_by") or "").strip()
    sort_dir = (request.GET.get("sort_dir") or "").strip()
    ALLOWED_SORT_FIELDS = {
        "user": "user__first_name",
        "username": "user__username",
        "email": "user__email",
        "status": "status",
        "correct": "correct_count",
        "wrong": "wrong_count",
        "score": "teacher_score",
        "start": "started_at",
        "end": "finished_at",
        "duration": "duration_seconds",
    }
    if sort_by in ALLOWED_SORT_FIELDS and sort_dir in ("asc", "desc"):
        order_field = ALLOWED_SORT_FIELDS[sort_by]
        if sort_dir == "desc":
            order_field = f"-{order_field}"
        attempts = attempts.order_by(order_field)
    else:
        sort_by = ""
        sort_dir = ""
        attempts = attempts.order_by("-started_at")

    return attempts, {
        "search_query": search_query,
        "status_filter": status_filter,
        "checked_filter": checked_filter,
        "date_from_raw": date_from_raw,
        "date_to_raw": date_to_raw,
        "group_id": group_id,
        "sort_by": sort_by,
        "sort_dir": sort_dir,
    }


@login_required
def teacher_exam_results(request, slug):
    """
    Müəllim üçün imtahan nəticələri:
    - solda bütün cəhdlər cədvəli
    - aşağıda/sağda seçilmiş cəhdin cavabları + qiymətləndirmə formu
    """
    _ensure_teacher(request.user)
    exam = get_teacher_exam_or_404(request, slug=slug)
    profile_return_url, navigation_params = _resolve_profile_navigation(request, default_section="my-exams")
    exam_navigation_query = urlencode(navigation_params)
    exam_detail_url = _append_query_params(
        reverse("exams:teacher_exam_detail", kwargs={"slug": exam.slug}),
        **navigation_params,
    )

    selected_attempt = None
    selected_answers = None

    # ---------- POST: müəllim bal + feedback saxlayır ----------
    if request.method == "POST":
        if not request_has_permission(request, "grade.input"):
            messages.error(
                request,
                pgettext_lazy("exams.view.results.message", "grading_permission_required"),
            )
            return redirect(request.path)

        attempt_id = request.POST.get("attempt_id")
        score_raw = request.POST.get("teacher_score", "").strip()
        feedback = request.POST.get("teacher_feedback", "").strip()
        from apps.notifications.services import notify_student_about_feedback

        selected_attempt = get_object_or_404(ExamAttempt, id=attempt_id, exam=exam)

        if score_raw:
            try:
                score_val = int(score_raw)
            except ValueError:
                messages.error(request, pgettext_lazy("exams.view.results.message", "score_must_be_integer"))
            else:
                if 0 <= score_val <= 100:
                    selected_attempt.teacher_score = score_val
                    selected_attempt.teacher_feedback = feedback
                    selected_attempt.mark_checked()
                    notify_student_about_feedback(
                        task=exam,
                        student=selected_attempt.user,
                        task_kind="exam",
                        extra_metadata={"attempt_id": selected_attempt.id},
                    )
                    messages.success(request, pgettext_lazy("exams.view.results.message", "score_feedback_saved"))
                    return redirect(
                        _append_query_params(
                            request.path,
                            attempt=selected_attempt.id,
                            **navigation_params,
                        )
                    )
                else:
                    messages.error(request, pgettext_lazy("exams.view.results.message", "score_range_0_100"))
        else:
            # yalnız feedback saxlanılır
            selected_attempt.teacher_score = None
            selected_attempt.teacher_feedback = feedback
            selected_attempt.checked_by_teacher = False
            selected_attempt.save(
                update_fields=[
                    "teacher_score",
                    "teacher_feedback",
                    "checked_by_teacher",
                ]
            )
            notify_student_about_feedback(
                task=exam,
                student=selected_attempt.user,
                task_kind="exam",
                extra_metadata={"attempt_id": selected_attempt.id},
            )
            messages.success(request, pgettext_lazy("exams.view.results.message", "feedback_saved"))
            return redirect(
                _append_query_params(
                    request.path,
                    attempt=selected_attempt.id,
                    **navigation_params,
                )
            )

    attempts, filter_state = _apply_results_filters(exam, request)
    search_query = filter_state["search_query"]
    status_filter = filter_state["status_filter"]
    checked_filter = filter_state["checked_filter"]
    date_from_raw = filter_state["date_from_raw"]
    date_to_raw = filter_state["date_to_raw"]
    group_id = filter_state["group_id"]
    sort_by = filter_state["sort_by"]
    sort_dir = filter_state["sort_dir"]

    max_score = 100 if exam.exam_type == "test" else exam.questions.aggregate(total=Sum("points")).get("total") or 0
    pending_count = 0 if exam.exam_type == "test" else attempts.filter(checked_by_teacher=False).count()
    graded_count = attempts.count() if exam.exam_type == "test" else attempts.filter(checked_by_teacher=True).count()
    review_stats = {
        "total": attempts.count(),
        "pending": pending_count,
        "graded": graded_count,
        "max_score": max_score,
    }
    paginator = Paginator(attempts, 12)
    page_obj = paginator.get_page(request.GET.get("page"))
    attempts_page = list(page_obj.object_list)

    # ---------- GET: hansı attempt seçilib? ----------
    if selected_attempt is None:
        attempt_param = request.GET.get("attempt")
        if attempt_param:
            selected_attempt = exam.attempts.filter(id=attempt_param).select_related("user").first()

    if selected_attempt:
        selected_answers = (
            ExamAnswer.objects.filter(attempt=selected_attempt)
            .select_related("question")
            .prefetch_related("question__options", "selected_options")
            .order_by("question__order", "question__id")
        )
        if exam.exam_type == "test":
            selected_attempt.test_result = calculate_test_attempt_result(
                selected_attempt,
                answers=list(selected_answers),
            )

    now = timezone.now()
    attempts_data = []

    for att in attempts_page:
        anonymous_name = _build_anonymous_name(attempt_id=att.id, user_id=att.user_id, exam_id=exam.id)

        can_view_name, identity_window_seconds = _resolve_attempt_name_visibility(att, current_time=now)
        review_window_seconds = _resolve_attempt_review_window_seconds(att, current_time=now)
        action_state = _resolve_attempt_action_state(
            att,
            can_view_name=can_view_name,
            review_window_seconds=review_window_seconds,
            identity_window_seconds=identity_window_seconds,
        )
        real_name = att.user.get_full_name() or att.user.username
        effective_finish, finish_inferred = _attempt_effective_finish(att, now=now)
        effective_duration = att.duration_seconds
        if effective_duration is None and effective_finish and att.started_at:
            effective_duration = max(int((effective_finish - att.started_at).total_seconds()), 0)

        test_result = calculate_test_attempt_result(att) if exam.exam_type == "test" else None
        if test_result is not None:
            delivered_count = test_result.delivered_count
        else:
            delivered_count = att.correct_count + att.wrong_count
        attempts_data.append(
            {
                "attempt": att,
                "test_result": test_result,
                "delivered_count": delivered_count,
                "anonymous_name": anonymous_name,
                "real_name": real_name,
                "can_view_name": can_view_name,
                "seconds_remaining": review_window_seconds or identity_window_seconds or 0,
                "action_label": action_state["label"],
                "action_url": _append_query_params(
                    reverse(action_state["url_name"], kwargs={"slug": exam.slug, "attempt_id": att.id}),
                    **navigation_params,
                ),
                "countdown_seconds": action_state["countdown_seconds"],
                "countdown_mode": action_state["countdown_mode"],
                "effective_finish": effective_finish,
                "finish_inferred": finish_inferred,
                "effective_duration_seconds": effective_duration,
            }
        )

    # ═══════════════════════════════════════════════════════════════════
    # Statistikalar (əvvəlki kimi)
    # ═══════════════════════════════════════════════════════════════════
    filtered_attempts = list(attempts)
    fastest_attempts = sorted([a for a in filtered_attempts if a.duration_seconds], key=lambda a: a.duration_seconds)[
        :5
    ]

    questions = exam.questions.all()
    hardest_questions = sorted(questions, key=lambda q: q.correct_ratio)[:5]

    group_filter_value = str(group_id) if group_id else ""
    pagination_query = urlencode(
        {
            key: value
            for key, value in {
                "q": search_query,
                "status": status_filter,
                "checked": checked_filter,
                "date_from": date_from_raw,
                "date_to": date_to_raw,
                "group": group_filter_value,
                "sort_by": sort_by,
                "sort_dir": sort_dir,
                "attempt": (request.GET.get("attempt") or "").strip(),
                "from_section": (request.GET.get("from_section") or "").strip(),
                "return_to": (request.GET.get("return_to") or "").strip(),
            }.items()
            if value not in ("", None)
        }
    )

    sort_base_query = urlencode(
        {
            key: value
            for key, value in {
                "q": search_query,
                "status": status_filter,
                "checked": checked_filter,
                "date_from": date_from_raw,
                "date_to": date_to_raw,
                "group": group_filter_value,
                "attempt": (request.GET.get("attempt") or "").strip(),
                "from_section": (request.GET.get("from_section") or "").strip(),
                "return_to": (request.GET.get("return_to") or "").strip(),
            }.items()
            if value not in ("", None)
        }
    )

    # Export URL — bütün filtrlər saxlanılır, pagination çıxarılır
    export_query = urlencode(
        {
            key: value
            for key, value in {
                "q": search_query,
                "status": status_filter,
                "checked": checked_filter,
                "date_from": date_from_raw,
                "date_to": date_to_raw,
                "group": group_filter_value,
                "sort_by": sort_by,
                "sort_dir": sort_dir,
            }.items()
            if value not in ("", None)
        }
    )
    export_xlsx_url = reverse("exams:export_exam_results_xlsx", kwargs={"slug": exam.slug})
    if export_query:
        export_xlsx_url = f"{export_xlsx_url}?{export_query}"

    available_groups = _available_groups_for_exam(exam)
    teacher_display = exam.author.get_full_name() or exam.author.username
    teacher_username = exam.author.username

    return render(
        request,
        "exams/teacher/teacher_exam_results.html",
        {
            "exam": exam,
            "attempts": page_obj.object_list,
            "attempts_data": attempts_data,
            "page_obj": page_obj,
            "fastest_attempts": fastest_attempts,
            "hardest_questions": hardest_questions,
            "selected_attempt": selected_attempt,
            "selected_answers": selected_answers,
            "profile_return_url": profile_return_url,
            "exam_detail_url": exam_detail_url,
            "exam_navigation_query": exam_navigation_query,
            "source_back_label": pgettext("exams.template.teacher_exam_detail", "action_back"),
            "review_stats": review_stats,
            "search_query": search_query,
            "status_filter": status_filter,
            "checked_filter": checked_filter,
            "date_from": date_from_raw,
            "date_to": date_to_raw,
            "sort_by": sort_by,
            "sort_dir": sort_dir,
            "sort_base_query": sort_base_query,
            "pagination_query": pagination_query,
            "can_delete_attempts": request_has_permission(request, "exam.delete"),
            "available_groups": available_groups,
            "group_filter": group_filter_value,
            "export_xlsx_url": export_xlsx_url,
            "teacher_display": teacher_display,
            "teacher_username": teacher_username,
        },
    )


@login_required
def export_exam_results_xlsx(request, slug):
    """İmtahan nəticələrini xlsx olaraq export et.

    - Eyni filtrlər tətbiq olunur (group, status, checked, date, q).
    - Pagination YOX — bütün uyğun cəhdlər.
    - Tenant isolation: get_teacher_exam_or_404 bunu təmin edir.
    """
    _ensure_teacher(request.user)
    exam = get_teacher_exam_or_404(request, slug=slug)

    # Eyni filtrləri tətbiq et
    attempts_qs, _ = _apply_results_filters(exam, request)
    attempts_list = list(attempts_qs)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            "openpyxl is not installed. Add openpyxl to requirements.",
            status=500,
            content_type="text/plain",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = (exam.title[:28] + "…") if len(exam.title) > 30 else exam.title

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    is_test = exam.exam_type == "test"
    is_written_or_coding = exam.exam_type in ("written", "coding")

    headers = [
        "#",
        "Qruplar",
        "Ad Soyad",
        "İstifadəçi adı",
        "E-poçt",
        "Status",
        "Başlama",
        "Bitmə",
        "Müddət (s:dq:sn)",
    ]
    if is_written_or_coding:
        headers += ["Müəllim balı", "Yoxlanıb"]
    if is_test:
        headers += ["Düzgün", "Səhv", "Cavabsız", "Verilmiş sual", "Bal", "Maks. bal", "Faiz"]
    else:
        headers += ["Düzgün", "Səhv", "Verilmiş sual"]

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws.row_dimensions[1].height = 28

    now = timezone.now()
    for row_idx, att in enumerate(attempts_list, start=2):
        effective_finish, _ = _attempt_effective_finish(att, now=now)
        effective_duration = att.duration_seconds
        if effective_duration is None and effective_finish and att.started_at:
            effective_duration = max(int((effective_finish - att.started_at).total_seconds()), 0)

        # Bütün allowed/iştirakçı qruplarından user-in üzv olduqları
        available_group_ids = list(_available_groups_for_exam(exam).values_list("id", flat=True))
        user_groups = ", ".join(
            list(att.user.student_groups_as_student.filter(id__in=available_group_ids).values_list("name", flat=True))
        )

        # Müddəti hh:mm:ss formatına çevir
        if effective_duration is not None:
            d_total = max(int(effective_duration), 0)
            d_h, d_m, d_s = d_total // 3600, (d_total % 3600) // 60, d_total % 60
            duration_str = f"{d_h:02d}:{d_m:02d}:{d_s:02d}"
        else:
            duration_str = ""

        row = [
            row_idx - 1,
            user_groups,
            att.user.get_full_name() or att.user.username,
            att.user.username,
            att.user.email or "",
            att.get_status_display(),
            att.started_at.replace(tzinfo=None) if att.started_at else "",
            effective_finish.replace(tzinfo=None) if effective_finish else "",
            duration_str,
        ]

        if is_written_or_coding:
            row += [
                att.teacher_score if att.teacher_score is not None else "",
                "Bəli" if att.checked_by_teacher else "Xeyr",
            ]

        if is_test:
            test_result = calculate_test_attempt_result(att)
            row += [
                test_result.correct_count,
                test_result.wrong_count,
                test_result.unanswered_count,
                test_result.delivered_count,
                float(test_result.score_display) if test_result.score_display else 0,
                float(test_result.max_score_display) if test_result.max_score_display else 0,
                float(test_result.percentage_display) if test_result.percentage_display else 0,
            ]
        else:
            delivered = att.correct_count + att.wrong_count
            row += [att.correct_count, att.wrong_count, delivered]

        # Sola düzlənən sütunlar: Qruplar (2), Ad Soyad (3), İstifadəçi adı (4), E-poçt (5), Başlama (7), Bitmə (8)
        left_columns = {2, 3, 4, 5, 7, 8}
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = left if col_idx in left_columns else center
            if isinstance(value, datetime):
                cell.number_format = "DD.MM.YYYY HH:MM"

    # Sütun genişlikləri (header sırasına uyğun)
    # #, Qruplar, Ad Soyad, İstifadəçi adı, E-poçt, Status, Başlama, Bitmə, Müddət
    widths = [5, 26, 26, 20, 28, 16, 20, 20, 16]
    if is_written_or_coding:
        widths += [14, 12]  # Müəllim balı, Yoxlanıb
    if is_test:
        widths += [10, 10, 12, 16, 8, 12, 8]  # Düzgün, Səhv, Cavabsız, Verilmiş, Bal, Maks, Faiz
    else:
        widths += [10, 10, 14]  # Düzgün, Səhv, Verilmiş sual
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{slugify(exam.title) or 'exam'}-results-{timezone.now().strftime('%Y%m%d-%H%M')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_http_methods(["POST"])
def delete_exam_attempts(request, slug):
    _ensure_teacher(request.user)
    exam = get_teacher_exam_or_404(request, slug=slug)
    redirect_url = _safe_same_origin_redirect_path(request, request.POST.get("next"))
    if not redirect_url:
        nav_params = {}
        from_section = (request.POST.get("from_section") or "").strip()
        return_to = _safe_same_origin_redirect_path(request, request.POST.get("return_to"))
        if from_section:
            nav_params["from_section"] = from_section
        if return_to:
            nav_params["return_to"] = return_to
        redirect_url = _append_query_params(
            reverse("exams:teacher_exam_results", kwargs={"slug": exam.slug}),
            **nav_params,
        )

    if not request_has_permission(request, "exam.delete"):
        messages.error(request, pgettext_lazy("exams.view.results.message", "delete_permission_required"))
        return redirect(redirect_url)

    raw_ids = request.POST.getlist("attempt_ids")
    single_attempt_id = (request.POST.get("attempt_id") or "").strip()
    if single_attempt_id:
        raw_ids.append(single_attempt_id)

    attempt_ids = sorted({int(raw_id) for raw_id in raw_ids if str(raw_id).isdigit()})
    if not attempt_ids:
        messages.warning(request, pgettext_lazy("exams.view.results.message", "select_attempt_to_delete"))
        return redirect(redirect_url)

    attempts_qs = exam.attempts.filter(id__in=attempt_ids)
    attempt_count = attempts_qs.count()
    if attempt_count == 0:
        messages.warning(request, pgettext_lazy("exams.view.results.message", "attempt_not_found"))
        return redirect(redirect_url)

    attempts_qs.delete()
    messages.success(
        request,
        pgettext_lazy("exams.view.results.message", "attempts_deleted").format(count=attempt_count),
    )
    return redirect(redirect_url)


@login_required
def teacher_view_attempt(request, slug, attempt_id):
    """
    ✅ Müəllim cavabları YALNIZ GÖRMƏK üçün (bal verə bilməz)
    Test və Yazılı hər ikisi üçün işləyir
    """
    _ensure_teacher(request.user)

    exam = get_teacher_exam_or_404(request, slug=slug)
    attempt = get_object_or_404(ExamAttempt, id=attempt_id, exam=exam)
    profile_return_url, navigation_params = _resolve_profile_navigation(request, default_section="my-exams")
    _sync_coding_answers_from_final_submissions(attempt)

    # Cavabları al
    answers_qs = (
        attempt.answers.select_related("question")
        .prefetch_related("files", "selected_options", "question__options")
        .order_by("id")
    )

    if not answers_qs.exists() and not attempt.is_finished:
        generate_random_questions_for_attempt(attempt)
        answers_qs = (
            attempt.answers.select_related("question")
            .prefetch_related("files", "selected_options", "question__options")
            .order_by("id")
        )

    qa_list = [_build_answer_review_item(a) for a in answers_qs]
    test_result = calculate_test_attempt_result(attempt, answers=list(answers_qs)) if exam.exam_type == "test" else None

    # Apellyasiya nəticəsində bal düzəlibsə effektiv balı + düzəlmiş sualları göstər.
    effective_score_info = None
    appeal_bonus_points = 0
    appeal_corrected_qids = {}
    if exam.exam_type == "test":
        from apps.appeals.services import appeal_score_state, effective_test_score

        effective_score_info = effective_test_score(attempt)
        _appeal_state = appeal_score_state(attempt)
        appeal_bonus_points = _appeal_state["bonus_points"]
        appeal_corrected_qids = {qid: True for qid in _appeal_state["credited_question_ids"]}
    can_view_name, identity_window_seconds = _resolve_attempt_name_visibility(attempt, current_time=timezone.now())
    if attempt.exam.exam_type == "test":
        student_display = attempt.user.get_full_name() or attempt.user.username
    else:
        anonymous_name = _build_anonymous_name(
            attempt_id=attempt.id,
            user_id=attempt.user_id,
            exam_id=attempt.exam_id,
        )
        student_display = attempt.user.get_full_name() or attempt.user.username if can_view_name else anonymous_name

    search_query = (request.GET.get("q") or "").strip()
    if search_query:
        search_token = search_query.lower()
        filtered = []
        for item in qa_list:
            question = item["question"]
            answer = item["answer"]
            question_text = (question.text or "").lower()
            answer_text = (getattr(answer, "text_answer", "") or "").lower()
            options_text = " ".join(opt.text for opt in question.options.all()).lower()
            if search_token in question_text or search_token in answer_text or search_token in options_text:
                filtered.append(item)
        qa_list = filtered

    questions_page = Paginator(qa_list, 6).get_page(request.GET.get("questions_page"))
    pagination_query = urlencode(
        {
            **navigation_params,
            "q": search_query,
        }
    )
    clear_search_url = _append_query_params(request.path, **navigation_params)

    context = {
        "exam": exam,
        "attempt": attempt,
        "attempt_timing": _build_attempt_timing_context(attempt),
        "qa_list": questions_page.object_list,
        "qa_page": questions_page,
        "qa_search_query": search_query,
        "test_result": test_result,
        "effective_score_info": effective_score_info,
        "appeal_bonus_points": appeal_bonus_points,
        "appeal_corrected_qids": appeal_corrected_qids,
        "qa_pagination_query": pagination_query,
        "qa_clear_search_url": clear_search_url,
        "read_only": True,  # ✅ Yalnız oxumaq rejimi
        "profile_return_url": profile_return_url,
        "source_back_label": pgettext("exams.template.teacher_exam_detail", "action_back"),
        "student_display": student_display,
        "exam_evaluator_display": _user_display_name(exam.author),
        "can_view_student_identity": can_view_name,
        "identity_window_seconds_left": identity_window_seconds,
    }

    return render(request, "exams/teacher/teacher_view_attempt.html", context)


@login_required
def teacher_check_attempt(request, slug, attempt_id):
    """
    Müəllim yazılı/praktiki imtahandakı BİR cəhdi sual-sual yoxlayır.

    ✅ MÜDAFİƏ: 5 dəqiqə keçibsə, yalnız oxumaq üçün yönləndir
    """
    _ensure_teacher(request.user)

    exam = get_teacher_exam_or_404(request, slug=slug)
    attempt = get_object_or_404(ExamAttempt, id=attempt_id, exam=exam)
    profile_return_url, navigation_params = _resolve_profile_navigation(request, default_section="my-exams")
    _sync_coding_answers_from_final_submissions(attempt)
    if navigation_params.get("from_section") == "pending-review":
        results_return_url = profile_return_url
    else:
        results_return_url = _append_query_params(
            reverse("exams:teacher_exam_results", kwargs={"slug": exam.slug}),
            **navigation_params,
        )
    view_attempt_url = _append_query_params(
        reverse("exams:teacher_view_attempt", kwargs={"slug": exam.slug, "attempt_id": attempt.id}),
        **navigation_params,
    )

    # ✅ 5 dəqiqə keçibsə, yalnız "bax" səhifəsinə yönləndir
    if _attempt_review_window_locked(attempt, current_time=timezone.now()):
        messages.warning(
            request,
            pgettext_lazy("exams.view.results.message", "cannot_edit_after_five_minutes"),
        )
        return redirect(view_attempt_url)

    # YALNIZ bu attempt-ə düşən suallar
    answers_qs = (
        attempt.answers.select_related("question")
        .prefetch_related("files", "selected_options", "question__options")
        .order_by("id")
    )

    if not answers_qs.exists() and not attempt.is_finished:
        generate_random_questions_for_attempt(attempt)
        answers_qs = (
            attempt.answers.select_related("question")
            .prefetch_related("files", "selected_options", "question__options")
            .order_by("id")
        )

    qa_list = [_build_answer_review_item(a) for a in answers_qs]
    can_view_name, identity_window_seconds = _resolve_attempt_name_visibility(attempt, current_time=timezone.now())
    if attempt.exam.exam_type == "test":
        student_display = attempt.user.get_full_name() or attempt.user.username
    else:
        anonymous_name = _build_anonymous_name(
            attempt_id=attempt.id,
            user_id=attempt.user_id,
            exam_id=attempt.exam_id,
        )
        student_display = attempt.user.get_full_name() or attempt.user.username if can_view_name else anonymous_name

    if request.method == "POST":
        if not request_has_permission(request, "grade.input"):
            messages.error(
                request,
                pgettext_lazy("exams.view.results.message", "grading_permission_required"),
            )
            return redirect(view_attempt_url)

        # ✅ DOUBLE-CHECK: POST zamanı da yoxla
        if _attempt_review_window_locked(attempt, current_time=timezone.now()):
            messages.error(
                request,
                pgettext_lazy("exams.view.results.message", "cannot_edit_after_five_minutes"),
            )
            return redirect(view_attempt_url)

        total_score = 0
        any_score = False
        from apps.notifications.services import notify_student_about_feedback

        for a in answers_qs:
            q = a.question

            score_raw = (request.POST.get(f"score_{q.id}") or "").strip()
            max_points_raw = (request.POST.get(f"max_points_{q.id}") or "").strip()
            feedback = (request.POST.get(f"feedback_{q.id}") or "").strip()

            if max_points_raw:
                try:
                    max_points_val = int(max_points_raw)
                except ValueError:
                    max_points_val = q.points
                max_points_val = max(1, max_points_val)
            else:
                max_points_val = max(1, q.points)

            if score_raw == "":
                a.teacher_score = None
            else:
                try:
                    score_val = int(score_raw)
                except ValueError:
                    score_val = 0
                score_val = max(0, score_val)
                if score_val > max_points_val:
                    max_points_val = score_val
                a.teacher_score = score_val
                total_score += score_val
                any_score = True

            if q.points != max_points_val:
                q.points = max_points_val
                q.save(update_fields=["points"])

            a.teacher_feedback = feedback
            a.save(update_fields=["teacher_score", "teacher_feedback", "updated_at"])

        # İlk yoxlama vaxtını saxla; 5 dəqiqəlik redaktə pəncərəsi bu vaxtdan hesablanır.
        attempt.teacher_score = total_score if any_score else None
        attempt.checked_by_teacher = True
        if not attempt.teacher_checked_at:
            attempt.teacher_checked_at = timezone.now()
        attempt.save(update_fields=["teacher_score", "checked_by_teacher", "teacher_checked_at"])
        notify_student_about_feedback(
            task=exam,
            student=attempt.user,
            task_kind="exam",
            extra_metadata={"attempt_id": attempt.id},
        )

        messages.success(request, pgettext_lazy("exams.view.results.message", "attempt_checked_success"))
        return redirect(results_return_url)

    context = {
        "exam": exam,
        "attempt": attempt,
        "attempt_timing": _build_attempt_timing_context(attempt),
        "qa_list": qa_list,
        "profile_return_url": profile_return_url,
        "results_return_url": results_return_url,
        "student_display": student_display,
        "exam_creator_display": _user_display_name(exam.author),
        "can_view_student_identity": can_view_name,
        "identity_window_seconds_left": identity_window_seconds,
    }
    return render(request, "exams/teacher/teacher_check_attempt.html", context)


@login_required
@require_http_methods(["POST"])
def ai_grade_answer(request, slug, attempt_id):
    """AJAX endpoint — AI grades a single written answer."""
    from django.http import JsonResponse

    from apps.exams.services.ai_grading import grade_written_answer

    _ensure_teacher(request.user)
    exam = get_teacher_exam_or_404(request, slug=slug)
    attempt = get_object_or_404(ExamAttempt, id=attempt_id, exam=exam)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    question_id = data.get("question_id")
    if not question_id:
        return JsonResponse({"ok": False, "error": "question_id required"}, status=400)

    answer = (
        attempt.answers.select_related("question").prefetch_related("files").filter(question_id=question_id).first()
    )
    if not answer:
        return JsonResponse({"ok": False, "error": "Answer not found"}, status=404)

    q = answer.question
    try:
        max_points = int(data.get("max_points", q.points) or q.points)
    except (TypeError, ValueError):
        max_points = q.points
    max_points = max(1, max_points)

    result = grade_written_answer(
        question_text=q.text,
        student_answer=answer.text_answer or "",
        max_points=int(max_points),
        correct_answer=getattr(q, "correct_answer", "") or "",
        language_code=request.LANGUAGE_CODE,
        user_id=request.user.pk,
        answer_files=answer.files.all(),
        paint_image=getattr(answer, "paint_image", None),
    )
    return JsonResponse(result)


@login_required
def teacher_pending_attempts(request):
    """
    Müəllimin bütün imtahanlarından yığılmış,
    yoxlanılmağı gözləyən (Pending) işlərin siyahısı.
    """
    _ensure_teacher(request.user)

    teacher_exams = tenant_scoped_exams(request, request.user.exams.all())

    # Yoxlanılacaq işləri tapırıq
    pending_attempts = (
        ExamAttempt.objects.filter(
            exam__in=teacher_exams,  # Bu müəllimin aktiv tenant imtahanları
            status__in=["submitted", "expired"],  # Bitmiş imtahanlar
            checked_by_teacher=False,  # Hələ yoxlanmayıb
        )
        .exclude(exam__exam_type="test")  # Testləri çıxarırıq
        .select_related("user", "exam")
        .order_by("finished_at")
    )

    now = timezone.now()
    attempts_data = []

    for att in pending_attempts:
        anonymous_name = _build_anonymous_name(attempt_id=att.id, user_id=att.user_id, exam_id=att.exam_id)

        can_view_name, seconds_remaining = _resolve_attempt_name_visibility(att, current_time=now)

        attempts_data.append(
            {
                "attempt": att,
                "anonymous_name": anonymous_name,
                "real_name": att.user.get_full_name() or att.user.username,
                "can_view_name": can_view_name,
                "seconds_remaining": seconds_remaining,
            }
        )

    # ═══════════════════════════════════════════════════════════════════
    # ✅ YENİ: Tip üzrə saylar (Yazılı və Praktiki)
    # ═══════════════════════════════════════════════════════════════════
    essay_count = sum(1 for att in pending_attempts if att.exam.exam_type == "written")
    practical_count = sum(1 for att in pending_attempts if att.exam.exam_type == "coding")

    context = {
        "pending_attempts": pending_attempts,
        "attempts_data": attempts_data,  # ✅ YENİ - anonim adlar
        "essay_count": essay_count,  # ✅ YENİ - yazılı say
        "practical_count": practical_count,  # ✅ YENİ - praktiki say
    }
    return render(request, "exams/teacher/teacher_pending_attempts.html", context)
