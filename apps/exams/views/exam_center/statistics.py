"""exam_center paketi — imtahan statistikaları / nəticələr (SPA bölməsi backend-i).

İmtahan mərkəzi üçün bütün imtahan cəhdlərinin (nəticələrin) analitikası:
* filtr — fənn, qrup, kafedra/fakültə, imtahan tipi, il, mətn axtarışı;
* sayğaclar (summary) + səhifələnən cədvəl (lazy, performanslı);
* Excel export.

Data AJAX ilə gəlir (``exam_center_stats_data``); Excel ``exam_center_stats_export``.
İcazə: imtahan mərkəzi (rəhbər + işçi) və superadmin.
"""

from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.utils.translation import pgettext, pgettext_lazy
from django.views.decorators.http import require_GET

from apps.exams.models import ExamAttempt
from apps.exams.services.access_policy import is_exam_center_user
from apps.exams.services.result_calculation import attach_test_result_summaries
from apps.exams.services.supervision import attach_attempt_interventions

from ._shared import supervisor_org_or_403

_FINISHED = ("submitted", "expired")
_PAGE_SIZE = 15

# Tədris ili sentyabrda başlayır (AZ universitetləri): sentyabr–avqust = bir
# tədris ili. Cəhdin təqvim tarixindən tədris ilinin BAŞLANĞIC ilini çıxarırıq
# (məs. 17.07.2026 → 2025/2026, çünki iyul sentyabrdan əvvəldir).
_ACADEMIC_YEAR_START_MONTH = 9


def _academic_year_start(dt):
    """datetime → tədris ilinin başlanğıc təqvim ili (lokal vaxta görə)."""
    local = timezone.localtime(dt)
    return local.year if local.month >= _ACADEMIC_YEAR_START_MONTH else local.year - 1


# Semestr ay qrupları — apellyasiya statistikası ilə eyni bölgü (bax
# apps/appeals/views/teacher/statistics.py), ona görə iki panel eyni cavabı verir.
_SEMESTERS = {
    "fall": (9, 10, 11, 12, 1),
    "spring": (2, 3, 4, 5, 6),
    "summer": (7, 8),
}
# DİQQƏT: modul səviyyəsində `pgettext` dili İMPORT anına dondurur (apellyasiya
# tərəfindəki mövcud tələ). `pgettext_lazy` sorğunun dilinə görə həll olunur.
_SEMESTER_LABELS = {
    "fall": pgettext_lazy("exams.center.stats.semester", "Payız semestri"),
    "spring": pgettext_lazy("exams.center.stats.semester", "Yaz semestri"),
    "summer": pgettext_lazy("exams.center.stats.semester", "Yay semestri"),
}

# Cəhd statusları — yalnız bitmiş cəhdlər statistikaya düşdüyü üçün siyahı
# `_FINISHED` ilə uyğun saxlanılır.
_ATTEMPT_STATUS_LABELS = {
    "submitted": pgettext_lazy("exams.center.stats.status", "Təhvil verilib"),
    "expired": pgettext_lazy("exams.center.stats.status", "Vaxtı bitib"),
}
_ATTEMPT_STATUS_VALUES = set(_ATTEMPT_STATUS_LABELS)


# Sütun → DB order ifadəsi (klik-sıralama üçün). "score" xam düzgün cavab sayı
# (correct_count) ilə sıralanır — bal-proksi (bərabər-çəkili suallar üçün).
_SORT_MAP = {
    "student": ("user__first_name", "user__last_name", "user__username"),
    "exam": ("exam__title",),
    "subject": ("exam__subject__name",),
    "type": ("exam__exam_type_extended",),
    "score": ("correct_count",),
    "status": ("status",),
    "date": ("started_at",),
}


def _stats_org(request):
    organization = supervisor_org_or_403(request)
    if not is_exam_center_user(request.user):
        raise PermissionDenied(pgettext("exams.final_center.permission", "Bu bölmə yalnız imtahan mərkəzi üçündür."))
    return organization


def _csv_ints(raw):
    return [int(x) for x in (raw or "").split(",") if x.strip().isdigit()]


def _csv_uuids(raw):
    """UUID PK-lı sahələr (Subject, OrgUnit) üçün. `_csv_ints` UUID-ləri
    ATIRDI (isdigit False) → fənn/fakültə/kafedra filtri səssizcə işləməzdi."""
    import uuid as _uuid

    out = []
    for x in (raw or "").split(","):
        x = x.strip()
        if not x:
            continue
        try:
            out.append(str(_uuid.UUID(x)))
        except ValueError:
            continue
    return out


def _unit_ids_with_children(organization, unit_ids):
    """Seçilmiş bölmələr + onların birbaşa alt-bölmələri (fakültə → kafedralar)."""
    if not unit_ids:
        return []
    from apps.organizations.models import OrgUnit

    rows = OrgUnit.objects.filter(organization=organization).filter(Q(id__in=unit_ids) | Q(parent_id__in=unit_ids))
    return list(rows.values_list("id", flat=True))


def _filtered_attempts(request, organization):
    """GET filtrlərinə görə bitmiş (real) imtahan cəhdlərinin queryset-i."""
    qs = (
        ExamAttempt.objects.filter(exam__organization=organization, is_trial=False, status__in=_FINISHED)
        .select_related("user", "exam", "exam__subject", "exam__author")
        .prefetch_related("exam__allowed_groups__org_unit__parent")
    )

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(user__first_name__icontains=q) | Q(user__last_name__icontains=q) | Q(user__username__icontains=q)
        )

    subject_ids = _csv_uuids(request.GET.get("subjects"))
    if subject_ids:
        qs = qs.filter(exam__subject_id__in=subject_ids)

    group_ids = _csv_ints(request.GET.get("groups"))
    if group_ids:
        qs = qs.filter(exam__allowed_groups__id__in=group_ids)

    # Köhnə birləşmiş "units" filtri (geriyə-uyğunluq — bookmarked URL-lər).
    unit_ids = _csv_uuids(request.GET.get("units"))
    if unit_ids:
        qs = qs.filter(exam__allowed_groups__org_unit_id__in=_unit_ids_with_children(organization, unit_ids))

    # Ayrı FAKÜLTƏ filtri: qrupun org_unit-inin VALİDEYNİ fakültədir.
    faculty_ids = _csv_uuids(request.GET.get("faculties"))
    if faculty_ids:
        qs = qs.filter(exam__allowed_groups__org_unit__parent_id__in=faculty_ids)

    # Ayrı KAFEDRA filtri: qrupun org_unit-i birbaşa kafedradır.
    department_ids = _csv_uuids(request.GET.get("departments"))
    if department_ids:
        qs = qs.filter(exam__allowed_groups__org_unit_id__in=department_ids)

    # MÜƏLLİM (imtahan müəllifi) filtri.
    teacher_ids = _csv_ints(request.GET.get("teachers"))
    if teacher_ids:
        qs = qs.filter(exam__author_id__in=teacher_ids)

    exam_type = (request.GET.get("type") or "").strip()
    if exam_type:
        qs = qs.filter(exam__exam_type_extended=exam_type)

    # İMTAHAN FORMASI (test/yazılı/praktik) — apellyasiya statistikasındakı
    # `format` filtri ilə eyni parametr adı və eyni sahə.
    exam_format = (request.GET.get("format") or "").strip()
    if exam_format:
        qs = qs.filter(exam__exam_type=exam_format)

    # STATUS: burada cəhdin statusudur (apellyasiyada müraciətin statusu idi).
    status = (request.GET.get("status") or "").strip()
    if status in _ATTEMPT_STATUS_VALUES:
        qs = qs.filter(status=status)

    # TƏDRİS İLİ filtri: `year` tədris ilinin başlanğıc təqvim ilidir (məs. 2025 =
    # 2025/2026). Sentyabr–avqust aralığına görə süzülür.
    year = (request.GET.get("year") or "").strip()
    if year.isdigit():
        from datetime import datetime

        start = int(year)
        lo = timezone.make_aware(datetime(start, _ACADEMIC_YEAR_START_MONTH, 1))
        hi = timezone.make_aware(datetime(start + 1, _ACADEMIC_YEAR_START_MONTH, 1))
        qs = qs.filter(started_at__gte=lo, started_at__lt=hi)

    # SEMESTR: təqvim ayı qruplarına görə (tədris ili filtri ilə birləşə bilir).
    semester = (request.GET.get("semester") or "").strip()
    if semester in _SEMESTERS:
        qs = qs.filter(started_at__month__in=_SEMESTERS[semester])

    return qs.distinct()


def _sorted(qs, request):
    """``?sort=score|-score|...`` → order_by; default ən yeni (-started_at)."""
    raw = (request.GET.get("sort") or "").strip()
    key = raw[1:] if raw.startswith("-") else raw
    desc = raw.startswith("-")
    fields = _SORT_MAP.get(key)
    if not fields:
        return qs.order_by("-started_at", "-id")
    prefix = "-" if desc else ""
    return qs.order_by(*[f"{prefix}{f}" for f in fields], "-id")


def _dedup(names):
    seen, out = set(), []
    for n in names:
        if n and n not in seen:
            seen.add(n)
            out.append(n)
    return out


def _row(attempt):
    result = getattr(attempt, "test_result", None)
    exam = attempt.exam
    subject = getattr(exam, "subject", None)
    groups = list(exam.allowed_groups.all())
    kafedras = _dedup(g.org_unit.name for g in groups if g.org_unit_id)
    faculties = _dedup(g.org_unit.parent.name for g in groups if g.org_unit_id and g.org_unit.parent_id)
    teacher = exam.author.get_full_name() or exam.author.username if exam.author_id else ""
    intervention = getattr(attempt, "exam_intervention", None)
    removed = bool(intervention and intervention.get("is_terminal"))
    return {
        "attempt_id": attempt.id,
        "exam_slug": exam.slug,
        "student": attempt.user.get_full_name() or attempt.user.username,
        "username": attempt.user.username,
        "group": ", ".join(_dedup(g.name for g in groups)),
        "kafedra": ", ".join(kafedras),
        "faculty": ", ".join(faculties),
        "teacher": teacher,
        "exam": exam.title,
        "subject": (f"{subject.code} — {subject.name}" if subject else ""),
        "type": exam.get_exam_type_extended_display() if exam.exam_type_extended else "",
        "percentage": (result.percentage_display if result else ""),
        "score": (f"{result.score}/{result.max_score}" if result else ""),
        "status": attempt.get_status_display(),
        # İmtahandan uzaqlaşdırılmış tələbə cədvəldə qırmızı görünsün + səbəb.
        "removed": removed,
        "removal_reason": (intervention.get("reason") if intervention else "") or "",
        "date": timezone.localtime(attempt.finished_at or attempt.started_at).strftime("%d.%m.%Y %H:%M"),
    }


@login_required
@require_GET
def exam_center_stats_data(request):
    """Filtrlənmiş, səhifələnən nəticələr + sayğaclar (JSON)."""
    organization = _stats_org(request)
    attempts = _filtered_attempts(request, organization)

    summary = {
        "total": attempts.count(),
        "students": attempts.values("user_id").distinct().count(),
        "exams": attempts.values("exam_id").distinct().count(),
        "by_type": list(
            attempts.values("exam__exam_type_extended")
            .order_by("exam__exam_type_extended")
            .annotate(n=Count("id", distinct=True))
        ),
    }

    page_obj = Paginator(_sorted(attempts, request), _PAGE_SIZE).get_page(request.GET.get("page"))
    page_attempts = list(page_obj.object_list)
    attach_test_result_summaries(page_attempts)
    attach_attempt_interventions(page_attempts)

    return JsonResponse(
        {
            "results": [_row(a) for a in page_attempts],
            "summary": summary,
            "pagination": {
                "page": page_obj.number,
                "num_pages": page_obj.paginator.num_pages,
                "has_prev": page_obj.has_previous(),
                "has_next": page_obj.has_next(),
                "count": page_obj.paginator.count,
            },
        }
    )


@login_required
@require_GET
def exam_center_stats_export(request):
    """Filtrlənmiş bütün nəticələrin Excel (.xlsx) ixracı."""
    organization = _stats_org(request)
    attempts = list(_sorted(_filtered_attempts(request, organization), request)[:5000])
    attach_test_result_summaries(attempts)
    attach_attempt_interventions(attempts)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl not installed", status=500)

    columns = [
        ("Tələbə", "student"),
        ("İstifadəçi", "username"),
        ("Qrup", "group"),
        ("Kafedra", "kafedra"),
        ("Fakültə / dekanlıq", "faculty"),
        ("Müəllim", "teacher"),
        ("İmtahan", "exam"),
        ("Fənn", "subject"),
        ("Tip", "type"),
        ("Bal", "score"),
        ("Faiz", "percentage"),
        ("Status", "status"),
        ("Uzaqlaşdırma səbəbi", "removal_reason"),
        ("Tarix", "date"),
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Statistika"
    for col, (title, _key) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for r, a in enumerate(attempts, start=2):
        row = _row(a)
        for c, (_title, key) in enumerate(columns, start=1):
            ws.cell(row=r, column=c, value=row[key])
    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    stamp = timezone.localtime().strftime("%Y%m%d-%H%M")
    response["Content-Disposition"] = f'attachment; filename="imtahan-statistika-{stamp}.xlsx"'
    wb.save(response)
    return response


def _lookup_bounds(request):
    """(offset, limit) — axtarışlı seçicilər üçün lazy səhifələmə (default 10)."""
    try:
        offset = max(0, int(request.GET.get("offset", 0)))
    except (TypeError, ValueError):
        offset = 0
    try:
        limit = int(request.GET.get("limit", 10))
    except (TypeError, ValueError):
        limit = 10
    return offset, max(1, min(limit, 50))


def _lookup_page(qs, offset, limit, serializer):
    """`limit + 1` gətir → `has_more` təyin et (infinite scroll üçün)."""
    window = list(qs[offset : offset + limit + 1])
    return [serializer(o) for o in window[:limit]], len(window) > limit


@login_required
@require_GET
def stats_faculty_search(request):
    """Fakültə/dekanlıq axtarışı (OrgUnit) — statistika filtrləri üçün, lazy."""
    organization = _stats_org(request)
    from apps.organizations.models import OrgUnit

    query = (request.GET.get("q") or "").strip()
    qs = OrgUnit.objects.filter(organization=organization, is_active=True, unit_type__in=("faculty", "deanery"))
    if query:
        qs = qs.filter(name__icontains=query)
    qs = qs.order_by("name")

    offset, limit = _lookup_bounds(request)
    results, has_more = _lookup_page(qs, offset, limit, lambda u: {"id": str(u.id), "text": u.name})
    return JsonResponse({"results": results, "has_more": has_more})


@login_required
@require_GET
def stats_department_search(request):
    """Kafedra/bölmə axtarışı — `?faculty=<id>` verilibsə həmin fakültəyə görə daralır."""
    organization = _stats_org(request)
    from apps.organizations.models import OrgUnit

    query = (request.GET.get("q") or "").strip()
    qs = OrgUnit.objects.filter(organization=organization, is_active=True, unit_type__in=("chair", "department"))
    faculty_ids = _csv_uuids(request.GET.get("faculty"))
    if faculty_ids:
        qs = qs.filter(parent_id__in=faculty_ids)
    if query:
        qs = qs.filter(name__icontains=query)
    qs = qs.order_by("name")

    offset, limit = _lookup_bounds(request)
    results, has_more = _lookup_page(qs, offset, limit, lambda u: {"id": str(u.id), "text": u.name})
    return JsonResponse({"results": results, "has_more": has_more})


@login_required
@require_GET
def stats_teacher_search(request):
    """Müəllim (imtahan müəllifi) axtarışı — ad/soyad/istifadəçi adı üzrə, lazy."""
    organization = _stats_org(request)
    from apps.exams.models import Exam

    query = (request.GET.get("q") or "").strip()
    rows = Exam.objects.filter(organization=organization, author__isnull=False)
    if query:
        rows = rows.filter(
            Q(author__first_name__icontains=query)
            | Q(author__last_name__icontains=query)
            | Q(author__username__icontains=query)
        )
    rows = (
        rows.values("author_id", "author__first_name", "author__last_name", "author__username")
        .distinct()
        .order_by("author__first_name", "author__last_name")
    )

    offset, limit = _lookup_bounds(request)
    results, has_more = _lookup_page(
        rows,
        offset,
        limit,
        lambda r: {
            "id": str(r["author_id"]),
            "text": (f"{r['author__first_name']} {r['author__last_name']}".strip() or r["author__username"]),
        },
    )
    return JsonResponse({"results": results, "has_more": has_more})


@login_required
@require_GET
def exam_center_stats_filters(request):
    """Filtr seçimləri üçün metadata (illər, tiplər, kafedralar)."""
    organization = _stats_org(request)
    from apps.exams.models import Exam
    from apps.organizations.models import OrgUnit

    exam_type_choices = Exam.EXAM_TYPE_EXTENDED_CHOICES

    # Tədris illəri (2025/2026 formatı) — cəhd tarixlərindən sentyabr sərhədi ilə.
    year_starts = sorted(
        {
            _academic_year_start(d)
            for d in ExamAttempt.objects.filter(
                exam__organization=organization, is_trial=False, status__in=_FINISHED
            ).values_list("started_at", flat=True)
        },
        reverse=True,
    )
    academic_years = [{"value": str(s), "label": f"{s}/{s + 1}"} for s in year_starts]
    units = list(
        OrgUnit.objects.filter(organization=organization, is_active=True)
        .order_by("name")
        .values("id", "name", "unit_type", "parent_id")
    )
    # Fakültə (faculty/deanery) və kafedra (chair/department) ayrı siyahılara bölünür
    # ki, iki müstəqil filtr olsun (kaskad üçün kafedrada parent_id qalır).
    faculty_types = {"faculty", "deanery"}
    department_types = {"chair", "department"}
    faculties = [{"id": u["id"], "name": u["name"]} for u in units if u["unit_type"] in faculty_types]
    departments = [
        {"id": u["id"], "name": u["name"], "parent_id": u["parent_id"]}
        for u in units
        if u["unit_type"] in department_types
    ]

    teachers = [
        {
            "id": row["author_id"],
            "name": (f"{row['author__first_name']} {row['author__last_name']}".strip() or row["author__username"]),
        }
        for row in (
            Exam.objects.filter(organization=organization, author__isnull=False)
            .values("author_id", "author__first_name", "author__last_name", "author__username")
            .distinct()
            .order_by("author__first_name", "author__last_name")
        )
    ]

    return JsonResponse(
        {
            "academic_years": academic_years,
            "units": units,  # geriyə-uyğunluq
            "faculties": faculties,
            "departments": departments,
            "teachers": teachers,
            "types": [{"value": v, "label": str(lbl)} for v, lbl in exam_type_choices],
            # Apellyasiya statistikası ilə paritet: forma/status/semestr.
            "formats": [{"value": v, "label": str(lbl)} for v, lbl in Exam.EXAM_TYPE_CHOICES],
            "statuses": [{"value": v, "label": str(lbl)} for v, lbl in _ATTEMPT_STATUS_LABELS.items()],
            "semesters": [{"value": k, "label": str(_SEMESTER_LABELS[k])} for k in ("fall", "spring", "summer")],
        }
    )


__all__ = [
    "exam_center_stats_data",
    "exam_center_stats_export",
    "exam_center_stats_filters",
    "stats_faculty_search",
    "stats_department_search",
    "stats_teacher_search",
]
