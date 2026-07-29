"""Final imtahan hesabatının Excel iş kitabını qurur (openpyxl rendering).

Data yığımı `xlsx_report.py`-dədir; burada yalnız vərəqlərin çəkilməsi var:
başlıq üslubu, dondurulmuş sətir, avtofiltr, sütun eni və status rəngləri.
"""

from __future__ import annotations

from collections import OrderedDict, defaultdict

from django.utils import timezone
from django.utils.translation import pgettext

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .xlsx_report import (
    INCIDENT_COLUMNS,
    ROOM_COLUMNS,
    _collect,
    _fmt_dt,
    _person,
    _sheet_title,
)

HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13, color="0F172A")
LABEL_FONT = Font(bold=True, size=10, color="334155")
THIN = Side(style="thin", color="D9E2EF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Tələbə sətrinin nəticəyə görə fon rəngi — hesabata gözlə baxanda
# çıxarılmış/gəlməyən tələbə dərhal seçilsin.
STATUS_FILLS = {
    "completed": PatternFill("solid", fgColor="F0FDF4"),
    "removed": PatternFill("solid", fgColor="FEF2F2"),
    "absent": PatternFill("solid", fgColor="F8FAFC"),
}

SEVERITY_FILLS = {
    "critical": PatternFill("solid", fgColor="FEE2E2"),
    "high": PatternFill("solid", fgColor="FFEDD5"),
    "medium": PatternFill("solid", fgColor="FEF9C3"),
}


def _write_header(sheet, columns, *, row=1):
    for index, (title, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=row, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[row].height = 30
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1)
    sheet.auto_filter.ref = f"A{row}:{get_column_letter(len(columns))}{row}"


def _write_summary(sheet, *, organization, meta, totals):
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 52

    sheet["A1"] = pgettext("exams.final_center.report", "Final imtahan hesabatı")
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = organization.name if organization else ""
    sheet["A2"].font = Font(size=11, color="475569")

    row = 4
    for label, value in meta:
        sheet.cell(row=row, column=1, value=label).font = LABEL_FONT
        sheet.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value=pgettext("exams.final_center.report", "Göstəricilər")).font = TITLE_FONT
    row += 1
    for label, value in totals:
        sheet.cell(row=row, column=1, value=label).font = LABEL_FONT
        cell = sheet.cell(row=row, column=2, value=value)
        cell.font = Font(bold=True, size=11)
        row += 1


def _write_room_sheet(sheet, rows):
    _write_header(sheet, ROOM_COLUMNS)
    for order, row in enumerate(rows, start=1):
        values = list(row["values"])
        values[0] = order
        fill = STATUS_FILLS.get(row["ticket"].status)
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=order + 1, column=index, value=value)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=index in (3, 4, 6, 32, 36))
            if fill:
                cell.fill = fill


def _incident_note(incident) -> str:
    metadata = incident.metadata if isinstance(incident.metadata, dict) else {}
    parts = []
    for key in ("key", "reason", "detail", "message", "shortcut", "url"):
        value = metadata.get(key)
        if value:
            parts.append(f"{key}: {value}")
    return "; ".join(parts)


def _write_incident_sheet(sheet, incident_rows):
    _write_header(sheet, INCIDENT_COLUMNS)
    for order, item in enumerate(incident_rows, start=1):
        incident = item["incident"]
        values = [
            order,
            _fmt_dt(incident.timestamp),
            _person(incident.student),
            incident.student.username,
            item["group"],
            item["exam"],
            item["room"],
            item["seat"],
            incident.get_event_type_display(),
            incident.get_severity_display(),
            incident.violation_count_at_time,
            _incident_note(incident),
        ]
        fill = SEVERITY_FILLS.get(incident.severity)
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=order + 1, column=index, value=value)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=index == 12)
            if fill:
                cell.fill = fill


def build_final_report_workbook(organization, tickets, *, meta_rows=None) -> Workbook:
    """Filtrlənmiş biletlərdən tam hesabat iş kitabı qur.

    *tickets* — `filter_tickets(...)` nəticəsi (select_related edilmiş queryset
    və ya siyahı). Hər zal öz vərəqinə düşür, pozuntular ayrıca jurnaldadır.
    """
    rows = _collect(organization, tickets)

    by_room = OrderedDict()
    for row in rows:
        room = row["room"]
        key = room.pk if room else 0
        by_room.setdefault(key, {"room": room, "rows": []})["rows"].append(row)

    incident_rows = []
    status_totals = defaultdict(int)
    exams = set()
    for row in rows:
        ticket = row["ticket"]
        status_totals[ticket.status] += 1
        if ticket.exam_id:
            exams.add(ticket.exam_id)
        for incident in row["incidents"]:
            incident_rows.append(
                {
                    "incident": incident,
                    "group": row["academic"].get("group", ""),
                    "exam": ticket.exam.title if ticket.exam_id else "",
                    "room": row["room"].name if row["room"] else "",
                    "seat": ticket.seat_number if ticket.seat_number is not None else "",
                }
            )
    incident_rows.sort(key=lambda item: item["incident"].timestamp)

    workbook = Workbook()
    summary = workbook.active
    summary.title = pgettext("exams.final_center.report", "Xülasə")

    totals = [
        (pgettext("exams.final_center.report", "Zal sayı"), len([r for r in by_room.values() if r["room"]])),
        (pgettext("exams.final_center.report", "İmtahan sayı"), len(exams)),
        (pgettext("exams.final_center.report", "Tələbə sayı"), len(rows)),
        (pgettext("exams.final_center.report", "İmtahanı bitirib"), status_totals.get("completed", 0)),
        (pgettext("exams.final_center.report", "İmtahandan çıxarılıb"), status_totals.get("removed", 0)),
        (pgettext("exams.final_center.report", "Gəlməyib"), status_totals.get("absent", 0)),
        (pgettext("exams.final_center.report", "Qeydə alınan hadisə"), len(incident_rows)),
    ]
    meta = list(meta_rows or [])
    meta.append(
        (
            pgettext("exams.final_center.report", "Hesabat yaradıldı"),
            timezone.localtime(timezone.now()).strftime("%d.%m.%Y %H:%M"),
        )
    )
    _write_summary(summary, organization=organization, meta=meta, totals=totals)

    used_titles = {summary.title}
    for entry in by_room.values():
        room = entry["room"]
        raw = f"{room.code or room.name}" if room else pgettext("exams.final_center.report", "Zalsız")
        _write_room_sheet(workbook.create_sheet(_sheet_title(raw, used_titles)), entry["rows"])

    if incident_rows:
        title = _sheet_title(pgettext("exams.final_center.report", "Pozuntular"), used_titles)
        _write_incident_sheet(workbook.create_sheet(title), incident_rows)

    return workbook


__all__ = ["build_final_report_workbook"]
