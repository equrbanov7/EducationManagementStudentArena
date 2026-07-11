"""
View tests for exams app.
"""

import base64
import json
from datetime import timedelta
from io import BytesIO
from unittest.mock import patch
from urllib.parse import quote, urlencode

from django.contrib.auth import get_user_model
from django.core.exceptions import PermissionDenied
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, RequestFactory, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from django.utils.translation import override

from apps.accounts.models import ProfileRole
from apps.courses.models import Course, CourseMembership
from apps.exams.models import (
    Exam,
    ExamAnswer,
    ExamAnswerFile,
    ExamAttempt,
    ExamQuestion,
    ExamQuestionOption,
    ExamSupervisionConfig,
    QuestionBlock,
    StudentGroup,
)
from apps.organizations.models import Membership, Organization
from core.constants import OrganizationType

User = get_user_model()
_TINY_PNG_BYTES = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+aF9sAAAAASUVORK5CYII="
)


def _assign_user_to_org(user, organization, profile_role, *, membership_role_name=None):
    if membership_role_name is None:
        if profile_role == ProfileRole.TEACHER:
            membership_role_name = "teacher"
        elif profile_role == ProfileRole.STUDENT:
            membership_role_name = "student"
        elif profile_role == ProfileRole.ORG_ADMIN:
            admin_role_candidates = (
                "vice_rector",
                "dean",
                "director",
                "deputy_director",
                "section_head",
                "chair_head",
                "manager",
                "senior_instructor",
            )
            membership_role_name = next(
                (
                    role_name
                    for role_name in admin_role_candidates
                    if organization.roles.filter(name=role_name, is_active=True).exists()
                ),
                "member",
            )
        elif profile_role == ProfileRole.ORG_OWNER:
            membership_role_name = (
                organization.roles.filter(is_active=True).order_by("-level").values_list("name", flat=True).first()
                or "member"
            )
        else:
            membership_role_name = "member"

    profile = user.profile
    profile.organization = organization
    profile.organization_type = organization.org_type
    profile.role = profile_role
    profile.save(update_fields=["organization", "organization_type", "role", "updated_at"])

    Membership.objects.update_or_create(
        user=user,
        organization=organization,
        defaults={
            "role": organization.roles.get(name=membership_role_name),
            "is_primary": True,
            "is_active": True,
        },
    )


def _login_with_org(client, user, organization):
    client.force_login(user)
    session = client.session
    session["active_organization"] = organization.slug
    session.save()


class MyGroupsTenantIsolationTest(TestCase):
    def setUp(self):
        self.teacher = User.objects.create_user("teacher_groups", "teacher_groups@example.com", "StrongPass123!")
        self.teacher_a2 = User.objects.create_user("teacher_a2", "teacher_a2@example.com", "StrongPass123!")
        self.teacher_a3 = User.objects.create_user("teacher_a3", "teacher_a3@example.com", "StrongPass123!")
        self.teacher_a4 = User.objects.create_user("teacher_a4", "teacher_a4@example.com", "StrongPass123!")
        self.teacher_b = User.objects.create_user("teacher_b", "teacher_b@example.com", "StrongPass123!")
        self.student_a = User.objects.create_user("student_a", "student_a@example.com", "StrongPass123!")
        self.student_b = User.objects.create_user("student_b", "student_b@example.com", "StrongPass123!")
        self.member_a = User.objects.create_user("member_a", "member_a@example.com", "StrongPass123!")
        self.org_admin_a = User.objects.create_user("org_admin_a", "org_admin_a@example.com", "StrongPass123!")
        self.superadmin = User.objects.create_superuser(
            "superadmin_groups", "superadmin_groups@example.com", "StrongPass123!"
        )

        self.org_a = Organization.objects.create(
            name="Org A",
            org_type=OrganizationType.SCHOOL,
            owner=self.teacher,
            status="active",
            is_active=True,
        )
        self.org_b = Organization.objects.create(
            name="Org B",
            org_type=OrganizationType.UNIVERSITY,
            owner=self.teacher_b,
            status="active",
            is_active=True,
        )

        self._assign_profile(self.teacher, self.org_a, ProfileRole.TEACHER)
        self._assign_profile(self.teacher_a2, self.org_a, ProfileRole.TEACHER)
        self._assign_profile(self.teacher_a3, self.org_a, ProfileRole.TEACHER)
        self._assign_profile(self.teacher_a4, self.org_a, ProfileRole.TEACHER)
        self._assign_profile(self.teacher_b, self.org_b, ProfileRole.TEACHER)
        self._assign_profile(self.student_a, self.org_a, ProfileRole.STUDENT)
        self._assign_profile(self.student_b, self.org_b, ProfileRole.STUDENT)
        self._assign_profile(self.member_a, self.org_a, ProfileRole.MEMBER)
        self._assign_profile(self.org_admin_a, self.org_a, ProfileRole.ORG_ADMIN)
        self._assign_profile(self.superadmin, self.org_a, ProfileRole.SUPERADMIN)

        self._login_as(self.teacher)
        self._set_active_org(self.org_a)

    def _assign_profile(self, user, organization, role):
        _assign_user_to_org(user, organization, role)

    def _login_as(self, user):
        self.client.force_login(user)

    def _set_active_org(self, organization):
        session = self.client.session
        session["active_organization"] = organization.slug
        session.save()

    def _group_payload(self, **overrides):
        payload = {
            "name": "Test Group",
            "students": [str(self.student_a.id)],
            "primary_teacher": str(self.teacher.id),
            "assigned_teachers": [str(self.teacher.id)],
        }
        payload.update(overrides)
        return payload

    def test_my_groups_page_links_to_create_group_template(self):
        # Qrup yaratma yalnız administratordadır — admin qutunu və linki görür.
        self._login_as(self.org_admin_a)
        self._set_active_org(self.org_a)
        response = self.client.get(reverse("exams:teacher_group_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("exams:create_student_group"))

        create_response = self.client.get(reverse("exams:create_student_group"))
        self.assertEqual(create_response.status_code, 200)
        self.assertTemplateUsed(create_response, "exams/teacher/create_student_group.html")
        self.assertContains(create_response, reverse("exams:teacher_create_group"))

    def test_teacher_cannot_create_or_manage_groups(self):
        # Bənd 2: adi müəllim (təşkilat sahibi/admini olmayan) qrup yarada/idarə
        # edə bilməz. teacher_a2 org_a-nın sahibi deyil — sadəcə müəllimdir.
        self._login_as(self.teacher_a2)
        self._set_active_org(self.org_a)
        # Yaratma səhifəsi və POST → 403.
        self.assertEqual(self.client.get(reverse("exams:create_student_group")).status_code, 403)
        self.assertEqual(
            self.client.post(reverse("exams:teacher_create_group"), self._group_payload(name="No")).status_code,
            403,
        )
        # Siyahı görünür (oxu), amma "yeni qrup" düyməsi gizlidir.
        list_response = self.client.get(reverse("exams:teacher_group_list"))
        self.assertEqual(list_response.status_code, 200)
        self.assertNotContains(list_response, reverse("exams:create_student_group"))

    def test_group_manage_permission_delegation(self):
        # Faza 3: superadmin "group.manage" icazəsini bir rola verə bilər →
        # həmin roldakı istifadəçi (adi müəllim olsa belə) qrup yarada bilər.
        from apps.organizations.models import Membership, Role
        from core.constants import RoleScopeType

        deleg_role = Role.objects.create(
            organization=self.org_a,
            name="group_manager",
            level=55,
            scope_type=RoleScopeType.ORGANIZATION,
            permissions=["group.manage"],
            is_active=True,
        )
        Membership.objects.create(user=self.teacher_a2, organization=self.org_a, role=deleg_role, is_active=True)
        self._login_as(self.teacher_a2)
        self._set_active_org(self.org_a)
        response = self.client.post(
            reverse("exams:teacher_create_group"),
            self._group_payload(
                name="Deleg Group",
                primary_teacher=str(self.teacher_a2.id),
                assigned_teachers=[str(self.teacher_a2.id)],
            ),
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(StudentGroup.objects.filter(name="Deleg Group").exists())

    def test_groups_are_created_and_listed_per_active_tenant(self):
        # Yaratma admin əməliyyatıdır; org_a-da org_admin qrup yaradır və
        # müəllimi (primary_teacher) təyin edir.
        self._login_as(self.org_admin_a)
        self._set_active_org(self.org_a)
        response = self.client.post(
            reverse("exams:teacher_create_group"),
            self._group_payload(
                name="A Group", primary_teacher=str(self.teacher.id), assigned_teachers=[str(self.teacher.id)]
            ),
        )
        self.assertEqual(response.status_code, 302)
        group_a = StudentGroup.objects.get(name="A Group")
        self.assertEqual(group_a.organization, self.org_a)
        self.assertEqual(group_a.teacher, self.teacher)

        # Org B-nin adminini modelləmirik; tenant-izolyasiyanı yoxlamaq üçün org_b
        # qrupunu birbaşa yaradırıq (müəllim yalnız öz təşkilatının qruplarını görür).
        group_b = StudentGroup.objects.create(teacher=self.teacher_b, organization=self.org_b, name="B Group")
        group_b.teachers.add(self.teacher_b)
        group_b.students.add(self.student_b)

        self._login_as(self.teacher)
        self._set_active_org(self.org_a)
        response = self.client.get(reverse("exams:teacher_group_list"))
        self.assertContains(response, "A Group")
        self.assertNotContains(response, "B Group")

        self._login_as(self.teacher_b)
        self._set_active_org(self.org_b)
        response = self.client.get(reverse("exams:teacher_group_list"))
        self.assertContains(response, "B Group")
        self.assertNotContains(response, "A Group")

    def test_group_creation_rejects_cross_tenant_students(self):
        self._login_as(self.org_admin_a)
        self._set_active_org(self.org_a)
        response = self.client.post(
            reverse("exams:teacher_create_group"),
            self._group_payload(name="Invalid Group", students=[str(self.student_b.id)]),
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(StudentGroup.objects.filter(name="Invalid Group").exists())

    def test_admin_assigns_subjects_to_group(self):
        # Faza 1: admin qrup yaradarkən registrar.Subject fənlərini təyin edir.
        from apps.registrar.models import Subject

        subject = Subject.objects.create(organization=self.org_a, code="MATH1", name="Riyaziyyat")
        self._login_as(self.org_admin_a)
        self._set_active_org(self.org_a)
        response = self.client.post(
            reverse("exams:teacher_create_group"),
            self._group_payload(
                name="Fenn Qrupu",
                primary_teacher=str(self.teacher.id),
                assigned_teachers=[str(self.teacher.id)],
                subjects=[str(subject.id)],
            ),
        )
        self.assertEqual(response.status_code, 302)
        group = StudentGroup.objects.get(name="Fenn Qrupu")
        self.assertIn(subject, list(group.subjects.all()))

    def test_group_list_contains_edit_and_delete_routes(self):
        # Redaktə/silmə düymələri yalnız administratora görünür.
        self._login_as(self.org_admin_a)
        self._set_active_org(self.org_a)
        group = StudentGroup.objects.create(teacher=self.teacher, organization=self.org_a, name="Route Group")
        group.students.add(self.student_a)
        group.teachers.add(self.teacher)

        response = self.client.get(reverse("exams:teacher_group_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("exams:teacher_update_group", args=[0]))
        self.assertContains(response, reverse("exams:teacher_delete_group", args=[group.id]))

    def test_add_single_student_to_group(self):
        # U6.2 — manual add (resit / transfer student) via the single-add endpoint.
        group = StudentGroup.objects.create(teacher=self.teacher, organization=self.org_a, name="Add Group")
        self._login_as(self.org_admin_a)
        self._set_active_org(self.org_a)
        resp = self.client.post(reverse("exams:teacher_add_student_to_group", args=[group.id, self.student_a.id]))
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(group.students.filter(id=self.student_a.id).exists())

    def test_add_student_rejects_cross_tenant(self):
        group = StudentGroup.objects.create(teacher=self.teacher, organization=self.org_a, name="XT Add Group")
        self._login_as(self.org_admin_a)
        self._set_active_org(self.org_a)
        # student_b belongs to org_b → not addable to an org_a group.
        resp = self.client.post(reverse("exams:teacher_add_student_to_group", args=[group.id, self.student_b.id]))
        self.assertEqual(resp.status_code, 404)
        self.assertFalse(group.students.filter(id=self.student_b.id).exists())

    def test_add_student_requires_post(self):
        group = StudentGroup.objects.create(teacher=self.teacher, organization=self.org_a, name="GET Add Group")
        self._login_as(self.teacher)
        self._set_active_org(self.org_a)
        resp = self.client.get(reverse("exams:teacher_add_student_to_group", args=[group.id, self.student_a.id]))
        self.assertEqual(resp.status_code, 405)

    def test_student_cannot_access_group_management_routes(self):
        group = StudentGroup.objects.create(teacher=self.teacher, organization=self.org_a, name="Protected Group")
        group.students.add(self.student_a)
        group.teachers.add(self.teacher)

        self._login_as(self.student_a)
        self._set_active_org(self.org_a)

        self.assertEqual(self.client.get(reverse("exams:teacher_group_list")).status_code, 403)
        self.assertEqual(self.client.get(reverse("exams:create_student_group")).status_code, 403)
        self.assertEqual(
            self.client.post(reverse("exams:teacher_create_group"), self._group_payload(name="Blocked")).status_code,
            403,
        )
        self.assertEqual(
            self.client.post(
                reverse("exams:teacher_update_group", args=[group.id]), self._group_payload(name="Blocked Update")
            ).status_code,
            403,
        )
        self.assertEqual(self.client.post(reverse("exams:teacher_delete_group", args=[group.id])).status_code, 403)

    def test_teacher_can_multi_assign_teachers(self):
        self._login_as(self.teacher)
        self._set_active_org(self.org_a)
        response = self.client.post(
            reverse("exams:teacher_create_group"),
            self._group_payload(
                name="Teacher Multi Allowed",
                primary_teacher=str(self.teacher.id),
                assigned_teachers=[str(self.teacher.id), str(self.teacher_a2.id)],
            ),
        )
        self.assertEqual(response.status_code, 302)
        group = StudentGroup.objects.get(name="Teacher Multi Allowed")
        self.assertSetEqual(set(group.teachers.values_list("id", flat=True)), {self.teacher.id, self.teacher_a2.id})

    def test_org_admin_can_access_teacher_groups_url_and_multi_assign(self):
        self._login_as(self.org_admin_a)
        self._set_active_org(self.org_a)
        response = self.client.get(reverse("exams:teacher_group_list"))
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            reverse("exams:teacher_create_group"),
            self._group_payload(
                name="Admin Group",
                primary_teacher=str(self.teacher.id),
                assigned_teachers=[str(self.teacher.id), str(self.teacher_a2.id)],
            ),
        )
        self.assertEqual(response.status_code, 302)
        group = StudentGroup.objects.get(name="Admin Group")
        self.assertEqual(group.organization, self.org_a)
        self.assertEqual(group.teacher, self.teacher)
        self.assertSetEqual(set(group.teachers.values_list("id", flat=True)), {self.teacher.id, self.teacher_a2.id})

    def test_org_admin_cannot_assign_more_than_three_teachers(self):
        self._login_as(self.org_admin_a)
        self._set_active_org(self.org_a)
        response = self.client.post(
            reverse("exams:teacher_create_group"),
            self._group_payload(
                name="Too Many Teachers",
                primary_teacher=str(self.teacher.id),
                assigned_teachers=[
                    str(self.teacher.id),
                    str(self.teacher_a2.id),
                    str(self.teacher_a3.id),
                    str(self.teacher_a4.id),
                ],
            ),
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(StudentGroup.objects.filter(name="Too Many Teachers").exists())

    def test_org_admin_cannot_assign_cross_tenant_teachers(self):
        self._login_as(self.org_admin_a)
        self._set_active_org(self.org_a)
        response = self.client.post(
            reverse("exams:teacher_create_group"),
            self._group_payload(
                name="Cross Tenant Teacher Block",
                primary_teacher=str(self.teacher_b.id),
                assigned_teachers=[str(self.teacher_b.id)],
            ),
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(StudentGroup.objects.filter(name="Cross Tenant Teacher Block").exists())

    def test_non_student_member_cannot_create_group(self):
        # Bənd 2: adi üzv (MEMBER) qrup yarada bilməz — yaratma yalnız
        # superadmin/təşkilat sahibi/administratorundadır.
        self._login_as(self.member_a)
        self._set_active_org(self.org_a)
        response = self.client.post(
            reverse("exams:teacher_create_group"),
            self._group_payload(name="Member Created Group", primary_teacher=str(self.teacher.id)),
        )
        self.assertEqual(response.status_code, 403)
        self.assertFalse(StudentGroup.objects.filter(name="Member Created Group").exists())

    def test_update_and_delete_routes_are_tenant_scoped(self):
        group_b = StudentGroup.objects.create(teacher=self.teacher_b, organization=self.org_b, name="OrgB Group")
        group_b.students.add(self.student_b)
        group_b.teachers.add(self.teacher_b)

        self._login_as(self.teacher)
        self._set_active_org(self.org_a)
        update_response = self.client.post(
            reverse("exams:teacher_update_group", args=[group_b.id]),
            self._group_payload(name="Should Not Update"),
        )
        delete_response = self.client.post(reverse("exams:teacher_delete_group", args=[group_b.id]))

        self.assertEqual(update_response.status_code, 404)
        self.assertEqual(delete_response.status_code, 404)
        self.assertTrue(StudentGroup.objects.filter(id=group_b.id).exists())

    def test_non_superadmin_cannot_switch_session_to_other_tenant(self):
        self._login_as(self.teacher)
        self._set_active_org(self.org_b)
        response = self.client.get(reverse("exams:teacher_group_list"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.client.session.get("active_organization"), self.org_a.slug)

    def test_superadmin_can_manage_any_active_tenant(self):
        self._login_as(self.superadmin)

        self._set_active_org(self.org_a)
        response_a = self.client.post(
            reverse("exams:teacher_create_group"),
            self._group_payload(
                name="Super A",
                students=[str(self.student_a.id)],
                primary_teacher=str(self.teacher.id),
                assigned_teachers=[str(self.teacher.id), str(self.teacher_a2.id)],
            ),
        )
        self.assertEqual(response_a.status_code, 302)

        self._set_active_org(self.org_b)
        response_b = self.client.post(
            reverse("exams:teacher_create_group"),
            {
                "name": "Super B",
                "students": [str(self.student_b.id)],
                "primary_teacher": str(self.teacher_b.id),
                "assigned_teachers": [str(self.teacher_b.id)],
            },
        )
        self.assertEqual(response_b.status_code, 302)

        self.assertTrue(StudentGroup.objects.filter(name="Super A", organization=self.org_a).exists())
        self.assertTrue(StudentGroup.objects.filter(name="Super B", organization=self.org_b).exists())


class TeacherExamListOwnershipFilteringTest(TestCase):
    def setUp(self):
        self.teacher = User.objects.create_user(
            username="teacher_exam_owner",
            email="teacher_exam_owner@example.com",
            password="StrongPass123!",
        )
        self.other_teacher = User.objects.create_user(
            username="teacher_exam_other",
            email="teacher_exam_other@example.com",
            password="StrongPass123!",
        )
        self.student = User.objects.create_user(
            username="teacher_exam_student",
            email="teacher_exam_student@example.com",
            password="StrongPass123!",
        )
        self.org_admin = User.objects.create_user(
            username="teacher_exam_org_admin",
            email="teacher_exam_org_admin@example.com",
            password="StrongPass123!",
        )

        self.org_a = Organization.objects.create(
            name="Exam Org A",
            org_type=OrganizationType.SCHOOL,
            owner=self.teacher,
            status="active",
            is_active=True,
        )
        self.org_b = Organization.objects.create(
            name="Exam Org B",
            org_type=OrganizationType.UNIVERSITY,
            owner=self.teacher,
            status="active",
            is_active=True,
        )

        _assign_user_to_org(self.teacher, self.org_a, ProfileRole.TEACHER)
        _assign_user_to_org(self.other_teacher, self.org_a, ProfileRole.TEACHER)
        _assign_user_to_org(self.student, self.org_a, ProfileRole.STUDENT)
        _assign_user_to_org(self.org_admin, self.org_a, ProfileRole.ORG_ADMIN)

        self.exam_visible = Exam.objects.create(
            author=self.teacher,
            title="Visible Exam",
            is_active=True,
        )
        self.exam_other_tenant = Exam.objects.create(
            author=self.teacher,
            title="Other Tenant Exam",
            organization=self.org_b,
            is_active=True,
        )
        self.exam_other_author = Exam.objects.create(
            author=self.other_teacher,
            title="Other Author Exam",
            is_active=True,
        )
        self.course = Course.objects.create(
            owner=self.teacher,
            title="Teacher Exam Course",
            status="published",
        )
        self.exam_question = ExamQuestion.objects.create(
            exam=self.exam_visible,
            text="What is Python?",
            order=1,
            answer_mode="single",
        )
        ExamQuestionOption.objects.create(question=self.exam_question, text="Language", is_correct=True)
        ExamQuestionOption.objects.create(question=self.exam_question, text="Browser", is_correct=False)

        self.client.force_login(self.teacher)
        session = self.client.session
        session["active_organization"] = self.org_a.slug
        session.save()

    def test_teacher_exam_list_redirects_to_profile_my_exams_section(self):
        response = self.client.get(reverse("exams:teacher_exam_list"))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, f"{reverse('accounts:profile')}?section=my-exams")

    def test_modal_create_exam_includes_course_hidden_field_when_requested_from_course_dashboard(self):
        response = self.client.get(
            reverse("exams:create_exam"),
            {"modal": "1", "course": str(self.course.id)},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="course_id"')
        self.assertContains(response, f'value="{self.course.id}"')

    def test_modal_create_exam_form_includes_random_question_count_with_default_fifty(self):
        response = self.client.get(
            reverse("exams:create_exam"),
            {"modal": "1"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertRegex(response.content.decode(), r'name="random_question_count"[^>]*value="50"')
        self.assertContains(response, 'name="fair_question_distribution_enabled"', html=False)
        self.assertContains(response, 'name="ai_difficulty_balance_enabled"', html=False)

    def test_modal_edit_written_exam_includes_random_question_count(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Written Random Count Exam",
            exam_type="written",
            organization=self.org_a,
            random_question_count=7,
        )

        response = self.client.get(
            reverse("exams:edit_exam", args=[written_exam.slug]),
            {"modal": "1"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn("data-random-question-group", content)
        self.assertRegex(content, r'name="random_question_count"[^>]*value="7"')

    def test_modal_edit_coding_exam_includes_random_question_count(self):
        coding_exam = Exam.objects.create(
            author=self.teacher,
            title="Practical Random Count Exam",
            exam_type="coding",
            organization=self.org_a,
            random_question_count=4,
        )

        response = self.client.get(
            reverse("exams:edit_exam", args=[coding_exam.slug]),
            {"modal": "1"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn("data-random-question-group", content)
        self.assertRegex(content, r'name="random_question_count"[^>]*value="4"')

    def test_modal_create_exam_links_new_exam_to_requested_course(self):
        response = self.client.post(
            reverse("exams:create_exam") + f"?modal=1&course={self.course.id}",
            {
                "modal": "1",
                "course_id": str(self.course.id),
                "title": "Linked From Course Dashboard",
                "description": "Created via dashboard modal",
                "exam_type": "test",
                "is_active": "on",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)

        created_exam = Exam.objects.get(title="Linked From Course Dashboard")
        self.assertEqual(created_exam.author, self.teacher)
        self.assertEqual(created_exam.course, self.course)
        self.assertEqual(created_exam.organization, self.org_a)

    def test_modal_create_exam_persists_custom_random_question_count(self):
        response = self.client.post(
            reverse("exams:create_exam") + "?modal=1",
            {
                "modal": "1",
                "title": "Custom Random Count Exam",
                "description": "Created with a custom student question count.",
                "exam_type": "test",
                "random_question_count": "20",
                "is_active": "on",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)

        created_exam = Exam.objects.get(title="Custom Random Count Exam")
        self.assertEqual(created_exam.random_question_count, 20)
        self.assertTrue(created_exam.fair_question_distribution_enabled)
        self.assertFalse(created_exam.ai_difficulty_balance_enabled)

    def test_modal_create_exam_can_disable_distribution_toggles(self):
        response = self.client.post(
            reverse("exams:create_exam") + "?modal=1",
            {
                "modal": "1",
                "title": "Manual Distribution Exam",
                "description": "Created with fairness toggles off.",
                "exam_type": "test",
                "random_question_count": "10",
                "fair_question_distribution_enabled": "false",
                "ai_difficulty_balance_enabled": "false",
                "is_active": "on",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)

        created_exam = Exam.objects.get(title="Manual Distribution Exam")
        self.assertFalse(created_exam.fair_question_distribution_enabled)
        self.assertFalse(created_exam.ai_difficulty_balance_enabled)

    def test_create_exam_requires_active_organization(self):
        from apps.exams.views.teacher.exams import createAndEditExamView

        request = RequestFactory().post(
            reverse("exams:create_exam"),
            {
                "title": "Blocked Without Org",
                "description": "Should not be created",
                "exam_type": "test",
                "is_active": "on",
            },
        )
        request.user = self.teacher
        request.organization = None
        request.org_memberships = []
        request.org_permissions = []

        with self.assertRaises(PermissionDenied):
            createAndEditExamView(request)

        self.assertFalse(Exam.objects.filter(title="Blocked Without Org").exists())

    def test_create_exam_full_page_redirects_after_single_membership_org_restore(self):
        session = self.client.session
        session.pop("active_organization", None)
        session.save()

        response = self.client.get(reverse("exams:create_exam"))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, f"{reverse('accounts:profile')}?section=my-exams")
        self.assertEqual(self.client.session.get("active_organization"), self.org_a.slug)

    def test_org_admin_can_open_and_submit_create_exam_modal(self):
        self.client.force_login(self.org_admin)
        session = self.client.session
        session["active_organization"] = self.org_a.slug
        session.save()

        modal_response = self.client.get(
            reverse("exams:create_exam") + "?modal=1",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(modal_response.status_code, 200)

        response = self.client.post(
            reverse("exams:create_exam") + "?modal=1",
            {
                "modal": "1",
                "title": "Org Admin Created Exam",
                "description": "Created by organization admin.",
                "exam_type": "test",
                "is_active": "on",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)

        created_exam = Exam.objects.get(title="Org Admin Created Exam")
        self.assertEqual(created_exam.organization, self.org_a)
        self.assertEqual(created_exam.author, self.org_admin)

    def test_create_exam_redirects_to_org_selector_without_active_org_when_multiple_orgs(self):
        _assign_user_to_org(self.teacher, self.org_b, ProfileRole.TEACHER)
        session = self.client.session
        session.pop("active_organization", None)
        session.save()

        response = self.client.get(reverse("exams:create_exam"))

        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.url.startswith(reverse("organizations:select")))
        self.assertIn("next=%2Fexams%2Fcreate%2F", response.url)

    def test_superadmin_with_profile_org_can_create_exam_when_session_org_missing(self):
        superadmin = User.objects.create_superuser(
            username="exam_superadmin_restore",
            email="exam_superadmin_restore@example.com",
            password="StrongPass123!",
        )
        profile = superadmin.profile
        profile.organization = self.org_a
        profile.organization_type = self.org_a.org_type
        profile.role = ProfileRole.SUPERADMIN
        profile.save(update_fields=["organization", "organization_type", "role", "updated_at"])

        self.client.force_login(superadmin)
        session = self.client.session
        session.pop("active_organization", None)
        session.save()

        response = self.client.post(
            reverse("exams:create_exam") + "?modal=1",
            {
                "modal": "1",
                "title": "Superadmin Restored Exam",
                "description": "Created after restoring organization from profile.",
                "exam_type": "test",
                "is_active": "on",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)

        created_exam = Exam.objects.get(title="Superadmin Restored Exam")
        self.assertEqual(created_exam.organization, self.org_a)
        self.assertEqual(created_exam.author, superadmin)
        self.assertEqual(self.client.session.get("active_organization"), self.org_a.slug)

    def test_superadmin_without_profile_org_can_choose_organization_in_create_exam_modal(self):
        superadmin = User.objects.create_superuser(
            username="exam_superadmin_modal",
            email="exam_superadmin_modal@example.com",
            password="StrongPass123!",
        )

        self.client.force_login(superadmin)
        session = self.client.session
        session.pop("active_organization", None)
        session.save()

        modal_response = self.client.get(
            reverse("exams:create_exam") + "?modal=1",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(modal_response.status_code, 200)
        self.assertContains(modal_response, 'name="organization"', html=False)
        self.assertContains(modal_response, self.org_a.name)
        self.assertContains(modal_response, self.org_b.name)

        response = self.client.post(
            reverse("exams:create_exam") + "?modal=1",
            {
                "modal": "1",
                "organization": str(self.org_b.pk),
                "title": "Superadmin Selected Org Exam",
                "description": "Created by explicit organization selection.",
                "exam_type": "test",
                "is_active": "on",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)

        created_exam = Exam.objects.get(title="Superadmin Selected Org Exam")
        self.assertEqual(created_exam.organization, self.org_b)
        self.assertEqual(created_exam.author, superadmin)
        self.assertEqual(self.client.session.get("active_organization"), self.org_b.slug)

    def test_modal_add_question_returns_partial_markup(self):
        response = self.client.get(
            reverse("exams:add_exam_question", args=[self.exam_visible.slug]),
            {"modal": "1"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="modal" value="1"')
        self.assertContains(response, 'name="text"')
        self.assertContains(response, 'name="option1_text"')
        self.assertContains(response, 'name="option2_text"')
        self.assertNotContains(response, 'name="option3_text"')
        self.assertContains(response, "Variant əlavə et")

    def test_ru_question_add_translations_use_add_not_delete_or_topic(self):
        with override("ru"):
            response = self.client.get(
                reverse("exams:add_exam_question", args=[self.exam_visible.slug]),
                {"modal": "1"},
                HTTP_X_REQUESTED_WITH="XMLHttpRequest",
                HTTP_ACCEPT_LANGUAGE="ru",
            )

            bank_response = self.client.get(
                reverse("exams:teacher_questions_bank", args=[self.exam_visible.slug]),
                HTTP_ACCEPT_LANGUAGE="ru",
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Вы добавляете новый вопрос.")
        self.assertContains(response, "Добавить вариант")
        self.assertNotContains(response, "Добавить новую тему")
        self.assertNotContains(response, "Yeni sual")

        self.assertEqual(bank_response.status_code, 200)
        self.assertContains(bank_response, 'questionCreateTitle: "Добавить новый вопрос"', html=False)
        self.assertNotContains(bank_response, 'questionCreateTitle: "Yeni sual elave et"', html=False)

    def test_modal_add_question_accepts_more_than_four_options(self):
        response = self.client.post(
            reverse("exams:add_exam_question", args=[self.exam_visible.slug]) + "?modal=1",
            {
                "modal": "1",
                "text": "Hansılar proqramlaşdırma dilləridir?",
                "answer_mode": "multiple",
                "time_limit_seconds": "60",
                "option1_text": "Python",
                "option1_is_correct": "on",
                "option2_text": "JavaScript",
                "option2_is_correct": "on",
                "option3_text": "Brauzer",
                "option4_text": "Kompilyator",
                "option5_text": "Go",
                "option5_is_correct": "on",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)

        created_question = self.exam_visible.questions.exclude(id=self.exam_question.id).latest("id")
        self.assertEqual(created_question.options.count(), 5)
        self.assertEqual(created_question.options.filter(is_correct=True).count(), 3)

    def test_modal_edit_question_updates_question_with_json_success(self):
        response = self.client.post(
            reverse("exams:edit_exam_question", args=[self.exam_visible.slug, self.exam_question.id]) + "?modal=1",
            {
                "modal": "1",
                "text": "Python nədir?",
                "answer_mode": "single",
                "time_limit_seconds": "45",
                "option1_text": "Proqramlaşdırma dili",
                "option1_is_correct": "on",
                "option2_text": "Brauzer",
                "option3_text": "",
                "option4_text": "",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)

        self.exam_question.refresh_from_db()
        self.assertEqual(self.exam_question.text, "Python nədir?")
        self.assertEqual(self.exam_question.time_limit_seconds, 45)
        self.assertEqual(self.exam_question.options.filter(is_correct=True).count(), 1)

    def test_modal_edit_question_can_reduce_option_count_to_two(self):
        ExamQuestionOption.objects.create(question=self.exam_question, text="Verilənlər bazası", is_correct=False)
        ExamQuestionOption.objects.create(question=self.exam_question, text="Əməliyyat sistemi", is_correct=False)

        response = self.client.post(
            reverse("exams:edit_exam_question", args=[self.exam_visible.slug, self.exam_question.id]) + "?modal=1",
            {
                "modal": "1",
                "text": "Python nədir?",
                "answer_mode": "single",
                "time_limit_seconds": "45",
                "option1_text": "Proqramlaşdırma dili",
                "option1_is_correct": "on",
                "option2_text": "Brauzer",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)

        self.exam_question.refresh_from_db()
        self.assertEqual(self.exam_question.options.count(), 2)

    def test_modal_edit_question_rejects_multiple_correct_options_in_single_mode(self):
        response = self.client.post(
            reverse("exams:edit_exam_question", args=[self.exam_visible.slug, self.exam_question.id]) + "?modal=1",
            {
                "modal": "1",
                "text": "Python nədir?",
                "answer_mode": "single",
                "time_limit_seconds": "45",
                "option1_text": "Proqramlaşdırma dili",
                "option1_is_correct": "on",
                "option2_text": "Brauzer",
                "option2_is_correct": "on",
                "option3_text": "",
                "option4_text": "",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["success"], False)
        self.assertIn(
            "Cavab rejimini tək seçimdən çoxlu seçimə dəyişmək lazımdır.",
            response.json()["html"],
        )

    def test_other_teacher_cannot_edit_or_delete_my_exam(self):
        self.client.force_login(self.other_teacher)
        session = self.client.session
        session["active_organization"] = self.org_a.slug
        session.save()

        edit_response = self.client.get(reverse("exams:edit_exam", args=[self.exam_visible.slug]))
        delete_response = self.client.post(reverse("exams:delete_exam", args=[self.exam_visible.slug]))

        self.assertEqual(edit_response.status_code, 403)
        self.assertEqual(delete_response.status_code, 403)
        self.assertTrue(Exam.objects.filter(id=self.exam_visible.id).exists())

    def test_student_cannot_delete_exam(self):
        self.client.force_login(self.student)
        session = self.client.session
        session["active_organization"] = self.org_a.slug
        session.save()

        response = self.client.post(reverse("exams:delete_exam", args=[self.exam_visible.slug]))
        self.assertEqual(response.status_code, 403)
        self.assertTrue(Exam.objects.filter(id=self.exam_visible.id).exists())

    def test_delete_exam_redirects_to_profile_my_exams_section(self):
        response = self.client.post(reverse("exams:delete_exam", args=[self.exam_visible.slug]))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, f"{reverse('accounts:profile')}?section=my-exams")
        # Soft delete: sətir qalır, yalnız işarələnir.
        self.exam_visible.refresh_from_db()
        self.assertTrue(self.exam_visible.is_deleted)
        self.assertIsNotNone(self.exam_visible.deleted_at)

    def test_delete_exam_soft_deletes_and_preserves_attempts(self):
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.exam_visible,
            status="submitted",
        )

        response = self.client.post(reverse("exams:delete_exam", args=[self.exam_visible.slug]))

        self.assertEqual(response.status_code, 302)
        # İmtahan sətri qorunur (yumşaq silmə) və silinmiş kimi işarələnir.
        self.exam_visible.refresh_from_db()
        self.assertTrue(self.exam_visible.is_deleted)
        self.assertFalse(self.exam_visible.is_active)
        # Nəticələr (attempts) qorunur — silinmə CASCADE etmir.
        self.assertTrue(ExamAttempt.objects.filter(id=attempt.id).exists())

    def test_soft_deleted_exam_is_hidden_from_teacher_edit_lookup(self):
        self.client.post(reverse("exams:delete_exam", args=[self.exam_visible.slug]))
        # Silinmiş imtahan redaktə lookup-ından çıxarılır (404).
        response = self.client.get(reverse("exams:edit_exam", args=[self.exam_visible.slug]))
        self.assertEqual(response.status_code, 404)

    def test_deleted_exams_list_lists_soft_deleted_exam(self):
        self.client.post(reverse("exams:delete_exam", args=[self.exam_visible.slug]))
        response = self.client.get(reverse("exams:deleted_exams_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.exam_visible.title)

    def test_restore_exam_brings_it_back(self):
        self.client.post(reverse("exams:delete_exam", args=[self.exam_visible.slug]))
        response = self.client.post(reverse("exams:restore_exam", args=[self.exam_visible.slug]))
        self.assertEqual(response.status_code, 302)
        self.exam_visible.refresh_from_db()
        self.assertFalse(self.exam_visible.is_deleted)
        self.assertIsNone(self.exam_visible.deleted_at)

    def test_permanent_delete_removes_exam_and_attempts(self):
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.exam_visible,
            status="submitted",
        )
        # Yalnız yumşaq silinmiş imtahanı birdəfəlik silmək olar.
        self.client.post(reverse("exams:delete_exam", args=[self.exam_visible.slug]))
        response = self.client.post(reverse("exams:permanent_delete_exam", args=[self.exam_visible.slug]))
        self.assertEqual(response.status_code, 302)
        self.assertFalse(Exam.objects.filter(id=self.exam_visible.id).exists())
        self.assertFalse(ExamAttempt.objects.filter(id=attempt.id).exists())

    def test_permanent_delete_rejects_non_deleted_exam(self):
        # Yumşaq silinməmiş imtahan üçün birdəfəlik silmə mövcud deyil (404).
        response = self.client.post(reverse("exams:permanent_delete_exam", args=[self.exam_visible.slug]))
        self.assertEqual(response.status_code, 404)
        self.assertTrue(Exam.objects.filter(id=self.exam_visible.id).exists())

    def test_view_results_of_deleted_exam_is_readonly(self):
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.exam_visible,
            status="submitted",
        )
        self.client.post(reverse("exams:delete_exam", args=[self.exam_visible.slug]))

        # Silinmiş imtahanın nəticələrinə "Zibil qutusu"ndan baxmaq olar.
        response = self.client.get(reverse("exams:teacher_exam_results", args=[self.exam_visible.slug]))
        self.assertEqual(response.status_code, 200)
        # Yalnız-oxu banneri "Zibil qutusu"na keçid göstərir.
        self.assertContains(response, reverse("exams:deleted_exams_list"))

        # Qiymətləndirmə POST-u bloklanır (redirect).
        post = self.client.post(
            reverse("exams:teacher_exam_results", args=[self.exam_visible.slug]),
            {"attempt_id": attempt.id, "teacher_score": "50"},
        )
        self.assertEqual(post.status_code, 302)
        attempt.refresh_from_db()
        self.assertIsNone(attempt.teacher_score)

    def test_edit_other_tenant_exam_is_not_found(self):
        response = self.client.get(reverse("exams:edit_exam", args=[self.exam_other_tenant.slug]))
        self.assertEqual(response.status_code, 404)

    def test_edit_exam_full_page_redirects_to_profile_my_exams(self):
        response = self.client.get(reverse("exams:edit_exam", args=[self.exam_visible.slug]))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, f"{reverse('accounts:profile')}?section=my-exams")

    def test_modal_edit_exam_updates_random_question_count(self):
        response = self.client.post(
            reverse("exams:edit_exam", args=[self.exam_visible.slug]) + "?modal=1",
            {
                "modal": "1",
                "title": self.exam_visible.title,
                "description": "Updated random draw count.",
                "exam_type": "test",
                "is_active": "on",
                "is_public": "on",
                "random_question_count": "25",
                "fair_question_distribution_enabled": "false",
                "ai_difficulty_balance_enabled": "false",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)
        self.exam_visible.refresh_from_db()
        self.assertEqual(self.exam_visible.random_question_count, 25)
        self.assertFalse(self.exam_visible.fair_question_distribution_enabled)
        self.assertFalse(self.exam_visible.ai_difficulty_balance_enabled)

    def test_teacher_exam_detail_defaults_to_generic_back_with_profile_fallback(self):
        response = self.client.get(reverse("exams:teacher_exam_detail", args=[self.exam_visible.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["profile_return_url"], f"{reverse('accounts:profile')}?section=my-exams")
        self.assertContains(response, "Geri")
        self.assertNotContains(response, "Profilə Qayıt")

    def test_teacher_exam_detail_uses_explicit_course_dashboard_return_url(self):
        return_to = reverse("courses:course_dashboard", args=[self.course.id])
        response = self.client.get(
            reverse("exams:teacher_exam_detail", args=[self.exam_visible.slug]),
            {"return_to": return_to},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["profile_return_url"], return_to)
        self.assertContains(response, "Geri")
        self.assertNotContains(response, "Profilə Qayıt")

    def test_teacher_exam_detail_questions_bank_link_preserves_return_to(self):
        return_to = reverse("courses:course_dashboard", args=[self.course.id])
        response = self.client.get(
            reverse("exams:teacher_exam_detail", args=[self.exam_visible.slug]),
            {"from_section": "my-courses", "return_to": return_to},
        )

        expected_query = urlencode({"from_section": "my-courses", "return_to": return_to})
        expected_href = f'{reverse("exams:teacher_questions_bank", args=[self.exam_visible.slug])}?{expected_query}'

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, expected_href.replace("&", "&amp;"), html=False)

    def test_teacher_exam_detail_bulk_add_link_preserves_return_to(self):
        return_to = reverse("courses:course_dashboard", args=[self.course.id])
        response = self.client.get(
            reverse("exams:teacher_exam_detail", args=[self.exam_visible.slug]),
            {"from_section": "my-courses", "return_to": return_to},
        )

        expected_query = urlencode({"from_section": "my-courses", "return_to": return_to})
        expected_href = f'{reverse("exams:test_question_bank", args=[self.exam_visible.slug])}?{expected_query}'

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, expected_href.replace("&", "&amp;"), html=False)

    def test_teacher_exam_detail_live_results_link_preserves_return_to(self):
        return_to = reverse("courses:course_dashboard", args=[self.course.id])
        response = self.client.get(
            reverse("exams:teacher_exam_detail", args=[self.exam_visible.slug]),
            {"from_section": "my-courses", "return_to": return_to},
        )

        expected_query = urlencode({"from_section": "my-courses", "return_to": return_to})
        expected_href = f'{reverse("liveExam:teacher_live_results", args=[self.exam_visible.slug])}?{expected_query}'

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, expected_href.replace("&", "&amp;"), html=False)

    def test_teacher_exam_detail_includes_archive_toggle(self):
        response = self.client.get(reverse("exams:teacher_exam_detail", args=[self.exam_visible.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Arxivlə")
        self.assertContains(response, reverse("exams:toggle_exam_archive", args=[self.exam_visible.slug]))

        self.exam_visible.is_archived = True
        self.exam_visible.save(update_fields=["is_archived"])

        response = self.client.get(reverse("exams:teacher_exam_detail", args=[self.exam_visible.slug]))
        self.assertContains(response, "Arxivdən çıxar")

    def test_teacher_can_archive_from_detail_and_stay_on_detail(self):
        detail_url = reverse("exams:teacher_exam_detail", args=[self.exam_visible.slug])
        response = self.client.post(
            reverse("exams:toggle_exam_archive", args=[self.exam_visible.slug]),
            {"next": detail_url},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, detail_url)
        self.exam_visible.refresh_from_db()
        self.assertTrue(self.exam_visible.is_archived)

    def test_teacher_exam_detail_initially_renders_first_question_batch(self):
        for order in range(2, 26):
            ExamQuestion.objects.create(
                exam=self.exam_visible,
                text=f"Lazy detail question {order}",
                order=order,
                points=1,
            )

        response = self.client.get(reverse("exams:teacher_exam_detail", args=[self.exam_visible.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["questions"]), 20)
        self.assertContains(response, "Lazy detail question 20")
        self.assertNotContains(response, "Lazy detail question 21")
        self.assertContains(
            response,
            reverse("exams:teacher_exam_detail_questions_page", args=[self.exam_visible.slug]),
        )
        self.assertContains(response, 'data-next-offset="20"', html=False)

    def test_teacher_exam_detail_questions_page_returns_next_batch(self):
        for order in range(2, 26):
            ExamQuestion.objects.create(
                exam=self.exam_visible,
                text=f"Lazy detail question {order}",
                order=order,
                points=1,
            )

        response = self.client.get(
            reverse("exams:teacher_exam_detail_questions_page", args=[self.exam_visible.slug]),
            {"offset": 20, "limit": 20},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertFalse(payload["has_more"])
        self.assertEqual(payload["next_offset"], 25)
        self.assertIn("Lazy detail question 21", payload["html"])
        self.assertIn("Lazy detail question 25", payload["html"])
        self.assertNotIn("Lazy detail question 20", payload["html"])

    def test_teacher_can_toggle_exam_results_visibility(self):
        response = self.client.post(
            reverse("exams:toggle_exam_results_visibility", args=[self.exam_visible.slug]),
            {
                "from_section": "my-exams",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.exam_visible.refresh_from_db()
        self.assertTrue(self.exam_visible.results_hidden_from_students)

        detail_response = self.client.get(reverse("exams:teacher_exam_detail", args=[self.exam_visible.slug]))
        self.assertContains(detail_response, "Nəticələri tələbələrə göstər")
        self.assertContains(detail_response, "Tələbələrdən gizlidir")

        self.client.post(reverse("exams:toggle_exam_results_visibility", args=[self.exam_visible.slug]))
        self.exam_visible.refresh_from_db()
        self.assertFalse(self.exam_visible.results_hidden_from_students)

    def test_hidden_exam_results_are_not_visible_to_student(self):
        self.exam_visible.results_hidden_from_students = True
        self.exam_visible.save(update_fields=["results_hidden_from_students"])
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.exam_visible,
            status="submitted",
            attempt_number=1,
        )

        _login_with_org(self.client, self.student, self.org_a)

        result_response = self.client.get(reverse("exams:exam_result", args=[self.exam_visible.slug, attempt.id]))
        self.assertEqual(result_response.status_code, 302)
        self.assertIn("section=my-results", result_response.url)

        history_response = self.client.get(
            reverse("exams:student_exam_history"),
            {"exam": self.exam_visible.slug},
        )
        self.assertEqual(history_response.status_code, 200)
        self.assertContains(history_response, "Gizlidir")
        self.assertNotContains(history_response, "Bax")

        profile_response = self.client.get(
            reverse("accounts:profile"),
            {"section": "my-results", "results_type": "exams"},
        )
        self.assertEqual(profile_response.status_code, 200)
        self.assertEqual(profile_response.context["my_result_counts"]["exams"], 0)

    def test_teacher_exam_detail_disables_live_start_when_exam_is_passive(self):
        self.exam_visible.is_active = False
        self.exam_visible.save(update_fields=["is_active"])

        response = self.client.get(reverse("exams:teacher_exam_detail", args=[self.exam_visible.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'aria-disabled="true"', html=False)
        self.assertContains(response, "Öncə imtahanı aktiv edin. İmtahan hazırda passivdir.")
        self.assertNotContains(
            response,
            reverse("liveExam:create_session_slug", kwargs={"slug": self.exam_visible.slug}),
        )

    def test_delete_exam_question_resequences_remaining_question_orders(self):
        second_question = ExamQuestion.objects.create(
            exam=self.exam_visible,
            text="Second question",
            order=2,
            answer_mode="single",
        )
        third_question = ExamQuestion.objects.create(
            exam=self.exam_visible,
            text="Third question",
            order=3,
            answer_mode="single",
        )

        response = self.client.post(
            reverse("exams:delete_exam_question", args=[self.exam_visible.slug, self.exam_question.id])
        )

        self.assertEqual(response.status_code, 302)
        remaining_orders = list(
            ExamQuestion.objects.filter(id__in=[second_question.id, third_question.id])
            .order_by("order", "id")
            .values_list("order", flat=True)
        )
        self.assertEqual(remaining_orders, [1, 2])

    def test_teacher_exam_detail_falls_back_to_safe_referer_when_return_to_missing(self):
        referer = reverse("exams:teacher_exam_list")
        response = self.client.get(
            reverse("exams:teacher_exam_detail", args=[self.exam_visible.slug]),
            HTTP_REFERER=referer,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["profile_return_url"], referer)
        self.assertContains(response, "Geri")
        self.assertNotContains(response, "Profilə Qayıt")

    def test_teacher_exam_detail_ignores_internal_question_bank_referer(self):
        referer = reverse("exams:teacher_questions_bank", args=[self.exam_visible.slug])
        response = self.client.get(
            reverse("exams:teacher_exam_detail", args=[self.exam_visible.slug]),
            HTTP_REFERER=referer,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["profile_return_url"], f"{reverse('accounts:profile')}?section=my-exams")
        self.assertNotEqual(response.context["profile_return_url"], referer)

    def test_teacher_exam_results_keeps_generic_source_back_label(self):
        return_to = reverse("courses:course_dashboard", args=[self.course.id])
        response = self.client.get(
            reverse("exams:teacher_exam_results", args=[self.exam_visible.slug]),
            {"return_to": return_to},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["profile_return_url"], return_to)
        self.assertContains(response, "Geri")
        self.assertNotContains(response, "Profilə Qayıt")
        self.assertNotContains(response, "Profile geri dön")
        self.assertContains(response, "results-filter-card")
        self.assertContains(response, "resultsFilterSearchInput")
        self.assertContains(response, "css/pagination.css")

    def test_teacher_exam_results_renders_bulk_delete_controls(self):
        ExamAttempt.objects.create(
            user=self.student,
            exam=self.exam_visible,
            status="submitted",
        )

        response = self.client.get(reverse("exams:teacher_exam_results", args=[self.exam_visible.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "selectedAttemptCount")
        self.assertContains(response, "deleteSelectedAttemptsBtn")
        self.assertContains(response, "js-attempt-checkbox")

    def test_delete_exam_attempts_removes_selected_attempts(self):
        first_attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.exam_visible,
            status="submitted",
        )
        second_attempt = ExamAttempt.objects.create(
            user=self.teacher,
            exam=self.exam_visible,
            status="draft",
        )

        response = self.client.post(
            reverse("exams:delete_exam_attempts", args=[self.exam_visible.slug]),
            {
                "attempt_ids": [str(first_attempt.id), str(second_attempt.id)],
                "next": reverse("exams:teacher_exam_results", args=[self.exam_visible.slug]),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(ExamAttempt.objects.filter(id=first_attempt.id).exists())
        self.assertFalse(ExamAttempt.objects.filter(id=second_attempt.id).exists())

    def test_teacher_exam_results_reveals_student_name_for_test_attempts(self):
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.exam_visible,
            status="submitted",
        )

        response = self.client.get(reverse("exams:teacher_exam_results", args=[self.exam_visible.slug]))

        self.assertEqual(response.status_code, 200)
        attempts_data = response.context["attempts_data"]
        matching_item = next(item for item in attempts_data if item["attempt"].id == attempt.id)
        self.assertTrue(matching_item["can_view_name"])
        self.assertEqual(matching_item["real_name"], self.student.username)
        self.assertEqual(matching_item["action_label"], "Bax")
        self.assertContains(response, self.student.username)
        self.assertNotContains(response, "<th>Nəticə (%)</th>", html=True)

    def test_test_exam_results_use_delivered_question_set_on_teacher_and_student_pages(self):
        correct_option = self.exam_question.options.filter(is_correct=True).first()
        wrong_question = ExamQuestion.objects.create(
            exam=self.exam_visible,
            text="Delivered wrong question",
            order=2,
            answer_mode="single",
        )
        ExamQuestionOption.objects.create(question=wrong_question, text="Correct", is_correct=True)
        wrong_option = ExamQuestionOption.objects.create(question=wrong_question, text="Wrong", is_correct=False)
        unanswered_question = ExamQuestion.objects.create(
            exam=self.exam_visible,
            text="Delivered unanswered question",
            order=3,
            answer_mode="single",
        )
        ExamQuestionOption.objects.create(question=unanswered_question, text="Correct", is_correct=True)
        ExamQuestionOption.objects.create(question=unanswered_question, text="Wrong", is_correct=False)
        for idx in range(4, 24):
            ExamQuestion.objects.create(
                exam=self.exam_visible,
                text=f"Bank only question {idx}",
                order=idx,
                answer_mode="single",
            )

        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.exam_visible,
            status="submitted",
        )
        correct_answer = ExamAnswer.objects.create(attempt=attempt, question=self.exam_question, is_correct=True)
        correct_answer.selected_options.add(correct_option)
        wrong_answer = ExamAnswer.objects.create(attempt=attempt, question=wrong_question, is_correct=False)
        wrong_answer.selected_options.add(wrong_option)
        ExamAnswer.objects.create(attempt=attempt, question=unanswered_question, is_correct=False)
        attempt.recalculate_score()

        teacher_results = self.client.get(reverse("exams:teacher_exam_results", args=[self.exam_visible.slug]))
        self.assertEqual(teacher_results.status_code, 200)
        result_item = next(
            item for item in teacher_results.context["attempts_data"] if item["attempt"].id == attempt.id
        )
        self.assertEqual(self.exam_visible.questions.count(), 23)
        self.assertEqual(result_item["test_result"].delivered_count, 3)
        self.assertEqual(result_item["test_result"].correct_count, 1)
        self.assertEqual(result_item["test_result"].wrong_count, 1)
        self.assertEqual(result_item["test_result"].unanswered_count, 1)
        self.assertEqual(result_item["test_result"].percentage_display, "33.3")

        teacher_detail = self.client.get(
            reverse("exams:teacher_view_attempt", args=[self.exam_visible.slug, attempt.id])
        )
        self.assertEqual(teacher_detail.status_code, 200)
        self.assertEqual(teacher_detail.context["test_result"].delivered_count, 3)
        self.assertContains(teacher_detail, "3 verilmiş sual")

        _login_with_org(self.client, self.student, self.org_a)
        student_result = self.client.get(reverse("exams:exam_result", args=[self.exam_visible.slug, attempt.id]))
        self.assertEqual(student_result.status_code, 200)
        self.assertEqual(student_result.context["test_result"].delivered_count, 3)
        self.assertEqual(student_result.context["test_result"].unanswered_count, 1)
        self.assertContains(student_result, "Verilmiş sual")

    def test_test_exam_result_falls_back_to_legacy_counts_without_rebuilding_finished_attempt(self):
        for idx in range(2, 12):
            ExamQuestion.objects.create(
                exam=self.exam_visible,
                text=f"Legacy bank question {idx}",
                order=idx,
                answer_mode="single",
            )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.exam_visible,
            status="submitted",
            correct_count=8,
            wrong_count=2,
        )

        teacher_detail = self.client.get(
            reverse("exams:teacher_view_attempt", args=[self.exam_visible.slug, attempt.id])
        )

        self.assertEqual(teacher_detail.status_code, 200)
        self.assertEqual(teacher_detail.context["test_result"].delivered_count, 10)
        self.assertTrue(teacher_detail.context["test_result"].used_legacy_fallback)
        self.assertFalse(attempt.answers.exists())

    def test_teacher_exam_results_keeps_written_student_name_hidden_until_review_is_completed(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Pending Written Exam",
            exam_type="written",
            is_active=True,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
            finished_at=timezone.now(),
        )

        response = self.client.get(reverse("exams:teacher_exam_results", args=[written_exam.slug]))

        self.assertEqual(response.status_code, 200)
        attempts_data = response.context["attempts_data"]
        matching_item = next(item for item in attempts_data if item["attempt"].id == attempt.id)
        self.assertFalse(matching_item["can_view_name"])
        self.assertEqual(matching_item["action_label"], "Yoxla")
        self.assertEqual(matching_item["countdown_seconds"], 0)
        self.assertContains(response, "Yoxla")
        self.assertContains(response, "Anonim görünüş")
        self.assertContains(response, "Bal")
        self.assertContains(response, "<strong>0</strong>", html=True)
        self.assertNotContains(response, "<strong>0%</strong>", html=True)

        attempt.finished_at = timezone.now() - timedelta(minutes=6)
        attempt.save(update_fields=["finished_at"])

        still_hidden_response = self.client.get(reverse("exams:teacher_exam_results", args=[written_exam.slug]))
        still_hidden_item = next(
            item for item in still_hidden_response.context["attempts_data"] if item["attempt"].id == attempt.id
        )
        self.assertFalse(still_hidden_item["can_view_name"])
        self.assertEqual(still_hidden_item["action_label"], "Yoxla")
        self.assertContains(still_hidden_response, "Anonim görünüş")
        self.assertNotContains(still_hidden_response, self.student.username)

    def test_teacher_exam_results_shows_recheck_then_view_for_written_attempts(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Recheck Written Exam",
            exam_type="written",
            is_active=True,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
            checked_by_teacher=True,
            teacher_score=74,
            teacher_checked_at=timezone.now(),
        )

        response = self.client.get(reverse("exams:teacher_exam_results", args=[written_exam.slug]))

        self.assertEqual(response.status_code, 200)
        attempts_data = response.context["attempts_data"]
        matching_item = next(item for item in attempts_data if item["attempt"].id == attempt.id)
        self.assertFalse(matching_item["can_view_name"])
        self.assertEqual(matching_item["action_label"], "Yenidən yoxla")
        self.assertContains(response, "Yenidən yoxla")

        attempt.teacher_checked_at = timezone.now() - timedelta(minutes=6)
        attempt.save(update_fields=["teacher_checked_at"])

        locked_response = self.client.get(reverse("exams:teacher_exam_results", args=[written_exam.slug]))
        locked_item = next(
            item for item in locked_response.context["attempts_data"] if item["attempt"].id == attempt.id
        )
        self.assertTrue(locked_item["can_view_name"])
        self.assertEqual(locked_item["action_label"], "Bax")
        self.assertContains(locked_response, "Bax")
        self.assertContains(locked_response, "Bal")
        self.assertContains(locked_response, "<strong>74</strong>", html=True)
        self.assertNotContains(locked_response, "<strong>74%</strong>", html=True)

    def test_teacher_exam_results_reveals_student_name_when_written_grade_is_visible_without_timestamp(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Written Visibility Exam",
            exam_type="written",
            is_active=True,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
            checked_by_teacher=True,
            teacher_score=74,
            teacher_checked_at=None,
        )

        response = self.client.get(reverse("exams:teacher_exam_results", args=[written_exam.slug]))

        self.assertEqual(response.status_code, 200)
        attempts_data = response.context["attempts_data"]
        matching_item = next(item for item in attempts_data if item["attempt"].id == attempt.id)
        self.assertTrue(matching_item["can_view_name"])
        self.assertEqual(matching_item["seconds_remaining"], 0)
        self.assertEqual(matching_item["real_name"], self.student.username)

    def test_teacher_exam_results_reveals_pending_written_student_name_when_org_override_enabled(self):
        self.org_a.set_written_exam_identity_reveal_enabled(True)
        self.org_a.save(update_fields=["settings", "updated_at"])
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Org Override Pending Written Exam",
            exam_type="written",
            is_active=True,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
            finished_at=timezone.now(),
        )

        response = self.client.get(reverse("exams:teacher_exam_results", args=[written_exam.slug]))

        self.assertEqual(response.status_code, 200)
        attempts_data = response.context["attempts_data"]
        matching_item = next(item for item in attempts_data if item["attempt"].id == attempt.id)
        self.assertTrue(matching_item["can_view_name"])
        self.assertEqual(matching_item["action_label"], "Yoxla")
        self.assertEqual(matching_item["countdown_seconds"], 0)
        self.assertContains(response, self.student.username)
        self.assertNotContains(response, "Anonim görünüş")

    def test_teacher_exam_results_keeps_recheck_window_when_org_override_enabled(self):
        self.org_a.set_written_exam_identity_reveal_enabled(True)
        self.org_a.save(update_fields=["settings", "updated_at"])
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Org Override Recheck Written Exam",
            exam_type="written",
            is_active=True,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
            checked_by_teacher=True,
            teacher_score=74,
            teacher_checked_at=timezone.now(),
        )

        response = self.client.get(reverse("exams:teacher_exam_results", args=[written_exam.slug]))

        self.assertEqual(response.status_code, 200)
        attempts_data = response.context["attempts_data"]
        matching_item = next(item for item in attempts_data if item["attempt"].id == attempt.id)
        self.assertTrue(matching_item["can_view_name"])
        self.assertEqual(matching_item["action_label"], "Yenidən yoxla")
        self.assertGreater(matching_item["countdown_seconds"], 0)
        self.assertContains(response, self.student.username)
        self.assertContains(response, "Yenidən yoxla")

    def test_teacher_view_attempt_keeps_generic_source_back_label(self):
        attempt = ExamAttempt.objects.create(
            user=self.teacher,
            exam=self.exam_visible,
            status="submitted",
        )
        return_to = reverse("courses:course_dashboard", args=[self.course.id])
        response = self.client.get(
            reverse("exams:teacher_view_attempt", args=[self.exam_visible.slug, attempt.id]),
            {"return_to": return_to},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["profile_return_url"], return_to)
        self.assertContains(response, "Geri")
        self.assertNotContains(response, "Profilə Qayıt")
        self.assertNotContains(response, "Profile geri dön")

    def test_teacher_view_attempt_keeps_written_student_name_anonymous_before_review(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Anonymous View Written Exam",
            exam_type="written",
            is_active=True,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
            finished_at=timezone.now() - timedelta(minutes=20),
        )

        response = self.client.get(reverse("exams:teacher_view_attempt", args=[written_exam.slug, attempt.id]))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["can_view_student_identity"])
        self.assertContains(response, response.context["student_display"])
        self.assertIn("#", response.context["student_display"])
        self.assertNotContains(response, self.student.username)
        self.assertNotContains(response, "Qalan vaxt:")

    def test_teacher_view_attempt_shows_finished_at_and_computed_duration(self):
        self.teacher.first_name = "View"
        self.teacher.last_name = "Teacher"
        self.teacher.save(update_fields=["first_name", "last_name"])
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Written Timing View Exam",
            exam_type="written",
            is_active=True,
        )
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Explain the timing",
            order=1,
            answer_mode="single",
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
        )
        ExamAnswer.objects.create(
            attempt=attempt,
            question=question,
            text_answer="Timing answer",
        )

        started_at = timezone.now() - timedelta(minutes=42)
        finished_at = started_at + timedelta(minutes=35)
        attempt.started_at = started_at
        attempt.finished_at = finished_at
        attempt.duration_seconds = None
        attempt.save(update_fields=["started_at", "finished_at", "duration_seconds"])

        response = self.client.get(reverse("exams:teacher_view_attempt", args=[written_exam.slug, attempt.id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["attempt_timing"]["finished_at"], finished_at)
        self.assertEqual(response.context["attempt_timing"]["duration_seconds"], 35 * 60)
        self.assertEqual(response.context["exam_evaluator_display"], "View Teacher")
        self.assertContains(response, "Yoxlayan müəllim:")
        self.assertContains(response, "View Teacher")
        self.assertContains(response, "Bitirib:")
        self.assertContains(response, "Ümumi müddət:")
        self.assertContains(response, "35 dəq")

    def test_teacher_check_attempt_includes_confirm_modal_and_integer_score_input(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Confirm Written Exam",
            exam_type="written",
            is_active=True,
        )
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Explain the solution",
            order=1,
            answer_mode="single",
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
        )
        ExamAnswer.objects.create(
            attempt=attempt,
            question=question,
            text_answer="Written answer",
            teacher_score=7,
            teacher_feedback="Saved score",
        )

        response = self.client.get(reverse("exams:teacher_check_attempt", args=[written_exam.slug, attempt.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "courseActionConfirmModal")
        self.assertContains(response, 'id="modalScoreInput"')
        self.assertContains(response, 'id="modalMaxPointsInput"')
        # EXAM-P0-04: max bal artıq POST edilmir — input display-only-dur.
        self.assertNotContains(response, f'name="max_points_{question.id}"')
        self.assertContains(response, f'id="hidden_max_points_{question.id}"')
        self.assertContains(response, 'step="1"')

    def test_teacher_check_attempt_shows_exam_timing_summary(self):
        self.teacher.first_name = "Teacher"
        self.teacher.last_name = "Owner"
        self.teacher.save(update_fields=["first_name", "last_name"])
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Timing Summary Exam",
            exam_type="written",
            is_active=True,
        )
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Explain the solution",
            order=1,
            answer_mode="single",
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
        )
        ExamAnswer.objects.create(
            attempt=attempt,
            question=question,
            text_answer="Written answer",
        )

        started_at = timezone.now() - timedelta(minutes=30)
        finished_at = started_at + timedelta(minutes=27)
        attempt.started_at = started_at
        attempt.finished_at = finished_at
        attempt.duration_seconds = None
        attempt.save(update_fields=["started_at", "finished_at", "duration_seconds"])

        response = self.client.get(reverse("exams:teacher_check_attempt", args=[written_exam.slug, attempt.id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["attempt_timing"]["finished_at"], finished_at)
        self.assertEqual(response.context["attempt_timing"]["duration_seconds"], 27 * 60)
        self.assertContains(response, "Başlayıb")
        self.assertContains(response, "Bitirib")
        self.assertContains(response, "Ümumi müddət")
        self.assertContains(response, "İmtahanı yaradan müəllim:")
        self.assertContains(response, "Teacher Owner")
        self.assertContains(response, "27 dəq")

    def test_teacher_check_attempt_shows_real_student_name_when_org_override_enabled(self):
        self.org_a.set_written_exam_identity_reveal_enabled(True)
        self.org_a.save(update_fields=["settings", "updated_at"])
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Override Identity Check Exam",
            exam_type="written",
            is_active=True,
        )
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Explain the solution",
            order=1,
            answer_mode="single",
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
        )
        ExamAnswer.objects.create(
            attempt=attempt,
            question=question,
            text_answer="Written answer",
        )

        response = self.client.get(reverse("exams:teacher_check_attempt", args=[written_exam.slug, attempt.id]))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["can_view_student_identity"])
        self.assertEqual(response.context["student_display"], self.student.username)
        self.assertContains(response, self.student.username)

    def test_teacher_check_attempt_returns_to_pending_review_when_opened_from_queue(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Pending Review Return Exam",
            exam_type="written",
            is_active=True,
        )
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Explain the solution",
            order=1,
            answer_mode="single",
            points=5,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
        )
        ExamAnswer.objects.create(
            attempt=attempt,
            question=question,
            text_answer="Written answer",
        )
        return_to = f"{reverse('accounts:profile')}?section=pending-review&pr_page=2"

        response = self.client.get(
            reverse("exams:teacher_check_attempt", args=[written_exam.slug, attempt.id]),
            {
                "from_section": "pending-review",
                "return_to": return_to,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["results_return_url"], return_to)
        self.assertContains(response, f'href="{return_to.replace("&", "&amp;")}"', html=False)

        save_response = self.client.post(
            reverse("exams:teacher_check_attempt", args=[written_exam.slug, attempt.id])
            + f"?from_section=pending-review&return_to={quote(return_to)}",
            {
                f"score_{question.id}": "4",
                f"max_points_{question.id}": "5",
                f"feedback_{question.id}": "Queue return flow",
            },
        )

        self.assertRedirects(save_response, return_to, fetch_redirect_response=False)

    def test_teacher_check_attempt_post_ignores_client_max_points(self):
        """EXAM-P0-04: grading POST sual tərifini (points) dəyişə bilməz."""
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Editable Max Written Exam",
            exam_type="written",
            is_active=True,
        )
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Explain the solution",
            order=1,
            answer_mode="single",
            points=5,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
        )
        answer = ExamAnswer.objects.create(
            attempt=attempt,
            question=question,
            text_answer="Written answer",
        )

        response = self.client.post(
            reverse("exams:teacher_check_attempt", args=[written_exam.slug, attempt.id]),
            {
                f"score_{question.id}": "4",
                f"max_points_{question.id}": "100",
                f"feedback_{question.id}": "Updated with new max",
            },
        )

        self.assertEqual(response.status_code, 302)
        question.refresh_from_db()
        answer.refresh_from_db()
        attempt.refresh_from_db()

        # Client-in göndərdiyi max_points sual tərifinə yazılmır.
        self.assertEqual(question.points, 5)
        self.assertEqual(answer.teacher_score, 4)
        self.assertEqual(answer.teacher_feedback, "Updated with new max")
        self.assertEqual(attempt.teacher_score, 4)
        self.assertTrue(attempt.checked_by_teacher)
        self.assertIsNotNone(attempt.teacher_checked_at)

    def test_teacher_check_attempt_post_clamps_score_to_question_max(self):
        """EXAM-P0-04: bal [0, max] aralığına clamp olunur — max artırılmır."""
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Free Score Written Exam",
            exam_type="written",
            is_active=True,
        )
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Clamp scores above max",
            order=1,
            answer_mode="single",
            points=1,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
        )
        answer = ExamAnswer.objects.create(
            attempt=attempt,
            question=question,
            text_answer="Written answer",
        )

        response = self.client.post(
            reverse("exams:teacher_check_attempt", args=[written_exam.slug, attempt.id]),
            {
                f"score_{question.id}": "10",
                f"max_points_{question.id}": "1",
                f"feedback_{question.id}": "Score must be clamped",
            },
        )

        self.assertEqual(response.status_code, 302)
        question.refresh_from_db()
        answer.refresh_from_db()
        attempt.refresh_from_db()

        self.assertEqual(question.points, 1)
        self.assertEqual(answer.teacher_score, 1)
        self.assertEqual(answer.teacher_feedback, "Score must be clamped")
        self.assertEqual(attempt.teacher_score, 1)
        self.assertTrue(attempt.checked_by_teacher)

    def test_teacher_check_attempt_post_uses_snapshot_points_over_live_question(self):
        """EXAM-P0-04 + INTEGRITY-001: clamp sərhədi çatdırılma snapshot-undan gəlir."""
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Snapshot Max Written Exam",
            exam_type="written",
            is_active=True,
        )
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Snapshot bounds grading",
            order=1,
            answer_mode="single",
            points=3,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
        )
        answer = ExamAnswer.objects.create(
            attempt=attempt,
            question=question,
            text_answer="Written answer",
            question_snapshot={"v": 1, "points": 3, "answer_mode": "single", "options": []},
        )

        # Sual sonradan redaktə olunub balı artırılsa da, keçmiş cavabın
        # qiymətləndirmə sərhədi çatdırılma anındakı 3 bal olaraq qalır.
        question.points = 50
        question.save(update_fields=["points"])

        response = self.client.post(
            reverse("exams:teacher_check_attempt", args=[written_exam.slug, attempt.id]),
            {
                f"score_{question.id}": "40",
                f"feedback_{question.id}": "Bounded by snapshot",
            },
        )

        self.assertEqual(response.status_code, 302)
        answer.refresh_from_db()
        attempt.refresh_from_db()

        self.assertEqual(answer.teacher_score, 3)
        self.assertEqual(attempt.teacher_score, 3)
        self.assertIsNotNone(attempt.teacher_checked_at)

    @patch("apps.exams.services.ai_grading.grade_written_answer")
    def test_ai_grade_answer_ignores_posted_max_points(self, mock_grade_written_answer):
        """EXAM-P0-04: AI qiymətləndirmə max balı client body-dən götürmür."""
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="AI Max Written Exam",
            exam_type="written",
            is_active=True,
        )
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Explain the solution",
            order=1,
            answer_mode="single",
            points=1,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
        )
        ExamAnswer.objects.create(
            attempt=attempt,
            question=question,
            text_answer="Written answer",
        )
        mock_grade_written_answer.return_value = {
            "ok": True,
            "score": 5,
            "explanation": "Looks good",
            "cached": False,
        }

        response = self.client.post(
            reverse("exams:ai_grade_answer", args=[written_exam.slug, attempt.id]),
            data='{"question_id": %d, "max_points": 500}' % question.id,
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["score"], 5)
        mock_grade_written_answer.assert_called_once()
        # Client body-dəki max_points ignor olunur; sualdan (points=1) gəlir.
        self.assertEqual(mock_grade_written_answer.call_args.kwargs["max_points"], 1)

    @patch("apps.exams.services.ai_grading.requests.post")
    def test_ai_grade_answer_accepts_image_only_written_submission(self, mock_post):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="AI Image Written Exam",
            exam_type="written",
            is_active=True,
        )
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Read and grade the handwritten answer",
            order=1,
            answer_mode="single",
            points=5,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
        )
        answer = ExamAnswer.objects.create(
            attempt=attempt,
            question=question,
            text_answer="",
        )
        ExamAnswerFile.objects.create(
            answer=answer,
            file=SimpleUploadedFile("answer-image.png", _TINY_PNG_BYTES, content_type="image/png"),
        )

        mock_response = mock_post.return_value
        mock_response.status_code = 200
        mock_response.json.return_value = {
            "candidates": [
                {"content": {"parts": [{"text": "SCORE: 5\nEXPLANATION: The uploaded handwritten answer is correct."}]}}
            ]
        }

        with self.settings(GEMINI_API_KEY="test-gemini-key"):
            response = self.client.post(
                reverse("exams:ai_grade_answer", args=[written_exam.slug, attempt.id]),
                data='{"question_id": %d, "max_points": 5}' % question.id,
                content_type="application/json",
                HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["score"], 5)
        mock_post.assert_called_once()

    def test_teacher_check_attempt_marks_image_only_answer_as_not_empty(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Image Only Review Exam",
            exam_type="written",
            is_active=True,
        )
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Check attached answer",
            order=1,
            answer_mode="single",
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="submitted",
        )
        answer = ExamAnswer.objects.create(
            attempt=attempt,
            question=question,
            text_answer="",
        )
        ExamAnswerFile.objects.create(
            answer=answer,
            file=SimpleUploadedFile("answer-image.png", _TINY_PNG_BYTES, content_type="image/png"),
        )

        response = self.client.get(reverse("exams:teacher_check_attempt", args=[written_exam.slug, attempt.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-has-answer="1"')
        self.assertContains(response, 'data-has-ai-gradable="1"')
        self.assertContains(response, "status-not-checked")
        self.assertNotContains(response, "status-empty")


class StudentExamVisibilityFilteringTest(TestCase):
    def setUp(self):
        self.teacher = User.objects.create_user(
            username="student_exam_teacher",
            email="student_exam_teacher@example.com",
            password="StrongPass123!",
        )
        self.student = User.objects.create_user(
            username="student_exam_student",
            email="student_exam_student@example.com",
            password="StrongPass123!",
        )
        self.viewer_teacher = User.objects.create_user(
            username="viewer_teacher",
            email="viewer_teacher@example.com",
            password="StrongPass123!",
        )
        self.superadmin = User.objects.create_superuser(
            username="viewer_superadmin",
            email="viewer_superadmin@example.com",
            password="StrongPass123!",
        )

        self.org_a = Organization.objects.create(
            name="Student Exam Org A",
            org_type=OrganizationType.SCHOOL,
            owner=self.teacher,
            status="active",
            is_active=True,
        )
        self.org_b = Organization.objects.create(
            name="Student Exam Org B",
            org_type=OrganizationType.SCHOOL,
            owner=self.superadmin,
            status="active",
            is_active=True,
        )

        _assign_user_to_org(self.teacher, self.org_a, ProfileRole.TEACHER)
        _assign_user_to_org(self.student, self.org_a, ProfileRole.STUDENT)
        _assign_user_to_org(self.viewer_teacher, self.org_a, ProfileRole.TEACHER)

        self.assigned_exam = Exam.objects.create(
            author=self.teacher,
            title="Assigned Student Exam",
            is_active=True,
            is_public=False,
        )
        self.assigned_exam.allowed_users.add(self.student)

        self.assigned_course = Course.objects.create(
            owner=self.teacher,
            title="Student Course Assignment",
            status="published",
        )
        CourseMembership.objects.create(
            course=self.assigned_course,
            user=self.student,
            role="student",
        )

        self.course_assigned_exam = Exam.objects.create(
            author=self.teacher,
            title="Course Assigned Exam",
            is_active=True,
            is_public=False,
            course=self.assigned_course,
        )
        ExamQuestion.objects.create(
            exam=self.course_assigned_exam,
            text="Course question",
            order=1,
            points=1,
        )

        self.course_code_exam = Exam.objects.create(
            author=self.teacher,
            title="Course Code Exam",
            is_active=True,
            is_public=False,
            access_code="777777",
            course=self.assigned_course,
        )
        ExamQuestion.objects.create(
            exam=self.course_code_exam,
            text="Course code question",
            order=1,
            points=1,
        )

        self.assigned_public_exam = Exam.objects.create(
            author=self.teacher,
            title="Assigned Public Exam",
            is_active=True,
            is_public=True,
        )
        self.assigned_public_exam.allowed_users.add(self.student)

        self.code_assigned_exam = Exam.objects.create(
            author=self.teacher,
            title="Code Assigned Exam",
            is_active=True,
            is_public=False,
            access_code="123456",
        )
        self.code_assigned_exam.allowed_users.add(self.student)
        ExamQuestion.objects.create(
            exam=self.code_assigned_exam,
            text="Code-protected question",
            order=1,
            points=1,
        )

        self.code_assigned_no_questions_exam = Exam.objects.create(
            author=self.teacher,
            title="Code Assigned No Questions Exam",
            is_active=True,
            is_public=False,
            access_code="654321",
        )
        self.code_assigned_no_questions_exam.allowed_users.add(self.student)

        self.code_unassigned_exam = Exam.objects.create(
            author=self.teacher,
            title="Code Unassigned Exam",
            is_active=True,
            is_public=False,
            access_code="123456",
        )

        self.unassigned_public_exam = Exam.objects.create(
            author=self.teacher,
            title="Unassigned Public Exam",
            is_active=True,
            is_public=True,
        )
        ExamQuestion.objects.create(
            exam=self.unassigned_public_exam,
            text="Public exam question",
            order=1,
            points=1,
        )

        self.unassigned_private_exam = Exam.objects.create(
            author=self.teacher,
            title="Unassigned Private Exam",
            is_active=True,
            is_public=False,
        )

        self.student_group = StudentGroup.objects.create(
            teacher=self.teacher,
            organization=self.org_a,
            name="Visibility Group A",
        )
        self.student_group.students.add(self.student)

        self.group_assigned_exam = Exam.objects.create(
            author=self.teacher,
            title="Group Assigned Exam",
            is_active=True,
            is_public=False,
        )
        self.group_assigned_exam.allowed_groups.add(self.student_group)
        ExamQuestion.objects.create(
            exam=self.group_assigned_exam,
            text="Group assignment question",
            order=1,
            points=1,
        )

        self.other_tenant_exam = Exam.objects.create(
            author=self.teacher,
            title="Assigned But Other Tenant",
            is_active=True,
            is_public=False,
            organization=self.org_b,
        )
        self.other_tenant_exam.allowed_users.add(self.student)

        _login_with_org(self.client, self.student, self.org_a)

    def test_student_available_exam_list_includes_public_exams_in_active_tenant(self):
        response = self.client.get(reverse("exams:student_exam_list"), {"q": self.course_assigned_exam.title})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.course_assigned_exam.title)
        self.assertNotContains(response, 'class="ex-eyebrow"')

        response = self.client.get(reverse("exams:student_exam_list"), {"q": self.unassigned_public_exam.title})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.unassigned_public_exam.title)

        response = self.client.get(reverse("exams:student_exam_list"), {"q": self.other_tenant_exam.title})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, self.other_tenant_exam.slug)

    def test_student_available_exam_list_hides_soft_deleted_exam(self):
        # Baseline — imtahan görünür.
        response = self.client.get(reverse("exams:student_exam_list"), {"q": self.course_assigned_exam.title})
        self.assertContains(response, self.course_assigned_exam.slug)

        # is_active toxunulmadan yalnız yumşaq silinir; siyahıdan çıxmasının
        # yeganə səbəbi is_deleted filtridir.
        self.course_assigned_exam.is_deleted = True
        self.course_assigned_exam.save(update_fields=["is_deleted"])

        response = self.client.get(reverse("exams:student_exam_list"), {"q": self.course_assigned_exam.title})
        self.assertNotContains(response, self.course_assigned_exam.slug)

    def test_student_exam_list_never_shows_final_or_midterm(self):
        # Açıq "İmtahanlar" siyahısı yalnız sınaq/canlı imtahanlar üçündür —
        # final/midterm heç bir halda görünmür (tablarda belə), tip filtri
        # ilə birbaşa URL yığılsa da sızmır.
        final_exam = Exam.objects.create(
            author=self.teacher,
            title="Hidden Final Exam",
            exam_type="test",
            exam_type_extended="final",
            is_active=True,
            is_public=True,
        )
        midterm_exam = Exam.objects.create(
            author=self.teacher,
            title="Hidden Midterm Exam",
            exam_type="test",
            exam_type_extended="midterm",
            is_active=True,
            is_public=True,
        )
        quiz_exam = Exam.objects.create(
            author=self.teacher,
            title="Visible Quiz Exam",
            exam_type="test",
            exam_type_extended="quiz",
            is_active=True,
            is_public=True,
        )
        ExamQuestion.objects.create(exam=quiz_exam, text="Quiz question", order=1, points=1)

        response = self.client.get(reverse("exams:student_exam_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, quiz_exam.title)
        self.assertNotContains(response, final_exam.title)
        self.assertNotContains(response, midterm_exam.title)
        tab_keys = {tab["key"] for tab in response.context["type_tabs"]}
        self.assertNotIn("final", tab_keys)
        self.assertNotIn("midterm", tab_keys)
        self.assertIn("quiz", tab_keys)

        forced = self.client.get(reverse("exams:student_exam_list"), {"type": "final"})
        self.assertEqual(forced.status_code, 200)
        self.assertNotContains(forced, final_exam.title)

    def test_student_exam_list_filters_practical_separately_from_written(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Visible Written Filter Exam",
            exam_type="written",
            is_active=True,
            is_public=True,
        )
        ExamQuestion.objects.create(exam=written_exam, text="Written filter question", order=1, points=1)

        coding_exam = Exam.objects.create(
            author=self.teacher,
            title="Visible Practical Filter Exam",
            exam_type="coding",
            is_active=True,
            is_public=True,
        )
        ExamQuestion.objects.create(exam=coding_exam, text="Practical filter question", order=1, points=1)

        response = self.client.get(reverse("exams:student_exam_list"), {"type": "coding"})

        self.assertEqual(response.status_code, 200)
        self.assertRegex(response.content.decode(), r'<a class="ex-tab on"[^>]*data-type="coding"')
        self.assertContains(response, coding_exam.title)
        self.assertNotContains(response, written_exam.title)

    def test_student_exam_type_counts_exclude_attempt_exhausted_exams(self):
        exhausted_exam = Exam.objects.create(
            author=self.teacher,
            title="Exhausted Filter Count Exam",
            exam_type="test",
            is_active=True,
            is_public=True,
            max_attempts_per_user=1,
        )
        ExamQuestion.objects.create(exam=exhausted_exam, text="Exhausted count question", order=1, points=1)
        ExamAttempt.objects.create(
            user=self.student,
            exam=exhausted_exam,
            status="submitted",
            attempt_number=1,
        )

        response = self.client.get(reverse("exams:student_exam_list"), {"q": exhausted_exam.title})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["type_counts"]["total"], 0)
        self.assertEqual(response.context["type_counts"]["test"], 0)
        self.assertNotContains(response, f'data-exam-slug="{exhausted_exam.slug}"')
        self.assertContains(response, "Nəticə tapılmadı")

    def test_student_exam_card_shows_category_and_mechanic_badges(self):
        # Açıq siyahıda final/midterm görünmür — kateqoriya nişanı sınaq (quiz)
        # imtahanı üzərində yoxlanılır.
        quiz_test_exam = Exam.objects.create(
            author=self.teacher,
            title="Quiz Test Badge Exam",
            exam_type="test",
            exam_type_extended="quiz",
            is_active=True,
            is_public=True,
        )
        ExamQuestion.objects.create(exam=quiz_test_exam, text="Quiz test badge question", order=1, points=1)

        response = self.client.get(reverse("exams:student_exam_list"), {"q": quiz_test_exam.title})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, quiz_test_exam.title)
        self.assertContains(response, '<span class="ex-badge" data-type="quiz">', html=False)
        self.assertContains(response, '<span class="ex-badge ex-badge--mechanic" data-type="test">', html=False)

    def test_student_exam_views_restore_profile_org_context_when_session_org_is_missing(self):
        session = self.client.session
        session.pop("active_organization", None)
        session.save()

        available_response = self.client.get(
            reverse("exams:student_exam_list"), {"q": self.unassigned_public_exam.title}
        )
        self.assertEqual(available_response.status_code, 200)
        self.assertContains(available_response, self.unassigned_public_exam.title)
        self.assertEqual(self.client.session.get("active_organization"), self.org_a.slug)

        grouped_available_response = self.client.get(
            reverse("exams:student_exam_list"),
            {"q": self.group_assigned_exam.title},
        )
        self.assertEqual(grouped_available_response.status_code, 200)
        self.assertContains(grouped_available_response, self.group_assigned_exam.title)

        assigned_response = self.client.get(reverse("exams:assigned_exam_list"), {"q": self.group_assigned_exam.title})
        self.assertEqual(assigned_response.status_code, 200)
        self.assertContains(assigned_response, self.group_assigned_exam.title)

        start_response = self.client.get(reverse("exams:start_exam", args=[self.group_assigned_exam.slug]))
        self.assertEqual(start_response.status_code, 302)
        self.assertTrue(self.group_assigned_exam.attempts.filter(user=self.student).exists())

    def test_student_assigned_exam_list_shows_only_assigned_in_active_tenant(self):
        response = self.client.get(reverse("exams:assigned_exam_list"), {"q": self.assigned_exam.title})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.assigned_exam.title)

        response = self.client.get(reverse("exams:assigned_exam_list"), {"q": self.course_assigned_exam.title})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.course_assigned_exam.title)

        response = self.client.get(reverse("exams:assigned_exam_list"), {"q": self.code_assigned_exam.title})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.code_assigned_exam.title)

        response = self.client.get(reverse("exams:assigned_exam_list"), {"q": self.group_assigned_exam.title})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.group_assigned_exam.title)

        response = self.client.get(reverse("exams:assigned_exam_list"), {"q": self.assigned_public_exam.title})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, self.assigned_public_exam.slug)

        response = self.client.get(reverse("exams:assigned_exam_list"), {"q": self.unassigned_public_exam.title})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, self.unassigned_public_exam.slug)

        response = self.client.get(reverse("exams:assigned_exam_list"), {"q": self.unassigned_private_exam.title})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, self.unassigned_private_exam.slug)

        response = self.client.get(reverse("exams:assigned_exam_list"), {"q": self.other_tenant_exam.title})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, self.other_tenant_exam.slug)

    def test_course_assigned_exam_can_be_started_by_student(self):
        response = self.client.get(reverse("exams:start_exam", args=[self.course_assigned_exam.slug]))
        self.assertEqual(response.status_code, 302)
        self.assertTrue(self.course_assigned_exam.attempts.filter(user=self.student).exists())

    def test_in_progress_exam_resumes_from_start_route_when_attempt_limit_is_one(self):
        self.course_assigned_exam.max_attempts_per_user = 1
        self.course_assigned_exam.save(update_fields=["max_attempts_per_user"])
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.course_assigned_exam,
            status="in_progress",
            attempt_number=1,
        )

        response = self.client.get(reverse("exams:start_exam", args=[self.course_assigned_exam.slug]))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("exams:take_exam", args=[self.course_assigned_exam.slug, attempt.id]))

    def test_in_progress_code_exam_resumes_without_reasking_for_code(self):
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.code_assigned_exam,
            status="in_progress",
            attempt_number=1,
        )

        response = self.client.get(reverse("exams:start_exam", args=[self.code_assigned_exam.slug]))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("exams:take_exam", args=[self.code_assigned_exam.slug, attempt.id]))

    def test_timed_out_test_attempt_no_longer_resumes_from_start_route(self):
        self.course_assigned_exam.total_duration_minutes = 30
        self.course_assigned_exam.max_attempts_per_user = 1
        self.course_assigned_exam.save(update_fields=["total_duration_minutes", "max_attempts_per_user"])
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.course_assigned_exam,
            status="in_progress",
            attempt_number=1,
        )
        attempt.started_at = timezone.now() - timedelta(minutes=31)
        attempt.save(update_fields=["started_at"])

        response = self.client.get(reverse("exams:start_exam", args=[self.course_assigned_exam.slug]))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("exams:exam_result", args=[self.course_assigned_exam.slug, attempt.id]))
        attempt.refresh_from_db()
        self.assertEqual(attempt.status, "expired")
        self.assertIsNotNone(attempt.finished_at)

    def test_timed_out_written_attempt_starts_new_attempt_when_attempts_remain(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Written Retry Exam",
            is_active=True,
            is_public=False,
            exam_type="written",
            total_duration_minutes=20,
            max_attempts_per_user=2,
        )
        written_exam.allowed_users.add(self.student)
        ExamQuestion.objects.create(
            exam=written_exam,
            text="Written retry question",
            order=1,
            points=1,
        )
        old_attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="in_progress",
            attempt_number=1,
        )
        old_attempt.started_at = timezone.now() - timedelta(minutes=21)
        old_attempt.save(update_fields=["started_at"])

        response = self.client.get(reverse("exams:start_exam", args=[written_exam.slug]))

        self.assertEqual(response.status_code, 302)
        old_attempt.refresh_from_db()
        self.assertEqual(old_attempt.status, "expired")
        new_attempt = written_exam.attempts.get(user=self.student, attempt_number=2)
        self.assertEqual(new_attempt.status, "in_progress")
        self.assertEqual(response.url, reverse("exams:take_exam", args=[written_exam.slug, new_attempt.id]))

    def test_take_exam_redirects_timed_out_written_attempt_to_result(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Written Direct Timeout Exam",
            is_active=True,
            is_public=False,
            exam_type="written",
            total_duration_minutes=15,
        )
        written_exam.allowed_users.add(self.student)
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Written timeout question",
            order=1,
            points=1,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="in_progress",
            attempt_number=1,
        )
        ExamAnswer.objects.create(attempt=attempt, question=question, text_answer="Draft answer")
        attempt.started_at = timezone.now() - timedelta(minutes=16)
        attempt.save(update_fields=["started_at"])

        response = self.client.get(reverse("exams:take_exam", args=[written_exam.slug, attempt.id]))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("exams:exam_result", args=[written_exam.slug, attempt.id]))
        attempt.refresh_from_db()
        self.assertEqual(attempt.status, "expired")
        self.assertIsNotNone(attempt.finished_at)

    def test_take_exam_finished_attempt_ajax_finish_returns_result_json(self):
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.course_assigned_exam,
            status="submitted",
            attempt_number=1,
            finished_at=timezone.now(),
        )

        response = self.client.post(
            reverse("exams:take_exam", args=[self.course_assigned_exam.slug, attempt.id]),
            {"submit_action": "finish"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertTrue(payload["finished"])
        self.assertTrue(payload["already_finished"])
        self.assertEqual(
            payload["redirect_url"],
            reverse("exams:exam_result", args=[self.course_assigned_exam.slug, attempt.id]),
        )

    def test_take_exam_finished_attempt_ajax_autosave_returns_result_json(self):
        question = self.course_assigned_exam.questions.first()
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.course_assigned_exam,
            status="submitted",
            attempt_number=1,
            finished_at=timezone.now(),
        )
        if question:
            ExamAnswer.objects.create(attempt=attempt, question=question, text_answer="Already submitted")

        response = self.client.post(
            reverse("exams:take_exam", args=[self.course_assigned_exam.slug, attempt.id]),
            {
                "submit_action": "autosave",
                "changed_questions[]": [str(question.id)] if question else [],
                f"q_{question.id}" if question else "q_0": "Late autosave",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertTrue(payload["finished"])
        self.assertTrue(payload["already_finished"])
        self.assertEqual(
            payload["redirect_url"],
            reverse("exams:exam_result", args=[self.course_assigned_exam.slug, attempt.id]),
        )

    @staticmethod
    def _take_exam_static_source(relative_path):
        """take_exam JS-i artıq inline deyil (refaktor 2026-07-02) — davranış
        müqaviləsi xarici static faylın məzmununda yoxlanılır."""
        from pathlib import Path

        from django.contrib.staticfiles import finders

        found = finders.find(relative_path)
        assert found, f"Static fayl tapılmadı: {relative_path}"
        return Path(found).read_text(encoding="utf-8")

    def test_take_exam_uses_deadline_based_timer_logic_for_background_tabs(self):
        self.course_assigned_exam.total_duration_minutes = 30
        self.course_assigned_exam.default_question_time_seconds = 45
        self.course_assigned_exam.save(update_fields=["total_duration_minutes", "default_question_time_seconds"])
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.course_assigned_exam,
            status="in_progress",
            attempt_number=1,
        )

        response = self.client.get(reverse("exams:take_exam", args=[self.course_assigned_exam.slug, attempt.id]))

        self.assertEqual(response.status_code, 200)
        # Səhifə taymer skriptlərini yükləyir + xəbərdarlıq modalı yerindədir.
        self.assertContains(response, "exams/js/take_exam/timers.js")
        self.assertContains(response, "examTimeWarningModal")
        self.assertContains(response, "exam_time_warning.js")
        # Davranış müqaviləsi: taymerlər tick-əsaslı yox, DEADLINE-əsaslıdır
        # (arxa fon tab-larında drift olmur).
        timers_source = self._take_exam_static_source("exams/js/take_exam/timers.js")
        self.assertIn("examTimerDeadlineMs = Date.now() + (remainingSeconds * 1000)", timers_source)
        self.assertIn("questionTimerDeadlineMs = Date.now() + (timeLimit * 1000)", timers_source)

    @override_settings(EXAM_AUTOSAVE_INTERVAL_MS=300000, EXAM_AUTOSAVE_JITTER_MS=60000)
    def test_take_exam_uses_five_minute_server_autosave_with_jitter(self):
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.course_assigned_exam,
            status="in_progress",
            attempt_number=1,
        )

        response = self.client.get(reverse("exams:take_exam", args=[self.course_assigned_exam.slug, attempt.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-autosave-interval-ms="300000"')
        self.assertContains(response, 'data-autosave-jitter-ms="60000"')
        self.assertContains(response, "exams/js/take_exam/config.js")
        # Davranış müqaviləsi: 5 dəqiqəlik server default-u + jitter əlavəsi.
        config_source = self._take_exam_static_source("exams/js/take_exam/config.js")
        self.assertIn("defaultAutoSaveIntervalMs = 300000", config_source)
        self.assertIn("autoSaveDelayMs = serverAutoSaveIntervalMs + autoSaveSpread(", config_source)

    def test_take_exam_time_warning_modal_strings_are_translated_for_supported_languages(self):
        self.course_assigned_exam.total_duration_minutes = 30
        self.course_assigned_exam.save(update_fields=["total_duration_minutes"])
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.course_assigned_exam,
            status="in_progress",
            attempt_number=1,
        )

        expected_titles = {
            "az": "İmtahanın bitməsinə 5 dəqiqə qalıb",
            "en": "5 minutes left until the exam ends",
            "ru": "До окончания экзамена осталось 5 минут",
            "tr": "Sınavın bitmesine 5 dakika kaldı",
        }
        for language_code, expected_title in expected_titles.items():
            with self.subTest(language_code=language_code), override(language_code):
                response = self.client.get(
                    reverse("exams:take_exam", args=[self.course_assigned_exam.slug, attempt.id]),
                    HTTP_ACCEPT_LANGUAGE=language_code,
                )
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, expected_title)

    def test_take_exam_autosave_updates_only_changed_questions(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Interval Autosave Written Exam",
            is_active=True,
            is_public=False,
            exam_type="written",
        )
        written_exam.allowed_users.add(self.student)
        first_question = ExamQuestion.objects.create(
            exam=written_exam,
            text="First written question",
            order=1,
            points=1,
        )
        second_question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Second written question",
            order=2,
            points=1,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="in_progress",
            attempt_number=1,
        )
        first_answer = ExamAnswer.objects.create(
            attempt=attempt,
            question=first_question,
            text_answer="Old first answer",
        )
        second_answer = ExamAnswer.objects.create(
            attempt=attempt,
            question=second_question,
            text_answer="Keep second answer",
        )

        response = self.client.post(
            reverse("exams:take_exam", args=[written_exam.slug, attempt.id]),
            {
                "submit_action": "autosave",
                "changed_questions[]": [str(first_question.id)],
                f"q_{first_question.id}": "New first answer",
                f"q_{second_question.id}": "Should not overwrite",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        first_answer.refresh_from_db()
        second_answer.refresh_from_db()
        self.assertEqual(first_answer.text_answer, "New first answer")
        self.assertEqual(second_answer.text_answer, "Keep second answer")

    @override_settings(EXAM_AUTOSAVE_BINARY_UPLOADS_ENABLED=False)
    def test_take_exam_autosave_ignores_file_and_paint_payloads(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Autosave Binary Guard Written Exam",
            is_active=True,
            is_public=False,
            exam_type="written",
            enable_paint=True,
        )
        written_exam.allowed_users.add(self.student)
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Upload-heavy written question",
            order=1,
            points=1,
            enable_paint=True,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="in_progress",
            attempt_number=1,
        )
        answer = ExamAnswer.objects.create(attempt=attempt, question=question)
        paint_data_url = "data:image/png;base64," + base64.b64encode(_TINY_PNG_BYTES).decode("ascii")
        uploaded_file = SimpleUploadedFile("answer.pdf", b"%PDF-1.4\n", content_type="application/pdf")

        response = self.client.post(
            reverse("exams:take_exam", args=[written_exam.slug, attempt.id]),
            {
                "submit_action": "autosave",
                "changed_questions[]": [str(question.id)],
                f"q_{question.id}": "Text should autosave",
                f"paint_enabled_{question.id}": "1",
                f"paint_data_{question.id}": paint_data_url,
                f"file_{question.id}[]": uploaded_file,
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        answer.refresh_from_db()
        self.assertEqual(answer.text_answer, "Text should autosave")
        self.assertFalse(answer.files.exists())
        self.assertFalse(answer.has_paint)
        self.assertFalse(bool(answer.paint_data_url))

    @override_settings(EXAM_ANSWER_MAX_FILES_PER_QUESTION=1, EXAM_ANSWER_FILE_MAX_SIZE_MB=1)
    def test_take_exam_rejects_too_many_files_for_one_written_answer(self):
        written_exam = Exam.objects.create(
            author=self.teacher,
            title="Written File Limit Exam",
            is_active=True,
            is_public=False,
            exam_type="written",
        )
        written_exam.allowed_users.add(self.student)
        question = ExamQuestion.objects.create(
            exam=written_exam,
            text="Upload limited question",
            order=1,
            points=1,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=written_exam,
            status="in_progress",
            attempt_number=1,
        )
        answer = ExamAnswer.objects.create(attempt=attempt, question=question)
        first_file = SimpleUploadedFile("first.pdf", b"%PDF-1.4\n", content_type="application/pdf")
        second_file = SimpleUploadedFile("second.pdf", b"%PDF-1.4\n", content_type="application/pdf")

        response = self.client.post(
            reverse("exams:take_exam", args=[written_exam.slug, attempt.id]),
            {
                "submit_action": "save_draft",
                f"q_{question.id}": "Draft with too many files",
                f"file_{question.id}[]": [first_file, second_file],
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("maksimum 1", response.json()["error"])
        answer.refresh_from_db()
        self.assertFalse(answer.files.exists())

    def test_take_exam_test_autosave_skips_full_score_recalculation(self):
        test_exam = Exam.objects.create(
            author=self.teacher,
            title="High Capacity Autosave Test Exam",
            is_active=True,
            is_public=False,
            exam_type="test",
        )
        test_exam.allowed_users.add(self.student)
        question = ExamQuestion.objects.create(
            exam=test_exam,
            text="Two plus two?",
            order=1,
            points=1,
            answer_mode="single",
        )
        correct_option = ExamQuestionOption.objects.create(
            question=question,
            label="A",
            text="4",
            is_correct=True,
        )
        ExamQuestionOption.objects.create(
            question=question,
            label="B",
            text="5",
            is_correct=False,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=test_exam,
            status="in_progress",
            attempt_number=1,
        )
        answer = ExamAnswer.objects.create(attempt=attempt, question=question)

        with patch.object(ExamAttempt, "recalculate_score", autospec=True) as recalculate_score:
            response = self.client.post(
                reverse("exams:take_exam", args=[test_exam.slug, attempt.id]),
                {
                    "submit_action": "autosave",
                    "changed_questions[]": [str(question.id)],
                    f"q_{question.id}": str(correct_option.id),
                },
                HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            )

        self.assertEqual(response.status_code, 200)
        recalculate_score.assert_not_called()
        answer.refresh_from_db()
        self.assertEqual(set(answer.selected_options.values_list("id", flat=True)), {correct_option.id})
        self.assertTrue(answer.is_correct)
        attempt.refresh_from_db()
        self.assertEqual(attempt.status, "in_progress")
        self.assertEqual(attempt.correct_count, 0)

    def test_course_dashboard_student_exam_actions_use_info_modal(self):
        response = self.client.get(reverse("courses:course_dashboard", args=[self.assigned_course.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="courseExamInfoBackdrop"')
        self.assertContains(response, "js-open-course-exam-modal")
        self.assertContains(response, f'data-exam-slug="{self.course_assigned_exam.slug}"')
        self.assertContains(response, f'data-exam-slug="{self.course_code_exam.slug}"')
        self.assertContains(response, 'data-requires-code="1"')
        self.assertContains(response, 'id="courseExamLanguageBlock"')
        self.assertContains(response, 'id="courseExamLanguageSelect"')
        self.assertContains(response, 'name="language" id="courseExamCodeLanguage"')
        self.assertContains(
            response,
            f'name="next" value="{reverse("courses:course_dashboard", args=[self.assigned_course.id])}"',
        )

    def test_course_dashboard_exam_modal_embeds_language_options(self):
        variant_az = self.course_assigned_exam.language_variants.create(
            language="az",
            display_name="Azərbaycan dili",
        )
        variant_en = self.course_assigned_exam.language_variants.create(language="en", display_name="English")
        self.course_assigned_exam.questions.update(language="az", language_variant=variant_az)
        ExamQuestion.objects.create(
            exam=self.course_assigned_exam,
            text="Course EN modal question",
            order=2,
            points=1,
            language="en",
            language_variant=variant_en,
        )

        response = self.client.get(reverse("courses:course_dashboard", args=[self.assigned_course.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f'data-language-options-id="course-exam-language-options-{self.course_assigned_exam.id}"',
        )
        self.assertContains(response, f'id="course-exam-language-options-{self.course_assigned_exam.id}"')
        self.assertContains(response, '"display_name": "English"', html=False)
        self.assertNotContains(response, '"count"', html=False)
        self.assertNotContains(response, "Tövsiyə olunan")

    def test_student_exam_list_actions_use_bootstrap_info_modal(self):
        available_response = self.client.get(reverse("exams:student_exam_list"), {"q": self.course_assigned_exam.title})

        self.assertEqual(available_response.status_code, 200)
        self.assertContains(available_response, 'id="examStartModal"')
        self.assertContains(available_response, "data-open-exam-start-modal")
        self.assertContains(available_response, f'data-exam-slug="{self.course_assigned_exam.slug}"')
        self.assertContains(
            available_response,
            f'data-start-url="{reverse("exams:start_exam", args=[self.course_assigned_exam.slug])}',
        )
        self.assertContains(available_response, 'data-requires-code="0"')
        self.assertContains(available_response, 'id="examStartCodeForm"')
        self.assertContains(available_response, 'name="next" value="')

        code_response = self.client.get(reverse("exams:student_exam_list"), {"q": self.code_assigned_exam.title})

        self.assertEqual(code_response.status_code, 200)
        self.assertContains(code_response, f'data-exam-slug="{self.code_assigned_exam.slug}"')
        self.assertContains(code_response, 'data-requires-code="1"')
        self.assertContains(
            code_response,
            'modalAccessCodeDescriptionWithTitle: "\\u0022{title}\\u0022 imtahanına başlamaq üçün giriş kodunu daxil edin."',
            html=False,
        )

    def test_student_exam_list_modal_embeds_language_select_for_multilingual_exam(self):
        multilingual_exam = Exam.objects.create(
            author=self.teacher,
            title="Student Modal Language Exam",
            is_active=True,
            is_public=True,
        )
        variant_az = multilingual_exam.language_variants.create(language="az", display_name="Azərbaycan dili")
        variant_en = multilingual_exam.language_variants.create(language="en", display_name="English")
        ExamQuestion.objects.create(
            exam=multilingual_exam,
            text="AZ modal question",
            order=1,
            points=1,
            language="az",
            language_variant=variant_az,
        )
        ExamQuestion.objects.create(
            exam=multilingual_exam,
            text="EN modal question",
            order=2,
            points=1,
            language="en",
            language_variant=variant_en,
        )

        response = self.client.get(reverse("exams:student_exam_list"), {"q": multilingual_exam.title})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="examStartLanguageBlock"')
        self.assertContains(response, 'id="examStartLanguageSelect"')
        self.assertContains(response, "İmtahan hansı dildə olsun?")
        self.assertContains(response, f'data-language-options-id="exam-language-options-{multilingual_exam.id}"')
        self.assertContains(response, f'id="exam-language-options-{multilingual_exam.id}"')
        self.assertContains(response, '"display_name": "English"', html=False)
        self.assertContains(response, "buildStartUrlWithLanguage")
        self.assertNotContains(response, '"count"', html=False)
        self.assertNotContains(response, "Tövsiyə olunan")

    def test_multilingual_exam_start_uses_selected_language_without_language_page(self):
        multilingual_exam = Exam.objects.create(
            author=self.teacher,
            title="Selected Language Start Exam",
            is_active=True,
            is_public=True,
        )
        variant_az = multilingual_exam.language_variants.create(language="az", display_name="Azərbaycan dili")
        variant_en = multilingual_exam.language_variants.create(language="en", display_name="English")
        ExamQuestion.objects.create(
            exam=multilingual_exam,
            text="AZ start question",
            order=1,
            points=1,
            language="az",
            language_variant=variant_az,
        )
        ExamQuestion.objects.create(
            exam=multilingual_exam,
            text="EN start question",
            order=2,
            points=1,
            language="en",
            language_variant=variant_en,
        )

        missing_language_response = self.client.get(reverse("exams:start_exam", args=[multilingual_exam.slug]))
        self.assertEqual(missing_language_response.status_code, 302)
        self.assertEqual(missing_language_response.url, reverse("exams:student_exam_list"))
        self.assertFalse(multilingual_exam.attempts.filter(user=self.student).exists())

        selected_language_response = self.client.get(
            reverse("exams:start_exam", args=[multilingual_exam.slug]),
            {"language": "en"},
        )

        self.assertEqual(selected_language_response.status_code, 302)
        attempt = multilingual_exam.attempts.get(user=self.student)
        self.assertEqual(attempt.language, "en")
        self.assertEqual(attempt.language_variant, variant_en)
        self.assertEqual(
            selected_language_response.url,
            reverse("exams:take_exam", args=[multilingual_exam.slug, attempt.id]),
        )

    def test_course_dashboard_student_history_button_shows_attempt_count(self):
        ExamAttempt.objects.create(
            user=self.student,
            exam=self.course_assigned_exam,
            status="submitted",
            attempt_number=1,
        )

        response = self.client.get(reverse("courses:course_dashboard", args=[self.assigned_course.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cavablarım (1)")
        self.assertContains(
            response,
            f'{reverse("exams:student_exam_history")}?exam={self.course_assigned_exam.slug}',
        )

    def test_unassigned_private_exam_cannot_be_started(self):
        response = self.client.get(reverse("exams:start_exam", args=[self.unassigned_private_exam.slug]))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("exams:student_exam_list"))
        self.assertFalse(self.unassigned_private_exam.attempts.filter(user=self.student).exists())

    def test_unassigned_private_exam_redirects_back_to_profile_assigned_section_when_requested(self):
        response = self.client.get(
            reverse("exams:start_exam", args=[self.unassigned_private_exam.slug]),
            {"from_section": "assigned-exams", "assigned_type": "exams"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, f"{reverse('accounts:profile')}?section=assigned-exams&assigned_type=exams")
        self.assertFalse(self.unassigned_private_exam.attempts.filter(user=self.student).exists())

    def test_assigned_exam_with_code_requires_code_before_start(self):
        response = self.client.get(reverse("exams:start_exam", args=[self.code_assigned_exam.slug]))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("exams:student_exam_list"))
        self.assertFalse(self.code_assigned_exam.attempts.filter(user=self.student).exists())

    def test_assigned_exam_with_code_starts_after_correct_code(self):
        response = self.client.post(
            reverse("exams:exam_code_check"),
            {"exam_slug": self.code_assigned_exam.slug, "access_code": "123456"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(self.code_assigned_exam.attempts.filter(user=self.student).exists())

    def test_assigned_exam_without_questions_cannot_be_started(self):
        response = self.client.get(
            reverse("exams:start_exam", args=[self.assigned_exam.slug]),
            {"from_section": "assigned-exams", "assigned_type": "exams"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, f"{reverse('accounts:profile')}?section=assigned-exams&assigned_type=exams")
        self.assertFalse(self.assigned_exam.attempts.filter(user=self.student).exists())

    def test_assigned_code_exam_without_questions_cannot_be_started(self):
        response = self.client.post(
            reverse("exams:exam_code_check"),
            {"exam_slug": self.code_assigned_no_questions_exam.slug, "access_code": "654321"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("exams:student_exam_list"))
        self.assertFalse(self.code_assigned_no_questions_exam.attempts.filter(user=self.student).exists())

    def test_take_exam_redirects_when_attempt_has_no_questions(self):
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.assigned_exam,
            status="in_progress",
        )

        response = self.client.get(reverse("exams:take_exam", args=[self.assigned_exam.slug, attempt.id]))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("exams:student_exam_list"))

    def test_exam_code_check_rejects_post_without_csrf_token(self):
        csrf_client = Client(enforce_csrf_checks=True)
        csrf_client.force_login(self.student)
        session = csrf_client.session
        session["active_organization"] = self.org_a.slug
        session.save()

        response = csrf_client.post(
            reverse("exams:exam_code_check"),
            {"exam_slug": self.code_assigned_exam.slug, "access_code": "123456"},
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(self.code_assigned_exam.attempts.filter(user=self.student).exists())

    def test_exam_code_check_sql_injection_payload_does_not_bypass_lookup(self):
        response = self.client.post(
            reverse("exams:exam_code_check"),
            {"exam_slug": "anything' OR 1=1 --", "access_code": "123456"},
        )

        self.assertEqual(response.status_code, 404)
        self.assertFalse(self.code_assigned_exam.attempts.filter(user=self.student).exists())

    def test_unassigned_exam_with_code_cannot_start_even_with_valid_code(self):
        response = self.client.post(
            reverse("exams:exam_code_check"),
            {"exam_slug": self.code_unassigned_exam.slug, "access_code": "123456"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("exams:student_exam_list"))
        self.assertFalse(self.code_unassigned_exam.attempts.filter(user=self.student).exists())

    def test_unassigned_exam_with_code_redirects_back_to_profile_assigned_section_when_requested(self):
        response = self.client.post(
            reverse("exams:exam_code_check"),
            {
                "exam_slug": self.code_unassigned_exam.slug,
                "access_code": "123456",
                "from_section": "assigned-exams",
                "assigned_type": "exams",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, f"{reverse('accounts:profile')}?section=assigned-exams&assigned_type=exams")
        self.assertFalse(self.code_unassigned_exam.attempts.filter(user=self.student).exists())

    def test_other_tenant_exam_cannot_be_started(self):
        response = self.client.get(reverse("exams:start_exam", args=[self.other_tenant_exam.slug]))
        self.assertEqual(response.status_code, 404)
        self.assertFalse(self.other_tenant_exam.attempts.filter(user=self.student).exists())

    def test_other_tenant_exam_code_check_is_not_found(self):
        code_exam = Exam.objects.create(
            author=self.teacher,
            title="Other Tenant Code Exam",
            is_active=True,
            is_public=False,
            access_code="654321",
            organization=self.org_b,
        )
        code_exam.allowed_users.add(self.student)

        response = self.client.post(
            reverse("exams:exam_code_check"),
            {"exam_slug": code_exam.slug, "access_code": "654321"},
        )
        self.assertEqual(response.status_code, 404)
        self.assertFalse(code_exam.attempts.filter(user=self.student).exists())

    def test_other_tenant_exam_result_is_not_found(self):
        other_tenant_attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.other_tenant_exam,
            status="submitted",
        )

        response = self.client.get(
            reverse("exams:exam_result", args=[self.other_tenant_exam.slug, other_tenant_attempt.id])
        )
        self.assertEqual(response.status_code, 404)

    def test_public_exams_are_visible_to_other_authenticated_roles(self):
        _login_with_org(self.client, self.viewer_teacher, self.org_a)
        teacher_response = self.client.get(reverse("exams:student_exam_list"))
        self.assertEqual(teacher_response.status_code, 200)
        self.assertContains(teacher_response, self.unassigned_public_exam.title)

        _login_with_org(self.client, self.superadmin, self.org_a)
        superadmin_response = self.client.get(reverse("exams:student_exam_list"))
        self.assertEqual(superadmin_response.status_code, 200)
        self.assertContains(superadmin_response, self.unassigned_public_exam.title)

    def test_az_student_exam_list_uses_localized_strings(self):
        with override("az"):
            response = self.client.get(
                reverse("exams:student_exam_list"),
                {"q": self.code_assigned_exam.title},
                HTTP_ACCEPT_LANGUAGE="az",
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "İmtahanlar")
        self.assertContains(response, "Giriş kodu tələb olunur")
        self.assertContains(response, "İmtahana başla")
        self.assertNotContains(response, "Start exam")
        self.assertNotContains(response, "Bu resurs haqqında qısa məlumat")
        self.assertContains(
            response,
            'modalAccessCodeDescriptionWithTitle: "\\u0022{title}\\u0022 imtahanına başlamaq üçün giriş kodunu daxil edin."',
            html=False,
        )
        self.assertNotContains(response, 'modalAccessCodeDescriptionWithTitle: ""{title}"', html=False)

    def test_student_exam_list_modal_strings_are_translated_for_supported_languages(self):
        cases = (
            ("en", "Exam details", "Access code required", "Start exam"),
            ("tr", "Sınav bilgileri", "Erişim kodu gerekiyor", "Sınavı başlat"),
            ("ru", "Информация об экзамене", "Требуется код доступа", "Начать экзамен"),
        )

        for language, modal_title, code_title, start_text in cases:
            with self.subTest(language=language), override(language):
                response = self.client.get(
                    reverse("exams:student_exam_list"),
                    {"q": self.code_assigned_exam.title},
                    HTTP_ACCEPT_LANGUAGE=language,
                )

                self.assertEqual(response.status_code, 200)
                self.assertContains(response, modal_title)
                self.assertContains(response, code_title)
                self.assertContains(response, start_text)
                self.assertNotContains(response, "modal_title_exam_info")
                self.assertNotContains(response, "modal_label_duration")
                self.assertNotContains(response, "modal_access_code_title")

    def test_az_exam_result_uses_localized_strings(self):
        exam = Exam.objects.create(
            author=self.teacher,
            title="Localized Result Exam",
            is_active=True,
            is_public=False,
        )
        exam.allowed_users.add(self.student)
        question = ExamQuestion.objects.create(
            exam=exam,
            text="Localized question",
            order=1,
            points=1,
        )
        correct_option = ExamQuestionOption.objects.create(
            question=question,
            text="Doğru variant",
            is_correct=True,
        )
        ExamQuestionOption.objects.create(
            question=question,
            text="Yanlış variant",
            is_correct=False,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=exam,
            status="submitted",
        )
        answer = ExamAnswer.objects.create(
            attempt=attempt,
            question=question,
            is_correct=True,
        )
        answer.selected_options.add(correct_option)
        attempt.recalculate_score()

        with override("az"):
            response = self.client.get(
                reverse("exams:exam_result", args=[exam.slug, attempt.id]),
                HTTP_ACCEPT_LANGUAGE="az",
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nəticəniz")
        self.assertContains(response, "İmtahan statusu")
        self.assertContains(response, "Təhvil verilib")
        self.assertContains(response, "Düzgün")
        self.assertContains(response, "Səhv")
        self.assertNotContains(response, "Subheading posts")
        self.assertNotContains(response, "Submitted")

    def test_exam_result_page_avoids_placeholder_copy_in_all_supported_languages(self):
        exam = Exam.objects.create(
            author=self.teacher,
            title="Localized Exam Result",
            is_active=True,
            is_public=False,
            course=self.assigned_course,
        )
        exam.allowed_users.add(self.student)
        question = ExamQuestion.objects.create(
            exam=exam,
            text="Localized question",
            order=1,
            points=1,
        )
        correct_option = ExamQuestionOption.objects.create(
            question=question,
            text="Correct option",
            is_correct=True,
        )
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=exam,
            status="submitted",
        )
        answer = ExamAnswer.objects.create(
            attempt=attempt,
            question=question,
            is_correct=True,
        )
        answer.selected_options.add(correct_option)
        attempt.recalculate_score()

        placeholder_strings = [
            "Subheading posts",
            "Answer unit",
            "Exam started! (Attempt #{attempt_number})",
            "Example: 60 (seconds). If empty, default is used.",
        ]

        for language in ["az", "en", "ru", "tr"]:
            with self.subTest(language=language):
                with override(language):
                    response = self.client.get(
                        reverse("exams:exam_result", args=[exam.slug, attempt.id]),
                        HTTP_ACCEPT_LANGUAGE=language,
                    )

                self.assertEqual(response.status_code, 200)
                for placeholder in placeholder_strings:
                    self.assertNotContains(response, placeholder)

    def test_filtered_exam_history_shows_only_selected_exam_attempts(self):
        selected_attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.course_assigned_exam,
            status="submitted",
            attempt_number=1,
        )
        other_attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.assigned_exam,
            status="submitted",
            attempt_number=1,
        )
        return_to = reverse("courses:course_dashboard", args=[self.assigned_course.id])

        response = self.client.get(
            reverse("exams:student_exam_history"),
            {"exam": self.course_assigned_exam.slug, "return_to": return_to},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["exam"], self.course_assigned_exam)
        self.assertEqual(response.context["back_url"], return_to)
        attempts = response.context["attempts"]
        self.assertEqual(len(attempts), 1)
        self.assertEqual(attempts[0].id, selected_attempt.id)
        self.assertNotContains(response, other_attempt.exam.title)
        self.assertContains(response, "Mənim cavablarım")
        self.assertContains(response, "Göndərdiyim cavablar")
        self.assertContains(response, "İstifadə olunmuş cəhd")
        self.assertContains(response, "Bax")
        self.assertNotContains(response, "İmtahana başla")

    def test_exam_result_defaults_back_to_course_dashboard_and_keeps_history_link_distinct(self):
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.course_assigned_exam,
            status="submitted",
            attempt_number=1,
        )
        expected_back_url = reverse("courses:course_dashboard", args=[self.assigned_course.id])

        response = self.client.get(reverse("exams:exam_result", args=[self.course_assigned_exam.slug, attempt.id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["back_url"], expected_back_url)
        self.assertEqual(
            response.context["history_url"],
            f'{reverse("exams:student_exam_history")}?{urlencode({"exam": self.course_assigned_exam.slug, "return_to": expected_back_url})}',
        )
        # İmtahan yenicə bitmiş görünüşdə naviqasiya render olunmur: alt
        # "İmtahanlara qayıt / Cavablarım" paneli silinib, "Profilə qayıt"
        # düyməsi isə yalnız profildən (my-results) baxılanda görünür.
        self.assertNotContains(response, "action-footer")
        self.assertNotContains(response, "result-back-btn")
        profile_response = self.client.get(
            reverse("exams:exam_result", args=[self.course_assigned_exam.slug, attempt.id]),
            {"from_section": "my-results"},
        )
        self.assertContains(profile_response, "result-back-btn")
        self.assertContains(profile_response, "section=my-results")

    def test_exam_result_restores_original_back_url_when_opened_from_history(self):
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.course_assigned_exam,
            status="submitted",
            attempt_number=1,
        )
        expected_back_url = reverse("courses:course_dashboard", args=[self.assigned_course.id])
        history_url = f'{reverse("exams:student_exam_history")}?{urlencode({"exam": self.course_assigned_exam.slug, "return_to": expected_back_url})}'

        response = self.client.get(
            reverse("exams:exam_result", args=[self.course_assigned_exam.slug, attempt.id]),
            {"return_to": history_url},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["history_url"], history_url)
        self.assertEqual(response.context["back_url"], expected_back_url)

    def test_take_exam_hides_previous_attempts_summary_while_attempt_is_active(self):
        previous_attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.course_assigned_exam,
            status="submitted",
            attempt_number=1,
        )

        start_response = self.client.get(
            reverse("exams:start_exam", args=[self.course_assigned_exam.slug]),
            {"next": reverse("courses:course_dashboard", args=[self.assigned_course.id])},
        )

        self.assertEqual(start_response.status_code, 302)
        take_response = self.client.get(start_response.url)

        self.assertEqual(take_response.status_code, 200)
        self.assertNotContains(take_response, "Əvvəlki cəhdlər")
        self.assertNotContains(take_response, "Cavablarım (1)")
        self.assertNotContains(
            take_response,
            reverse("exams:exam_result", args=[self.course_assigned_exam.slug, previous_attempt.id]),
        )
        self.assertContains(take_response, "exams/js/paint_answer.js")


class TeacherViewAttemptSearchPaginationTest(TestCase):
    def setUp(self):
        self.teacher = User.objects.create_user(
            username="view_attempt_teacher",
            email="view_attempt_teacher@example.com",
            password="StrongPass123!",
        )
        self.student = User.objects.create_user(
            username="view_attempt_student",
            email="view_attempt_student@example.com",
            password="StrongPass123!",
        )

        self.organization = Organization.objects.create(
            name="View Attempt Org",
            org_type=OrganizationType.SCHOOL,
            owner=self.teacher,
            status="active",
            is_active=True,
        )
        _assign_user_to_org(self.teacher, self.organization, ProfileRole.TEACHER)
        _assign_user_to_org(self.student, self.organization, ProfileRole.STUDENT)

        self.exam = Exam.objects.create(
            author=self.teacher,
            title="View Attempt Search Exam",
            exam_type="test",
            is_active=True,
        )
        self.attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.exam,
            status="submitted",
        )

        for index in range(1, 8):
            question = ExamQuestion.objects.create(
                exam=self.exam,
                text=f"Sual mətn test {index}",
                order=index,
                points=1,
            )
            correct_option = ExamQuestionOption.objects.create(
                question=question,
                text=f"Doğru cavab {index}",
                is_correct=True,
            )
            ExamQuestionOption.objects.create(
                question=question,
                text=f"Səhv cavab {index}",
                is_correct=False,
            )
            answer = ExamAnswer.objects.create(
                attempt=self.attempt,
                question=question,
                is_correct=True,
            )
            answer.selected_options.add(correct_option)

        _login_with_org(self.client, self.teacher, self.organization)

    def test_teacher_view_attempt_supports_search_and_questions_pagination(self):
        response = self.client.get(
            reverse("exams:teacher_view_attempt", args=[self.exam.slug, self.attempt.id]),
            {
                "q": "test 7",
                "questions_page": 1,
                "from_section": "review-results",
                "return_to": "/accounts/profile/?section=review-results",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["qa_search_query"], "test 7")
        self.assertEqual(len(response.context["qa_list"]), 1)
        self.assertIn("q=test+7", response.context["qa_pagination_query"])
        self.assertIn("from_section=review-results", response.context["qa_pagination_query"])
        self.assertIn("section%3Dreview-results", response.context["qa_pagination_query"])
        self.assertIn("from_section=review-results", response.context["qa_clear_search_url"])
        self.assertIn("section=review-results", response.context["qa_clear_search_url"])
        self.assertContains(response, 'class="attempt-search-group input-group"')

        response_page_two = self.client.get(
            reverse("exams:teacher_view_attempt", args=[self.exam.slug, self.attempt.id]),
            {"questions_page": 2},
        )
        self.assertEqual(response_page_two.status_code, 200)
        self.assertEqual(response_page_two.context["qa_page"].number, 2)
        self.assertEqual(len(response_page_two.context["qa_list"]), 1)

    def test_teacher_view_attempt_search_matches_option_text(self):
        response = self.client.get(
            reverse("exams:teacher_view_attempt", args=[self.exam.slug, self.attempt.id]),
            {"q": "Doğru cavab 5"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["qa_list"]), 1)
        self.assertContains(response, "Sual mətn test 5")


class StudentExamResultVisibilityWindowTest(TestCase):
    def setUp(self):
        from django.utils import timezone

        self.teacher = User.objects.create_user(
            username="result_visibility_teacher",
            email="result_visibility_teacher@example.com",
            password="StrongPass123!",
        )
        self.student = User.objects.create_user(
            username="result_visibility_student",
            email="result_visibility_student@example.com",
            password="StrongPass123!",
        )

        self.organization = Organization.objects.create(
            name="Exam Result Visibility Org",
            org_type=OrganizationType.SCHOOL,
            owner=self.teacher,
            status="active",
            is_active=True,
        )
        _assign_user_to_org(self.teacher, self.organization, ProfileRole.TEACHER)
        _assign_user_to_org(self.student, self.organization, ProfileRole.STUDENT)

        self.exam = Exam.objects.create(
            author=self.teacher,
            title="Visibility Written Exam",
            exam_type="written",
            is_active=True,
        )
        self.attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.exam,
            status="submitted",
            checked_by_teacher=True,
            teacher_score=82,
            teacher_checked_at=timezone.now(),
        )

        _login_with_org(self.client, self.student, self.organization)

    def test_exam_result_hidden_while_teacher_review_window_open(self):
        response = self.client.get(reverse("exams:exam_result", args=[self.exam.slug, self.attempt.id]))
        self.assertEqual(response.status_code, 302)
        self.assertIn("section=my-results", response.url)

    def test_exam_result_visible_after_teacher_review_window_closes(self):
        self.attempt.teacher_checked_at = timezone.now() - timedelta(minutes=6)
        self.attempt.save(update_fields=["teacher_checked_at"])

        response = self.client.get(reverse("exams:exam_result", args=[self.exam.slug, self.attempt.id]))
        self.assertEqual(response.status_code, 200)

    def _create_test_exam_with_answered_attempt(self, *, end_delta):
        """Bir sual + 2 variantlı test imtahanı və cavablı submitted attempt."""
        exam = Exam.objects.create(
            author=self.teacher,
            title=f"Release Lock Test Exam {end_delta}",
            exam_type="test",
            is_active=True,
            start_datetime=timezone.now() - timedelta(hours=2),
            end_datetime=timezone.now() + end_delta,
        )
        question = ExamQuestion.objects.create(
            exam=exam,
            text="Locked window question",
            order=1,
            answer_mode="single",
            points=1,
        )
        correct = ExamQuestionOption.objects.create(question=question, text="Correct option", is_correct=True)
        ExamQuestionOption.objects.create(question=question, text="Wrong option", is_correct=False)
        attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=exam,
            status="submitted",
        )
        answer = ExamAnswer.objects.create(attempt=attempt, question=question)
        answer.selected_options.add(correct)
        return exam, attempt

    def test_correct_answers_hidden_while_exam_window_open(self):
        """EXAM-P0-05: pəncərə açıq olduqca variant düzgünlüyü və verdikt sızmır."""
        exam, attempt = self._create_test_exam_with_answered_attempt(end_delta=timedelta(hours=2))

        response = self.client.get(reverse("exams:exam_result", args=[exam.slug, attempt.id]))
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["answers_release_locked"])
        self.assertTrue(response.context["hide_test_answer_correctness"])
        self.assertEqual(response.context["answer_verdict_by_qid"], {})
        self.assertNotContains(response, "correct-option")
        self.assertNotContains(response, "Correct option")

    def test_correct_answers_visible_after_exam_window_closes(self):
        exam, attempt = self._create_test_exam_with_answered_attempt(end_delta=timedelta(hours=-1))

        response = self.client.get(reverse("exams:exam_result", args=[exam.slug, attempt.id]))
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["answers_release_locked"])
        self.assertFalse(response.context["hide_test_answer_correctness"])
        self.assertContains(response, "correct-option")

    def test_open_ended_exam_is_not_release_locked(self):
        """end_datetime olmayan (məşq) imtahanda cavab analizi dərhal açıqdır."""
        exam = Exam.objects.create(
            author=self.teacher,
            title="Open Ended Practice Exam",
            exam_type="test",
            is_active=True,
        )
        question = ExamQuestion.objects.create(
            exam=exam,
            text="Practice question",
            order=1,
            answer_mode="single",
            points=1,
        )
        ExamQuestionOption.objects.create(question=question, text="Correct option", is_correct=True)
        attempt = ExamAttempt.objects.create(user=self.student, exam=exam, status="submitted")
        ExamAnswer.objects.create(attempt=attempt, question=question)

        response = self.client.get(reverse("exams:exam_result", args=[exam.slug, attempt.id]))
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["answers_release_locked"])

    def test_written_ideal_answer_hidden_while_exam_window_open(self):
        """EXAM-P0-05: yoxlanılmamış yazılı cavabda ideal cavab pəncərə açıq ikən görünmür."""
        exam = Exam.objects.create(
            author=self.teacher,
            title="Locked Written Exam",
            exam_type="written",
            is_active=True,
            start_datetime=timezone.now() - timedelta(hours=2),
            end_datetime=timezone.now() + timedelta(hours=2),
        )
        ExamQuestion.objects.create(
            exam=exam,
            text="Explain thoroughly",
            order=1,
            answer_mode="open",
            points=5,
            correct_answer="Ideal model answer text",
        )
        attempt = ExamAttempt.objects.create(user=self.student, exam=exam, status="submitted")
        question = exam.questions.first()
        ExamAnswer.objects.create(attempt=attempt, question=question, text_answer="My essay")

        response = self.client.get(reverse("exams:exam_result", args=[exam.slug, attempt.id]))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Ideal model answer text")

        exam.end_datetime = timezone.now() - timedelta(minutes=1)
        exam.save(update_fields=["end_datetime"])
        response = self.client.get(reverse("exams:exam_result", args=[exam.slug, attempt.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ideal model answer text")


class TeacherQuestionsBankViewTest(TestCase):
    def setUp(self):
        self.teacher = User.objects.create_user(
            username="questions_bank_teacher",
            email="questions_bank_teacher@example.com",
            password="StrongPass123!",
        )

        self.organization = Organization.objects.create(
            name="Questions Bank Org",
            org_type=OrganizationType.SCHOOL,
            owner=self.teacher,
            status="active",
            is_active=True,
        )
        _assign_user_to_org(self.teacher, self.organization, ProfileRole.TEACHER)

        self.exam = Exam.objects.create(
            author=self.teacher,
            title="Questions Bank Exam",
            exam_type="test",
            is_active=True,
        )

        self.questions = []
        for index in range(1, 15):
            q = ExamQuestion.objects.create(
                exam=self.exam,
                text=f"Alpha Question {index}",
                order=index,
                points=1,
                is_active=True,
            )
            self.questions.append(q)

        self.questions[2].is_active = False
        self.questions[2].save(update_fields=["is_active"])

        _login_with_org(self.client, self.teacher, self.organization)

    def test_questions_bank_supports_search_status_filter_and_pagination(self):
        response = self.client.get(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {"q": "Alpha", "status": "active", "page": 1},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "exams/teacher/teacher_questions_bank.html")
        self.assertEqual(response.context["status_filter"], "active")
        self.assertEqual(response.context["search_query"], "Alpha")
        self.assertEqual(response.context["page_obj"].paginator.num_pages, 2)
        self.assertContains(response, "questionSearchInput")
        self.assertContains(response, "js-open-question-form-modal")
        self.assertContains(response, 'id="questionFormModal"')
        self.assertContains(response, 'id="singleQuestionActionForm"')
        self.assertContains(response, "js-bulk-action-btn")
        self.assertContains(response, "disabled")

    def test_questions_bank_truncates_overlong_search_query(self):
        long_query = "A" * 320

        response = self.client.get(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {"q": long_query},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["search_query"], "A" * 240)
        self.assertContains(response, 'maxlength="240"', html=False)

    def test_questions_bank_drops_nested_exam_return_to_from_filter_forms(self):
        nested_return_to = (
            f"{reverse('exams:teacher_questions_bank', args=[self.exam.slug])}?return_to=/accounts/profile/"
        )

        response = self.client.get(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {"from_section": "my-exams", "return_to": nested_return_to},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["navigation_return_to"], "")
        self.assertNotContains(response, 'name="return_to"', html=False)

    def test_questions_bank_bulk_deactivate_selected(self):
        selected = [str(self.questions[0].id), str(self.questions[1].id)]
        response = self.client.post(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {
                "bulk_action": "deactivate",
                "selected_question_ids": selected,
                "status": "all",
                "q": "",
                "page": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(ExamQuestion.objects.get(id=self.questions[0].id).is_active)
        self.assertFalse(ExamQuestion.objects.get(id=self.questions[1].id).is_active)

    def test_questions_bank_single_question_activate_request_works(self):
        self.questions[2].is_active = False
        self.questions[2].save(update_fields=["is_active"])

        response = self.client.post(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {
                "bulk_action": "activate",
                "selected_question_ids": str(self.questions[2].id),
                "status": "all",
                "q": "",
                "sort": "newest",
                "page": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(ExamQuestion.objects.get(id=self.questions[2].id).is_active)

    def test_questions_bank_bulk_delete_selected(self):
        to_delete = [str(self.questions[3].id), str(self.questions[4].id)]
        response = self.client.post(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {
                "bulk_action": "delete",
                "selected_question_ids": to_delete,
                "status": "all",
                "q": "",
                "page": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(ExamQuestion.objects.filter(id=self.questions[3].id).exists())
        self.assertFalse(ExamQuestion.objects.filter(id=self.questions[4].id).exists())

    def test_questions_bank_filters_by_language(self):
        ExamQuestion.objects.filter(id__in=[self.questions[0].id, self.questions[1].id]).update(language="ru")

        response = self.client.get(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {"language": "ru"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["language_filter"], "ru")
        self.assertEqual(response.context["total_questions"], 2)
        object_ids = {q.id for q in response.context["page_obj"].object_list}
        self.assertEqual(object_ids, {self.questions[0].id, self.questions[1].id})
        self.assertContains(response, 'name="language"', html=False)
        self.assertContains(response, 'id="questionLangDeleteForm"', html=False)
        self.assertContains(response, "Bu dildəki bütün sualları sil")

    def test_questions_bank_deletes_selected_language_questions(self):
        ExamQuestion.objects.filter(id__in=[self.questions[0].id, self.questions[1].id]).update(language="ru")
        ExamQuestion.objects.filter(id=self.questions[2].id).update(language="en")

        response = self.client.post(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {
                "bulk_action": "delete_language",
                "language": "ru",
                "status": "all",
                "q": "",
                "sort": "newest",
                "page": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn("language=ru", response.url)
        self.assertFalse(self.exam.questions.filter(language="ru").exists())
        self.assertTrue(self.exam.questions.filter(language="az").exists())
        self.assertTrue(self.exam.questions.filter(language="en").exists())

    def test_questions_bank_deletes_all_exam_questions(self):
        ExamQuestion.objects.filter(id__in=[self.questions[0].id, self.questions[1].id]).update(language="ru")

        response = self.client.post(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {
                "bulk_action": "delete_all",
                "language": "ru",
                "status": "all",
                "q": "",
                "sort": "newest",
                "page": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.exam.questions.count(), 0)

    def test_questions_bank_search_matches_option_text_without_duplicates(self):
        ExamQuestionOption.objects.create(
            question=self.questions[0],
            label="A",
            text="Variant debounce açar sözü",
            is_correct=True,
        )
        ExamQuestionOption.objects.create(
            question=self.questions[0],
            label="B",
            text="Variant debounce açar sözü ikinci",
            is_correct=False,
        )

        response = self.client.get(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {"q": "debounce açar", "sort": "az"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["sort_filter"], "az")
        object_ids = [q.id for q in response.context["page_obj"].object_list]
        self.assertEqual(object_ids, [self.questions[0].id])

    def test_questions_bank_bulk_redirect_preserves_sort_filter(self):
        selected = [str(self.questions[0].id)]
        response = self.client.post(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {
                "bulk_action": "deactivate",
                "selected_question_ids": selected,
                "status": "active",
                "q": "Alpha",
                "sort": "za",
                "page": "2",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn("q=Alpha", response.url)
        self.assertIn("status=active", response.url)
        self.assertIn("sort=za", response.url)
        self.assertIn("page=2", response.url)

    def test_questions_bank_back_link_preserves_original_return_to(self):
        return_to = f"{reverse('accounts:profile')}?section=my-courses"
        response = self.client.get(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {"from_section": "my-courses", "return_to": return_to},
        )

        expected_query = urlencode({"from_section": "my-courses", "return_to": return_to})
        expected_href = f'{reverse("exams:teacher_exam_detail", args=[self.exam.slug])}?{expected_query}'

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, expected_href.replace("&", "&amp;"), html=False)

    def test_questions_bank_bulk_add_link_preserves_original_return_to(self):
        return_to = f"{reverse('accounts:profile')}?section=my-courses"
        response = self.client.get(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {"from_section": "my-courses", "return_to": return_to},
        )

        expected_query = urlencode({"from_section": "my-courses", "return_to": return_to})
        expected_href = f'{reverse("exams:test_question_bank", args=[self.exam.slug])}?{expected_query}'

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, expected_href.replace("&", "&amp;"), html=False)

    def test_test_question_bank_view_bank_link_preserves_original_return_to(self):
        return_to = f"{reverse('accounts:profile')}?section=my-courses"
        response = self.client.get(
            reverse("exams:test_question_bank", args=[self.exam.slug]),
            {"from_section": "my-courses", "return_to": return_to},
        )

        expected_query = urlencode({"from_section": "my-courses", "return_to": return_to})
        expected_href = f'{reverse("exams:teacher_questions_bank", args=[self.exam.slug])}?{expected_query}'

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, expected_href.replace("&", "&amp;"), html=False)

    def test_test_question_bank_back_link_preserves_original_return_to(self):
        return_to = f"{reverse('accounts:profile')}?section=my-courses"
        response = self.client.get(
            reverse("exams:test_question_bank", args=[self.exam.slug]),
            {"from_section": "my-courses", "return_to": return_to},
        )

        expected_query = urlencode({"from_section": "my-courses", "return_to": return_to})
        expected_href = f'{reverse("exams:teacher_exam_detail", args=[self.exam.slug])}?{expected_query}'

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, expected_href.replace("&", "&amp;"), html=False)

    def test_test_question_bank_shows_ai_generation_panel(self):
        response = self.client.get(reverse("exams:test_question_bank", args=[self.exam.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "AI ilə test sualı yarat")
        self.assertContains(response, 'data-ai-context="test"', html=False)
        self.assertContains(response, reverse("exams:ai_generate_question_bank", args=[self.exam.slug]), html=False)
        self.assertContains(response, "data-ai-file-input", html=False)
        self.assertContains(response, "data-bootstrap-select", html=False)
        self.assertContains(response, "Mətnin sonuna əlavə et")
        self.assertContains(response, "Mətn sahəsini əvəz et")

    def _end_question_import_text(self, count):
        return "\n\n".join(
            [
                "\n".join(
                    [
                        f"{index}. Import question {index}?",
                        f"Correct answer {index}",
                        f"Wrong answer {index}",
                        f"Other answer {index}",
                        f"Alternative answer {index}",
                        f"Extra answer {index}",
                        "END_QUESTION",
                    ]
                )
                for index in range(1, count + 1)
            ]
        )

    def test_test_question_bank_preview_checks_all_questions_and_adds_compact_save_fields(self):
        response = self.client.post(
            reverse("exams:test_question_bank", args=[self.exam.slug]),
            {
                "action": "preview",
                "raw_text": self._end_question_import_text(2),
            },
        )

        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        self.assertIn('id="selectedIndicesInput"', html)
        self.assertIn('id="pointsPayloadInput"', html)
        self.assertEqual(html.count('class="custom-checkbox qcheck"'), 2)
        self.assertEqual(html.count("q-card is-selected"), 2)

    def test_test_question_bank_downloads_problem_report_xlsx(self):
        raw_text = """
1. Same imported question?
A) Correct answer
B) Wrong answer
C) Other answer
D) Alternative answer
E) Extra answer
Cavab: A

2. Same imported question?
A) Correct answer
B) Wrong answer
C) Other answer
D) Alternative answer
E) Extra answer
Cavab: A
"""

        response = self.client.post(
            reverse("exams:test_question_bank", args=[self.exam.slug]),
            {
                "action": "download_report",
                "raw_text": raw_text,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("problem-report", response["Content-Disposition"])

        if response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            from openpyxl import load_workbook

            workbook = load_workbook(BytesIO(response.content), read_only=True)
            self.assertIn("Xülasə", workbook.sheetnames)
            self.assertIn("Problemlər", workbook.sheetnames)

            problems = workbook["Problemlər"]
            headers = [cell.value for cell in next(problems.iter_rows(min_row=1, max_row=1))]
            self.assertIn("Feedback", headers)
            rows_text = "\n".join(
                " ".join(str(cell.value or "") for cell in row)
                for row in problems.iter_rows(min_row=2, max_row=problems.max_row)
            )
        else:
            self.assertEqual(
                response["Content-Type"],
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
            from docx import Document

            document = Document(BytesIO(response.content))
            rows_text = "\n".join(paragraph.text for paragraph in document.paragraphs)
            self.assertIn("Feedback", rows_text)
        self.assertIn("Dublikat", rows_text)
        self.assertIn("Təkrarlanan suallardan birini silin", rows_text)

    def test_test_question_bank_preview_keeps_filter_counts_for_warning_types(self):
        raw_text = """
1. google.biz
A) Bu tip feyk vebsaytin hazirlanaraq hucum teskil edilmesi hansi hucum novune aiddir:
B) CyberSquaiting
C) BitSquatting
D) Paket manipulyasiyasi
E) BitManipulating
Cavab: A

2. google.biz
A) Bu tip feyk vebsaytin hazirlanaraq hucum teskil edilmesi hansi hucum novune aiddir:
B) CyberSquaiting
C) BitSquatting
D) Paket manipulyasiyasi
E) BitManipulating
Cavab: A
"""

        response = self.client.post(
            reverse("exams:test_question_bank", args=[self.exam.slug]),
            {
                "action": "preview",
                "raw_text": raw_text,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["category_counts"]["duplicates"], 2)
        self.assertEqual(response.context["category_counts"]["balance"], 2)
        self.assertEqual(response.context["category_counts"]["errors"], 2)

        html = response.content.decode()
        self.assertIn('data-filter="has-dup"', html)
        self.assertIn('data-filter="has-balance"', html)
        self.assertNotIn('filter-chip--duplicate" data-filter="has-dup" disabled', html)
        self.assertNotIn('filter-chip--balance" data-filter="has-balance" disabled', html)

    @override_settings(DATA_UPLOAD_MAX_NUMBER_FIELDS=1000)
    @patch("apps.exams.services.difficulty.schedule_ai_question_difficulty_warmup")
    def test_test_question_bank_saves_500_end_question_import_with_compact_payload(self, mock_schedule_warmup):
        question_count = 500
        points_payload = {str(index): "2" for index in range(1, question_count + 1)}
        points_payload["17"] = "5"

        response = self.client.post(
            reverse("exams:test_question_bank", args=[self.exam.slug]),
            {
                "action": "save",
                "raw_text": self._end_question_import_text(question_count),
                "selected_indices": ",".join(str(index) for index in range(1, question_count + 1)),
                "points_payload": json.dumps(points_payload),
                "random_question_count": str(question_count),
                "default_points": "2",
            },
        )

        self.assertEqual(response.status_code, 302)
        imported_questions = self.exam.questions.filter(order__gte=15).order_by("order")
        self.assertEqual(imported_questions.count(), question_count)
        self.assertEqual(ExamQuestionOption.objects.filter(question__in=imported_questions).count(), 2500)
        self.assertEqual(imported_questions.get(order=31).points, 5)
        first_imported = imported_questions.first()
        self.assertEqual(first_imported.options.get(label="A").text, "Correct answer 1")
        self.assertTrue(first_imported.options.get(label="A").is_correct)
        mock_schedule_warmup.assert_called_once()

    @patch("apps.exams.services.difficulty.schedule_ai_question_difficulty_warmup")
    def test_test_question_bank_compact_empty_selection_saves_no_questions(self, mock_schedule_warmup):
        response = self.client.post(
            reverse("exams:test_question_bank", args=[self.exam.slug]),
            {
                "action": "save",
                "raw_text": self._end_question_import_text(3),
                "selected_indices": "",
                "points_payload": "{}",
                "random_question_count": "3",
                "default_points": "2",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.exam.questions.count(), 14)
        mock_schedule_warmup.assert_called_once()

    @patch("apps.exams.services.difficulty.schedule_ai_question_difficulty_warmup")
    def test_test_question_bank_empty_compact_field_falls_back_to_legacy_selection(self, mock_schedule_warmup):
        response = self.client.post(
            reverse("exams:test_question_bank", args=[self.exam.slug]),
            {
                "action": "save",
                "raw_text": self._end_question_import_text(2),
                "selected_indices": "",
                "selected": ["2"],
                "points_payload": "",
                "points_2": "4",
                "random_question_count": "2",
                "default_points": "2",
            },
        )

        self.assertEqual(response.status_code, 302)
        imported_questions = self.exam.questions.filter(order__gte=15).order_by("order")
        self.assertEqual(imported_questions.count(), 1)
        self.assertEqual(imported_questions.first().text, "Import question 2?")
        self.assertEqual(imported_questions.first().points, 4)
        mock_schedule_warmup.assert_called_once()

    @patch("apps.exams.services.difficulty.schedule_ai_question_difficulty_warmup")
    def test_test_question_bank_saves_large_import_with_long_option_text(self, mock_schedule_warmup):
        long_option_text = " ".join(["Uzun variant mətni"] * 20)
        raw_text = "\n\n".join(
            [
                "\n".join(
                    [
                        f"{index}. Import sualı {index}?",
                        f"A) {long_option_text if index == 21 else f'Düzgün cavab {index}'}",
                        f"B) Yanlış cavab {index}",
                        f"C) Başqa cavab {index}",
                        f"D) Alternativ cavab {index}",
                        f"E) Əlavə cavab {index}",
                        "Cavab: A",
                    ]
                )
                for index in range(1, 301)
            ]
        )

        response = self.client.post(
            reverse("exams:test_question_bank", args=[self.exam.slug]),
            {
                "action": "save",
                "raw_text": raw_text,
                "selected": [str(index) for index in range(1, 301)],
                "random_question_count": "300",
                "default_points": "2",
            },
        )

        self.assertEqual(response.status_code, 302)
        imported_questions = self.exam.questions.filter(order__gte=15).order_by("order")
        self.assertEqual(imported_questions.count(), 300)
        self.assertEqual(ExamQuestionOption.objects.filter(question__in=imported_questions).count(), 1500)
        self.assertEqual(imported_questions.first().points, 2)
        self.assertGreater(
            len(imported_questions.get(order=35).options.get(label="A").text),
            255,
        )
        mock_schedule_warmup.assert_called_once()

    @patch("apps.exams.views.teacher.question_bank.generate_question_bank_text")
    def test_ai_generate_question_bank_passes_prompt_and_uploaded_source_to_service(self, mock_generate):
        mock_generate.return_value = {
            "ok": True,
            "text": "1. Sual\nA) A\nB) B\nC) C\nD) D\nE) E\nCavab: A",
            "question_count": 1,
        }

        response = self.client.post(
            reverse("exams:ai_generate_question_bank", args=[self.exam.slug]),
            {
                "prompt": "Python funksiyaları",
                "question_count": "1",
                "difficulty": "hard",
                "source_file": SimpleUploadedFile("lecture.txt", b"Funksiyalar movzusu", content_type="text/plain"),
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["ok"])
        mock_generate.assert_called_once()
        kwargs = mock_generate.call_args.kwargs
        self.assertEqual(kwargs["exam_type"], "test")
        self.assertEqual(kwargs["prompt_text"], "Python funksiyaları")
        self.assertEqual(kwargs["question_count"], "1")
        self.assertEqual(kwargs["difficulty"], "hard")
        self.assertIn("Funksiyalar movzusu", kwargs["source_text"])

    def test_questions_bank_bulk_delete_resequences_remaining_question_orders(self):
        response = self.client.post(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {
                "bulk_action": "delete",
                "selected_question_ids": str(self.questions[0].id),
                "status": "all",
                "q": "",
                "sort": "newest",
                "page": "1",
            },
        )

        self.assertEqual(response.status_code, 302)
        remaining_orders = list(self.exam.questions.order_by("order", "id").values_list("order", flat=True))
        self.assertEqual(remaining_orders, list(range(1, len(remaining_orders) + 1)))

    def test_questions_bank_bulk_redirect_preserves_return_navigation(self):
        response = self.client.post(
            reverse("exams:teacher_questions_bank", args=[self.exam.slug]),
            {
                "bulk_action": "activate",
                "selected_question_ids": str(self.questions[2].id),
                "status": "all",
                "q": "",
                "sort": "newest",
                "page": "1",
                "from_section": "my-courses",
                "return_to": f"{reverse('accounts:profile')}?section=my-courses",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn("from_section=my-courses", response.url)
        self.assertIn("return_to=%2Faccounts%2Fprofile%2F%3Fsection%3Dmy-courses", response.url)

    def test_create_question_bank_cancel_link_preserves_original_return_to(self):
        return_to = f"{reverse('accounts:profile')}?section=my-courses"
        response = self.client.get(
            reverse("exams:create_question_bank", args=[self.exam.slug]),
            {"from_section": "my-courses", "return_to": return_to},
        )

        expected_query = urlencode({"from_section": "my-courses", "return_to": return_to})
        expected_href = f'{reverse("exams:teacher_exam_detail", args=[self.exam.slug])}?{expected_query}'

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, expected_href.replace("&", "&amp;"), html=False)
        self.assertContains(response, 'class="btn-back-main"', html=False)

    def test_create_question_bank_shows_student_question_count_setting_for_written_exam(self):
        self.exam.exam_type = "written"
        self.exam.random_question_count = 6
        self.exam.save(update_fields=["exam_type", "random_question_count"])

        response = self.client.get(reverse("exams:create_question_bank", args=[self.exam.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Tələbəyə düşəcək sual sayı")
        self.assertRegex(response.content.decode(), r'name="random_question_count"[^>]*value="6"')

    def test_create_question_bank_shows_student_question_count_setting_for_coding_exam(self):
        self.exam.exam_type = "coding"
        self.exam.random_question_count = 3
        self.exam.save(update_fields=["exam_type", "random_question_count"])

        response = self.client.get(reverse("exams:create_question_bank", args=[self.exam.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Tələbəyə düşəcək sual sayı")
        self.assertContains(response, "Praktiki sual bankı")
        self.assertNotContains(response, "Yazılı sual bankı")
        self.assertRegex(response.content.decode(), r'name="random_question_count"[^>]*value="3"')

    def test_process_question_bank_success_redirect_preserves_return_navigation(self):
        return_to = f"{reverse('accounts:profile')}?section=my-courses"
        response = self.client.post(
            reverse("exams:process_question_bank", args=[self.exam.slug]),
            {
                "from_section": "my-courses",
                "return_to": return_to,
                "random_question_count": "0",
                "block_name_1": "Blok 1",
                "block_content_1": "1. Test sual",
                "block_time_1": "",
                "block_db_id_1": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn("from_section=my-courses", response.url)
        self.assertIn("return_to=%2Faccounts%2Fprofile%2F%3Fsection%3Dmy-courses", response.url)


class WrittenExamPaintInheritanceTest(TestCase):
    def setUp(self):
        self.teacher = User.objects.create_user(
            username="written_paint_teacher",
            email="written_paint_teacher@example.com",
            password="StrongPass123!",
        )
        self.student = User.objects.create_user(
            username="written_paint_student",
            email="written_paint_student@example.com",
            password="StrongPass123!",
        )

        self.organization = Organization.objects.create(
            name="Written Paint Org",
            org_type=OrganizationType.SCHOOL,
            owner=self.teacher,
            status="active",
            is_active=True,
        )
        _assign_user_to_org(self.teacher, self.organization, ProfileRole.TEACHER)
        _assign_user_to_org(self.student, self.organization, ProfileRole.STUDENT)

        self.exam = Exam.objects.create(
            author=self.teacher,
            title="Written Paint Exam",
            exam_type="written",
            is_active=True,
            enable_paint=True,
        )
        self.block = QuestionBlock.objects.create(
            exam=self.exam,
            name="Bölmə 1",
            order=1,
            enable_paint=True,
        )
        self.hidden_question = ExamQuestion.objects.create(
            exam=self.exam,
            block=self.block,
            text="Paint gizli qalsın",
            order=1,
            answer_mode="single",
            disable_paint=True,
        )
        self.visible_question = ExamQuestion.objects.create(
            exam=self.exam,
            block=self.block,
            text="Paint görünsün",
            order=2,
            answer_mode="single",
        )
        self.attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.exam,
            status="in_progress",
            attempt_number=1,
        )
        self.hidden_answer = ExamAnswer.objects.create(
            attempt=self.attempt,
            question=self.hidden_question,
            has_paint=True,
            paint_data_url="stale",
        )
        ExamAnswer.objects.create(attempt=self.attempt, question=self.visible_question)

    def test_take_exam_hides_paint_for_question_with_explicit_disable(self):
        _login_with_org(self.client, self.student, self.organization)

        response = self.client.get(reverse("exams:take_exam", args=[self.exam.slug, self.attempt.id]))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, f'name="paint_enabled_{self.hidden_question.id}"', html=False)
        self.assertContains(response, f'name="paint_enabled_{self.visible_question.id}"', html=False)

    def test_take_exam_post_clears_hidden_question_paint_even_if_payload_forces_it(self):
        _login_with_org(self.client, self.student, self.organization)

        response = self.client.post(
            reverse("exams:take_exam", args=[self.exam.slug, self.attempt.id]),
            {
                f"q_{self.hidden_question.id}": "Cavab",
                f"paint_enabled_{self.hidden_question.id}": "1",
                f"paint_data_{self.hidden_question.id}": "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAAB",
                f"q_{self.visible_question.id}": "Digər cavab",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.hidden_answer.refresh_from_db()
        self.assertFalse(self.hidden_answer.has_paint)
        self.assertFalse(bool(self.hidden_answer.paint_data_url))

    def test_edit_question_unchecked_paint_creates_question_level_disable(self):
        _login_with_org(self.client, self.teacher, self.organization)

        response = self.client.post(
            reverse("exams:edit_exam_question", args=[self.exam.slug, self.visible_question.id]),
            {
                "text": self.visible_question.text,
                "block": str(self.block.id),
                "time_limit_seconds": "",
                "correct_answer": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.visible_question.refresh_from_db()
        self.assertTrue(self.visible_question.disable_paint)
        self.assertFalse(self.visible_question.paint_enabled_effective)

    def test_edit_question_form_shows_checked_paint_checkbox_when_state_comes_from_block(self):
        _login_with_org(self.client, self.teacher, self.organization)

        response = self.client.get(
            reverse("exams:edit_exam_question", args=[self.exam.slug, self.visible_question.id]) + "?modal=1",
            {"modal": "1"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertRegex(response.content.decode(), r'name="enable_paint"[^>]*checked')
        self.assertContains(response, "Mövzu bloku ayarı")
        self.assertContains(response, "data-bootstrap-select", html=False)
        self.assertNotContains(response, "Ümumi", html=False)

    def test_edit_question_form_shows_question_source_when_question_override_disables_paint(self):
        _login_with_org(self.client, self.teacher, self.organization)

        response = self.client.get(
            reverse("exams:edit_exam_question", args=[self.exam.slug, self.hidden_question.id]) + "?modal=1",
            {"modal": "1"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotRegex(response.content.decode(), r'name="enable_paint"[^>]*checked')
        self.assertContains(response, "Bu sualın öz ayarı")

    def test_add_written_question_without_blocks_creates_default_block_and_assigns_question(self):
        _login_with_org(self.client, self.teacher, self.organization)
        exam = Exam.objects.create(
            author=self.teacher,
            title="Written Without Blocks",
            exam_type="written",
            is_active=True,
            enable_paint=True,
        )

        response = self.client.post(
            reverse("exams:add_exam_question", args=[exam.slug]) + "?modal=1",
            {
                "modal": "1",
                "text": "Default bloka düşən sual",
                "time_limit_seconds": "",
                "correct_answer": "",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)
        block = exam.question_blocks.get()
        question = exam.questions.get(text="Default bloka düşən sual")
        self.assertEqual(block.name, "Bölmə 1")
        self.assertEqual(question.block, block)

    def test_add_written_question_requires_block_when_blocks_exist(self):
        _login_with_org(self.client, self.teacher, self.organization)

        response = self.client.post(
            reverse("exams:add_exam_question", args=[self.exam.slug]) + "?modal=1",
            {
                "modal": "1",
                "text": "Bloksuz yazılı sual",
                "time_limit_seconds": "",
                "correct_answer": "",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["success"], False)
        self.assertFalse(self.exam.questions.filter(text="Bloksuz yazılı sual").exists())

    def test_process_question_bank_preserves_question_override_and_saves_block_paint(self):
        _login_with_org(self.client, self.teacher, self.organization)

        response = self.client.post(
            reverse("exams:process_question_bank", args=[self.exam.slug]),
            {
                "random_question_count": "0",
                "block_name_1": self.block.name,
                "block_time_1": "",
                "block_enable_paint_1": "on",
                "block_content_1": "1. Yenilənmiş 1\n2) Yenilənmiş 2",
                "block_db_id_1": str(self.block.id),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.block.refresh_from_db()
        self.hidden_question.refresh_from_db()
        self.visible_question.refresh_from_db()

        self.assertTrue(self.block.enable_paint)
        self.assertEqual(self.hidden_question.text, "Yenilənmiş 1")
        self.assertTrue(self.hidden_question.disable_paint)
        self.assertFalse(self.hidden_question.paint_enabled_effective)
        self.assertEqual(self.visible_question.text, "Yenilənmiş 2")

    def test_create_question_bank_shows_ai_panel_per_written_block(self):
        _login_with_org(self.client, self.teacher, self.organization)

        response = self.client.get(reverse("exams:create_question_bank", args=[self.exam.slug]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "AI ilə bu blok üçün sual yarat")
        self.assertContains(response, 'data-ai-context="written"', html=False)
        self.assertContains(response, "data-written-question-textarea", html=False)
        self.assertContains(response, "data-ai-file-input", html=False)
        self.assertContains(response, "data-bootstrap-select", html=False)
        self.assertContains(response, "Mətnin sonuna əlavə et")
        self.assertContains(response, "Mətn sahəsini əvəz et")


class SupervisionTeacherApiTest(TestCase):
    def setUp(self):
        self.teacher = User.objects.create_user(
            username="supervision_api_teacher",
            email="supervision_api_teacher@example.com",
            password="StrongPass123!",
        )
        self.student = User.objects.create_user(
            username="supervision_api_student",
            email="supervision_api_student@example.com",
            password="StrongPass123!",
        )
        self.org = Organization.objects.create(
            name="Supervision API Org",
            org_type=OrganizationType.SCHOOL,
            owner=self.teacher,
            status="active",
            is_active=True,
        )
        _assign_user_to_org(self.teacher, self.org, ProfileRole.TEACHER)
        _assign_user_to_org(self.student, self.org, ProfileRole.STUDENT)
        self.exam = Exam.objects.create(
            author=self.teacher,
            organization=self.org,
            title="Timed Supervision Exam",
            exam_type="test",
            is_active=True,
            total_duration_minutes=60,
        )
        ExamSupervisionConfig.objects.create(
            exam=self.exam,
            enabled=True,
            recovery_policy="teacher_controlled",
        )
        self.attempt = ExamAttempt.objects.create(
            user=self.student,
            exam=self.exam,
            status="in_progress",
            attempt_number=1,
            supervision_status="locked",
            supervision_violation_count=1,
        )
        _login_with_org(self.client, self.teacher, self.org)

    def test_teacher_resume_api_rejects_attempt_after_exam_duration(self):
        self.attempt.started_at = timezone.now() - timedelta(minutes=61)
        self.attempt.save(update_fields=["started_at"])

        response = self.client.post(
            reverse("exams:supervision_resume", args=[self.attempt.id]),
            data="{}",
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 403)
        self.attempt.refresh_from_db()
        self.assertEqual(self.attempt.status, "expired")
        self.assertEqual(self.attempt.supervision_status, "locked")
        self.assertIsNotNone(self.attempt.finished_at)

    def test_student_status_api_keeps_manual_lock_visible_after_exam_duration(self):
        self.attempt.supervision_manual_lock = True
        self.attempt.started_at = timezone.now() - timedelta(minutes=61)
        self.attempt.save(update_fields=["supervision_manual_lock", "started_at"])
        self.client.force_login(self.student)

        response = self.client.get(reverse("exams:supervision_status_api", args=[self.attempt.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["manual_lock"])
        self.assertFalse(payload["is_finished"])
        self.assertEqual(payload["supervision_status"], "locked")
        self.attempt.refresh_from_db()
        self.assertEqual(self.attempt.status, "in_progress")

    def test_student_status_api_reports_manual_lock_without_supervision_config(self):
        self.exam.supervision_config.delete()
        self.attempt.supervision_manual_lock = True
        self.attempt.save(update_fields=["supervision_manual_lock"])
        self.client.force_login(self.student)

        response = self.client.get(reverse("exams:supervision_status_api", args=[self.attempt.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertFalse(payload["supervised"])
        self.assertTrue(payload["manual_lock"])
        self.assertEqual(payload["supervision_status"], "locked")

    def test_unsupervised_student_exam_loads_manual_lock_listener(self):
        self.exam.supervision_config.delete()
        self.attempt.supervision_status = "active"
        self.attempt.supervision_manual_lock = False
        self.attempt.save(update_fields=["supervision_status", "supervision_manual_lock"])
        question = ExamQuestion.objects.create(exam=self.exam, text="Manual lock listener?", order=1)
        ExamQuestionOption.objects.create(question=question, text="Yes", is_correct=True)
        ExamQuestionOption.objects.create(question=question, text="No", is_correct=False)
        ExamAnswer.objects.create(attempt=self.attempt, question=question)
        _login_with_org(self.client, self.student, self.org)

        response = self.client.get(reverse("exams:take_exam", args=[self.exam.slug, self.attempt.id]))

        self.assertEqual(response.status_code, 200)
        # Refaktor 2026-07-02: monolit exam_supervision.js → exam_supervision/ paketi.
        self.assertContains(response, "exam_supervision.entry.js")
        self.assertContains(response, "supervised: false")

    @override_settings(EXAM_SUPERVISION_ENABLED=False)
    def test_student_exam_omits_supervision_listener_when_feature_disabled(self):
        self.attempt.supervision_status = "active"
        self.attempt.supervision_manual_lock = False
        self.attempt.save(update_fields=["supervision_status", "supervision_manual_lock"])
        question = ExamQuestion.objects.create(exam=self.exam, text="Feature disabled?", order=1)
        ExamQuestionOption.objects.create(question=question, text="Yes", is_correct=True)
        ExamQuestionOption.objects.create(question=question, text="No", is_correct=False)
        ExamAnswer.objects.create(attempt=self.attempt, question=question)
        _login_with_org(self.client, self.student, self.org)

        response = self.client.get(reverse("exams:take_exam", args=[self.exam.slug, self.attempt.id]))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "exam_supervision.js")
        self.assertNotContains(response, "supervised: false")

    def test_supervision_monitor_shows_exam_without_config_or_violations(self):
        fresh_exam = Exam.objects.create(
            author=self.teacher,
            organization=self.org,
            title="Fresh Monitor Exam",
            exam_type="test",
            is_active=True,
        )

        response = self.client.get(reverse("exams:supervision_monitor"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, fresh_exam.title)
        self.assertContains(response, "Bu imtahana hələ heç kim qoşulmayıb.")


# ════════════════════════════════════════════════════════════════════════════
# Tenant Isolation: Null Organization Edge-Case Tests
# ════════════════════════════════════════════════════════════════════════════


class ExamOrganizationRequiredTest(TestCase):
    """
    Tenant isolation: Exam cannot be created without an organization.
    These tests cover the null-organization edge cases for tenant boundary enforcement.
    """

    def setUp(self):
        self.teacher = User.objects.create_user(
            username="exam_org_req_teacher",
            email="exam_org_req_teacher@example.com",
            password="StrongPass123!",
        )
        self.org = Organization.objects.create(
            name="Exam OrgRequired Test Org",
            org_type=OrganizationType.SCHOOL,
            owner=self.teacher,
            status="active",
            is_active=True,
        )
        _assign_user_to_org(self.teacher, self.org, ProfileRole.TEACHER)

    def test_exam_model_raises_validation_error_without_organization(self):
        """Exam.save() raises ValidationError when organization cannot be resolved."""
        from django.core.exceptions import ValidationError

        # Create a user with NO organization profile
        orphan_teacher = User.objects.create_user(
            username="orphan_exam_teacher",
            email="orphan_exam_teacher@example.com",
            password="StrongPass123!",
        )

        with self.assertRaises(ValidationError):
            Exam.objects.create(
                author=orphan_teacher,
                title="Orphan Exam",
                exam_type="test",
                # organization intentionally omitted and profile has no org
            )

    def test_exam_model_auto_assigns_organization_from_author_profile(self):
        """Exam.save() auto-assigns organization from author's profile when not set explicitly."""
        exam = Exam.objects.create(
            author=self.teacher,
            title="Auto Org Exam",
            exam_type="test",
            # organization intentionally omitted
        )
        self.assertEqual(exam.organization, self.org)

    def test_exam_model_auto_assigns_organization_from_course(self):
        """Exam.save() auto-assigns organization from the linked course's organization."""
        from apps.courses.models import Course

        course = Course.objects.create(
            owner=self.teacher,
            title="Source Course for Exam",
            status="published",
            organization=self.org,
        )
        exam = Exam.objects.create(
            author=self.teacher,
            course=course,
            title="Course-Linked Exam",
            exam_type="test",
            # organization intentionally omitted — should be derived from course
        )
        self.assertEqual(exam.organization, self.org)

    def test_create_exam_view_raises_permission_denied_without_active_organization(self):
        """createAndEditExamView raises PermissionDenied when request has no organization."""
        from apps.exams.views.teacher.exams import createAndEditExamView

        request = RequestFactory().post(
            reverse("exams:create_exam"),
            {
                "title": "No Org Exam",
                "description": "Should not be created",
                "exam_type": "test",
                "is_active": "on",
            },
        )
        request.user = self.teacher
        request.organization = None
        request.org_memberships = []
        request.org_permissions = []

        with self.assertRaises(PermissionDenied):
            createAndEditExamView(request)

        self.assertFalse(Exam.objects.filter(title="No Org Exam").exists())

    def test_exam_with_explicit_organization_is_created_successfully(self):
        """Exam explicitly bound to an organization is created without errors."""
        exam = Exam.objects.create(
            author=self.teacher,
            title="Explicitly Bound Exam",
            exam_type="test",
            organization=self.org,
        )
        self.assertEqual(exam.organization, self.org)
        self.assertIsNotNone(exam.pk)
