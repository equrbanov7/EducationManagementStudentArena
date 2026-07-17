"""İmtahandan uzaqlaşdırılmış tələbənin nəticəsi — səbəb hər yerdə görünsün.

- Müəllim Excel ixracında «Uzaqlaşdırma səbəbi» sütunu.
- İmtahan mərkəzi statistika sətrində `removed` bayrağı + səbəb (qırmızı UI).
Tələbənin öz nəticə səhifəsi onsuz da qırmızı göstərir (mövcud davranış).
"""

from django.contrib.auth import get_user_model
from django.test import TestCase

from apps.exams.models import Exam, ExamAttempt, SupervisionIncident
from apps.exams.services.supervision import attach_attempt_interventions
from apps.organizations.models import Organization
from core.constants import OrganizationType

User = get_user_model()

REASON = "Telefondan istifadə — imtahandan uzaqlaşdırıldı"


class RemovedStudentResultsTests(TestCase):
    def setUp(self):
        self.teacher = User.objects.create_user("rmv_teacher", "rmv_t@test.az", "StrongPass123!")
        self.student = User.objects.create_user("rmv_student", "rmv_s@test.az", "StrongPass123!")
        self.org = Organization.objects.create(
            name="RMV Org",
            org_type=OrganizationType.UNIVERSITY,
            owner=self.teacher,
            status="active",
            is_active=True,
        )
        self.exam = Exam.objects.create(
            title="RMV Exam", author=self.teacher, organization=self.org, exam_type="test", is_active=True
        )
        self.attempt = ExamAttempt.objects.create(
            user=self.student, exam=self.exam, status="submitted", supervision_status="removed"
        )
        SupervisionIncident.objects.create(
            organization=self.org,
            exam=self.exam,
            attempt=self.attempt,
            student=self.student,
            event_type="auto_locked",
            teacher_action="teacher_force_stopped",
            metadata={"display_reason": REASON},
        )

    def test_intervention_marks_removed_with_reason(self):
        attach_attempt_interventions([self.attempt])
        intervention = self.attempt.exam_intervention
        self.assertTrue(intervention["is_terminal"])
        self.assertEqual(intervention["reason"], REASON)

    def test_exam_center_stats_row_flags_removed(self):
        from apps.exams.views.exam_center.statistics import _row

        attach_attempt_interventions([self.attempt])
        row = _row(self.attempt)
        self.assertTrue(row["removed"])
        self.assertEqual(row["removal_reason"], REASON)

    def test_teacher_xlsx_export_includes_removal_reason(self):
        from io import BytesIO

        from openpyxl import load_workbook

        from apps.exams.views.teacher.results._export_builder import build_exam_results_xlsx_export

        result = build_exam_results_xlsx_export(self.exam, [self.attempt])
        # (filename, content_type, bytes)
        content = result[-1]
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertIn("Uzaqlaşdırma səbəbi", headers)
        col = headers.index("Uzaqlaşdırma səbəbi") + 1
        self.assertEqual(ws.cell(row=2, column=col).value, REASON)
