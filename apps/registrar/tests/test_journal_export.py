"""Tests for the journal xlsx export (U14)."""

import datetime
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import Client, TestCase
from django.urls import reverse

from apps.organizations.models import AcademicPeriod, Membership, Organization, OrgUnit
from apps.registrar import finals, gradebook, services
from apps.registrar.models import Curriculum, CurriculumSubject, LessonKind, Program, StudentAcademicRecord, Subject
from core.constants import AcademicPeriodType, OrganizationType, OrgUnitType
from core.rls import bypass_rls

User = get_user_model()


class JournalExportTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.owner = User.objects.create_user("je_owner", "je_owner@qku.edu.az", "pw")
        with bypass_rls():
            cls.org = Organization.objects.create(
                name="JE Univ",
                slug="je-univ",
                org_type=OrganizationType.UNIVERSITY,
                owner=cls.owner,
                status="active",
                is_active=True,
            )
            cls.faculty = OrgUnit.objects.create(
                organization=cls.org,
                name="Fakültə",
                slug="je-faculty",
                unit_type=OrgUnitType.FACULTY,
            )
            cls.chair_unit = OrgUnit.objects.create(
                organization=cls.org,
                name="Kafedra",
                slug="je-chair",
                unit_type=OrgUnitType.CHAIR,
                parent=cls.faculty,
            )
            cls.group = OrgUnit.objects.create(
                organization=cls.org,
                name="G1",
                slug="je-g1",
                unit_type=OrgUnitType.GROUP,
                parent=cls.chair_unit,
            )
            cls.period = AcademicPeriod.objects.create(
                organization=cls.org,
                name="2024/2025 Payız",
                period_type=AcademicPeriodType.SEMESTER,
                academic_year="2024/2025",
                start_date="2024-09-01",
                end_date="2025-01-31",
                is_current=True,
            )
            cls.program = Program.objects.create(organization=cls.org, code="CS", name="Kompüter elmləri")
            cls.curriculum = Curriculum.objects.create(organization=cls.org, program=cls.program, admission_year=2024)
            cls.subject = Subject.objects.create(organization=cls.org, code="CS101", name="Proqramlaşdırma", ects=6)
            CurriculumSubject.objects.create(
                organization=cls.org, curriculum=cls.curriculum, subject=cls.subject, semester_number=1
            )
            cls.teacher = User.objects.create_user("je_teacher", "je_teacher@qku.edu.az", "pw")
            cls.chair = User.objects.create_user("je_chair", "je_chair@qku.edu.az", "pw")
            cls.student = User.objects.create_user(
                "je_student", "je_student@qku.edu.az", "pw", first_name="Əli", last_name="Vəliyev"
            )
            Membership.objects.create(
                user=cls.teacher,
                organization=cls.org,
                role=cls.org.roles.get(name="teacher"),
                is_primary=True,
                is_active=True,
            )
            Membership.objects.create(
                user=cls.chair,
                organization=cls.org,
                role=cls.org.roles.get(name="chair_head"),
                scope_unit=cls.chair_unit,
                is_primary=True,
                is_active=True,
            )
            Membership.objects.create(
                user=cls.student,
                organization=cls.org,
                role=cls.org.roles.get(name="student"),
                is_primary=True,
                is_active=True,
            )
            record = StudentAcademicRecord.objects.create(
                organization=cls.org,
                student=cls.student,
                program=cls.program,
                curriculum=cls.curriculum,
                group=cls.group,
                admission_year=2024,
            )
            services.enroll_mandatory_subjects(record=record, period=cls.period, semester_number=1)
            cls.offering = cls.student.enrollments.get().offering
            cls.offering.instructor = cls.teacher
            cls.offering.save(update_fields=["instructor"])
            cls.enrollment = cls.offering.enrollments.get()
            lesson = gradebook.create_lesson(
                allow_past=True, offering=cls.offering, date=datetime.date(2024, 10, 1), kind=LessonKind.SEMINAR
            )
            gradebook.save_marks(
                enforce_day=False,
                offering=cls.offering,
                entries=[
                    {"lesson_id": lesson.id, "enrollment_id": cls.enrollment.id, "status": "present", "score": 40}
                ],
                by_user=cls.teacher,
            )
            finals.set_exam_score(enrollment=cls.enrollment, score=45, by_user=cls.teacher)

    def _client(self, user):
        client = Client()
        client.force_login(user)
        session = client.session
        session["active_organization"] = self.org.slug
        session.save()
        return client

    def _url(self):
        return reverse("registrar:journal_xlsx", args=[self.offering.id])

    def test_teacher_downloads_xlsx_with_content(self):
        from openpyxl import load_workbook

        resp = self._client(self.teacher).get(self._url())
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        self.assertIn("jurnal-CS101.xlsx", resp["Content-Disposition"])
        workbook = load_workbook(BytesIO(resp.content))
        self.assertEqual(len(workbook.sheetnames), 2)
        texts = "\n".join(str(c.value) for ws in workbook.worksheets for row in ws.iter_rows() for c in row if c.value)
        self.assertIn("Əli Vəliyev", texts)
        self.assertIn("CS101", texts)

    def test_student_cannot_export(self):
        resp = self._client(self.student).get(self._url())
        self.assertEqual(resp.status_code, 404)

    def test_chair_cannot_export(self):
        """Təsdiq zənciri ləğv edildi — kafedra müdirinin «rəyçi» ixracı da yoxdur."""
        self.assertEqual(self._client(self.chair).get(self._url()).status_code, 404)

    def test_export_is_audited(self):
        from apps.audit.models import AuditLog

        self._client(self.teacher).get(self._url())
        with bypass_rls():
            self.assertTrue(
                AuditLog.objects.filter(
                    resource_type="registrar.journal_export", resource_id=str(self.offering.pk)
                ).exists()
            )


class JournalTenantIsolationTest(JournalExportTest):
    """P0 reqressiya: jurnal səthləri AKTİV TƏŞKİLATA bağlı olmalıdır.

    Əvvəl ``get_object_or_404(CourseOffering, pk=...)`` org filtri OLMADAN
    işləyirdi — tenant sərhədi tamamilə RLS-ə qalırdı (non-Postgres backend-də
    no-op, ``rolbypassrls`` daşıyan DB rolunda isə mühərrik səviyyəsində keçilir).
    """

    def _foreign_actor_client(self):
        """Başqa təşkilatın sahibi + öz org kontekstində sessiya."""
        with bypass_rls():
            other_owner = User.objects.create_user("je_other_owner", "je_other@qku.edu.az", "pw")
            other_org = Organization.objects.create(
                name="Other Univ",
                slug="je-other-univ",
                org_type=OrganizationType.UNIVERSITY,
                owner=other_owner,
                status="active",
                is_active=True,
            )
            Membership.objects.create(
                user=other_owner,
                organization=other_org,
                role=other_org.roles.order_by("-level").first(),
                is_primary=True,
                is_active=True,
            )

        client = Client()
        client.force_login(other_owner)
        session = client.session
        session["active_organization"] = other_org.slug
        session.save()
        return client

    def test_foreign_org_cannot_open_journal_detail(self):
        response = self._foreign_actor_client().get(reverse("registrar:journal_detail", args=[self.offering.id]))
        self.assertEqual(response.status_code, 404)

    def test_foreign_org_cannot_export_journal_xlsx(self):
        response = self._foreign_actor_client().get(self._url())
        self.assertEqual(response.status_code, 404)
