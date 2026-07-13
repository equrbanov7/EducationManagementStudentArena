"""Jurnal xlsx ixracı (U14) — müəllim/rəhbərlik üçün rəsmi cədvəl.

Bir offering-in tam jurnalını iki vərəqdə ixrac edir (openpyxl — artıq base
asılılıqdır): «Davamiyyət» (dərs-bə-dərs iə/qb + bal + qayıb + giriş balı) və
«Yekun» (giriş + imtahan + ümumi + hərf + nəticə). Sənəd anında yaradılır,
saxlanmır; hər ixrac audit-ə yazılır (view qatında).

Dizayn (2026-07-13): başlıq zolağı (birləşmiş), rəng-kodlu davamiyyət
(iştirak = yaşıl, qayıb = qırmızı), tam sərhədlər, növbələnən sətir kölgəsi,
əfsanə və imza sətri — çap/təhvil üçün səliqəli görünüş.
"""

from __future__ import annotations

from io import BytesIO

from django.utils import timezone
from django.utils.translation import pgettext


def _(text):
    return pgettext("registrar.pdf", text)  # PDF ilə eyni kontekst — etiketlər üst-üstə düşür


# WCU rəng palitrası (UI --ems-* tokenləri ilə uyğun).
_NAVY = "0F2A5C"  # başlıq zolağı
_PRIMARY = "2563EB"  # sütun başlıqları
_HEADER_FONT = "FFFFFF"
_PRESENT_FILL = "DCFCE7"  # yaşıl-100 (iştirak)
_PRESENT_FONT = "15803D"
_ABSENT_FILL = "FEE2E2"  # qırmızı-100 (qayıb)
_ABSENT_FONT = "B91C1C"
_ZEBRA_FILL = "F1F5F9"  # növbələnən sətir
_SUBTLE = "64748B"


def build_journal_workbook(*, offering, journal, finals) -> bytes:
    """Offering jurnalı → xlsx bytes. ``journal``/`finals`` — mövcud servis
    strukturları (gradebook.get_offering_journal / finals.get_offering_results)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy_fill = PatternFill("solid", fgColor=_NAVY)
    header_fill = PatternFill("solid", fgColor=_PRIMARY)
    header_font = Font(name="Calibri", color=_HEADER_FONT, bold=True, size=11)
    title_font = Font(name="Calibri", color=_HEADER_FONT, bold=True, size=15)
    sub_font = Font(name="Calibri", color=_HEADER_FONT, size=10)
    zebra = PatternFill("solid", fgColor=_ZEBRA_FILL)
    present_fill = PatternFill("solid", fgColor=_PRESENT_FILL)
    present_font = Font(color=_PRESENT_FONT, bold=True)
    absent_fill = PatternFill("solid", fgColor=_ABSENT_FILL)
    absent_font = Font(color=_ABSENT_FONT, bold=True)
    barred_font = Font(color=_ABSENT_FONT, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    workbook = Workbook()

    # ── Vərəq 1: Davamiyyət və ballar ────────────────────────────────────────
    ws = workbook.active
    ws.title = pgettext("registrar.journal", "Davamiyyət və ballar")[:31]
    ws.sheet_view.showGridLines = False

    lessons = journal["lessons"]
    total_cols = 1 + len(lessons) + 2  # tələbə + dərslər + qayıb + giriş balı
    last_col = get_column_letter(total_cols)

    # Başlıq zolağı (birləşmiş): fənn + qrup + semestr.
    org_name = getattr(getattr(offering, "organization", None), "name", "") or ""
    ws.merge_cells(f"A1:{last_col}1")
    tcell = ws.cell(row=1, column=1, value=f"{offering.subject.code} — {offering.subject.name}")
    tcell.fill = navy_fill
    tcell.font = title_font
    tcell.alignment = center
    ws.row_dimensions[1].height = 28

    sub_parts = [
        p
        for p in [
            org_name,
            getattr(getattr(offering, "group", None), "name", ""),
            getattr(getattr(offering, "period", None), "name", ""),
        ]
        if p
    ]
    ws.merge_cells(f"A2:{last_col}2")
    scell = ws.cell(row=2, column=1, value="  ·  ".join(sub_parts))
    scell.fill = navy_fill
    scell.font = sub_font
    scell.alignment = center
    ws.row_dimensions[2].height = 18
    ws.append([])  # boş sətir

    # Sütun başlıqları.
    head = [pgettext("registrar.journal", "Soyad, ad")]
    head += [f"{i + 1}\n{lesson.date.strftime('%d.%m')}" for i, lesson in enumerate(lessons)]
    head += [pgettext("registrar.journal", "Qayıb (q/b)"), pgettext("registrar.journal", "Giriş balı")]
    ws.append(head)
    header_row = ws.max_row
    for col in range(1, len(head) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[header_row].height = 30

    for r_idx, row in enumerate(journal["rows"]):
        values = [row["student"].get_full_name() or row["student"].username]
        for cell_data in row["cells"]:
            mark = cell_data["mark"]
            if mark is None:
                values.append("—")
            elif mark.status == "absent":
                values.append("q/b")
            elif mark.score is not None:
                values.append(float(mark.score))
            else:
                values.append("i/e")
        values.append(int(row.get("absence_count", 0)))
        values.append(float(row["entry_score"]))
        ws.append(values)
        excel_row = ws.max_row
        row_fill = zebra if r_idx % 2 else None
        for col in range(1, len(values) + 1):
            c = ws.cell(row=excel_row, column=col)
            c.border = border
            c.alignment = left if col == 1 else center
            if row_fill and col > 1:
                c.fill = row_fill
        # Davamiyyət xanalarını rəng-kodla.
        for idx, cell_data in enumerate(row["cells"], start=2):
            mark = cell_data["mark"]
            if mark is None:
                continue
            c = ws.cell(row=excel_row, column=idx)
            if mark.status == "absent":
                c.fill = absent_fill
                c.font = absent_font
            elif mark.score is None:
                c.fill = present_fill
                c.font = present_font
        if row["barred"]:
            ws.cell(row=excel_row, column=1).font = barred_font

    ws.column_dimensions["A"].width = 30
    for col in range(2, len(head) - 1):
        ws.column_dimensions[get_column_letter(col)].width = 7
    ws.column_dimensions[get_column_letter(total_cols - 1)].width = 12
    ws.column_dimensions[get_column_letter(total_cols)].width = 12
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    # Əfsanə.
    ws.append([])
    legend = ws.cell(
        row=ws.max_row + 1,
        column=1,
        value=_("Əfsanə:  i/e — iştirak edib   ·   q/b — qayıb   ·   Qayıb sütunu q/b sayını göstərir"),
    )
    legend.font = Font(color=_SUBTLE, italic=True, size=10)
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=total_cols)

    # ── Vərəq 2: Yekun nəticə ────────────────────────────────────────────────
    ws2 = workbook.create_sheet(pgettext("registrar.finals", "Yekun nəticə")[:31])
    ws2.sheet_view.showGridLines = False
    head2 = [
        _("Tələbə"),
        _("Giriş"),
        _("İmtahan"),
        _("Yekun"),
        _("Hərf"),
        pgettext("registrar.finals", "Bonus"),
        _("Nəticə"),
        pgettext("registrar.finals", "Təkrar imtahan"),
        pgettext("registrar.finals", "Rəy"),
    ]
    last2 = get_column_letter(len(head2))
    ws2.merge_cells(f"A1:{last2}1")
    t2 = ws2.cell(row=1, column=1, value=f"{offering.subject.code} — {_('Yekun nəticə')}")
    t2.fill = navy_fill
    t2.font = title_font
    t2.alignment = center
    ws2.row_dimensions[1].height = 26
    ws2.append([])
    ws2.append(head2)
    head2_row = ws2.max_row
    for col in range(1, len(head2) + 1):
        cell = ws2.cell(row=head2_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for r_idx, row in enumerate(finals["rows"]):
        result = row["result"]
        if result["barred"] or result["failed"]:
            outcome = _("Kəsildi")
        elif result["passed"]:
            outcome = _("Keçdi")
        else:
            outcome = _("Davam edir")
        resit = result.get("resit")
        ws2.append(
            [
                row["student"].get_full_name() or row["student"].username,
                float(result["entry_score"]),
                float(result["exam_score"]) if result["exam_score"] is not None else "",
                float(result["total"]) if result["graded"] else "",
                result["letter"] if result["graded"] else "",
                float(result["bonus"]) if result["bonus"] else "",
                str(outcome),
                float(resit.resit_score) if resit is not None and resit.resit_score is not None else "",
                result.get("comment", ""),
            ]
        )
        excel_row = ws2.max_row
        row_fill = zebra if r_idx % 2 else None
        for col in range(1, len(head2) + 1):
            c = ws2.cell(row=excel_row, column=col)
            c.border = border
            c.alignment = left if col in (1, 9) else center
            if row_fill:
                c.fill = row_fill
        if result["barred"] or result["failed"]:
            ws2.cell(row=excel_row, column=7).font = barred_font
        elif result["passed"]:
            ws2.cell(row=excel_row, column=7).font = present_font

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["I"].width = 28
    for col in range(2, len(head2)):
        ws2.column_dimensions[get_column_letter(col)].width = 13
    ws2.freeze_panes = f"A{head2_row + 1}"

    stamp = timezone.localtime().strftime("%d.%m.%Y %H:%M")
    ws2.append([])
    footer = ws2.cell(
        row=ws2.max_row + 1,
        column=1,
        value=f"{_('Bu sənəd sistem tərəfindən yaradılıb və elektron formada etibarlıdır.')} · {stamp}",
    )
    footer.font = Font(color=_SUBTLE, italic=True, size=10)
    ws2.merge_cells(start_row=ws2.max_row, start_column=1, end_row=ws2.max_row, end_column=len(head2))

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
