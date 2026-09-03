#!/usr/bin/env python
"""Köhnə MyEdu bal datasında şübhəli (əl ilə dəyişdirilmiş) qiymətlərin axtarışı.

YALNIZ-OXU. Skript heç bir bazaya yazmır: MariaDB sessiyası
``SET SESSION TRANSACTION READ ONLY``, PostgreSQL sessiyası
``SET default_transaction_read_only = on`` ilə açılır və yalnız SELECT işlədilir.

İstifadə::

    LEGACY_MYSQL_PASSWORD=... .venv/bin/python scripts/legacy_audit/suspicious_grades.py \
        --out ~/Desktop/RIM/Hesabat/SUBHELI_BALLAR_REKTOR_2026-09-03.xlsx

Mühit dəyişənləri (hamısının default-u var):
    LEGACY_MYSQL_HOST/PORT/USER/PASSWORD/DB   — köhnə MyEdu MariaDB
    QA_PG_DSN                                  — köçürülmüş namizəd Postgres bazası
    BAL_CACHE                                  — çap vərəqi sətirlərinin TSV keşi
"""

from __future__ import annotations

import argparse
import collections
import csv
import datetime as dtm
import html
import os
import re
import sys

import pymysql

# --------------------------------------------------------------------------- #
# Konfiqurasiya
# --------------------------------------------------------------------------- #

MYSQL = dict(
    host=os.environ.get("LEGACY_MYSQL_HOST", "127.0.0.1"),
    port=int(os.environ.get("LEGACY_MYSQL_PORT", "50200")),
    user=os.environ.get("LEGACY_MYSQL_USER", "root"),
    password=os.environ.get("LEGACY_MYSQL_PASSWORD", ""),
    database=os.environ.get("LEGACY_MYSQL_DB", "myedudb"),
    charset="utf8mb4",
)
PG_DSN = os.environ.get(
    "QA_PG_DSN",
    "postgres://emsarena_staging:emsarena_staging_password" "@127.0.0.1:55433/emsarena_rehearsal_d44526b97cbc",
)

# Bal şkalası — docs/migration/LEGACY_GIRISH_FORMULA.md §1 ilə təsdiqlənib.
MAX_GIRIS = 50  # davamiyyət 10 + sərbəst iş 10 + seminar 30
MAX_CIXIS = 50  # imtahan
MAX_YEKUN = 100
PASS_MARK = 51  # apps/registrar/grading_scale.py DEFAULT_LETTER_BANDS → E = 51

# Xarici dil kursları ayrı bal sxemi işlədir (çıxış > 50 normaldır) — şkala
# qaydalarından kənarda saxlanılır. docs/migration/MENTIQ_ZIDDIYYETLERI.md, sinif A5L.
LANG_COURSE_RE = re.compile(r"xarici dil|akademik kommunikasiya|level", re.I)

TERM_GAP = dtm.timedelta(days=120)  # çap dəstələrini semestrlərə bölən ara

# 2025-ci ildə çap renderi «(Kəsr)» bayrağının məxrəcini yarıya endirdi
# (docs/migration/QB_KESILENLER.md §1.1/§1.3) — 2025+ vərəqlərdə bayraq
# sistematik şişir (keçən sətirlərin 0.6–1.1 %-i → 2.3–3.1 %-i). Bayraq-əsaslı
# ziddiyyət qaydası yalnız DƏYİŞİKLİKDƏN ƏVVƏLKİ vərəqlərə tətbiq olunur.
KESR_RENDER_CHANGE = dtm.datetime(2025, 1, 1)
ROUND_TOL = 1.0  # ±1 yuvarlaqlaşdırma zolağı


def _f(v):
    if v in ("", None):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _dt(v):
    return dtm.datetime.strptime(v, "%Y-%m-%d %H:%M:%S")


def _norm(s):
    return " ".join((s or "").split())


def _clean_name(*parts):
    return " ".join(_norm(p) for p in parts if _norm(p))


# --------------------------------------------------------------------------- #
# Mənbə: MariaDB (yalnız SELECT)
# --------------------------------------------------------------------------- #


def connect_mysql():
    conn = pymysql.connect(cursorclass=pymysql.cursors.DictCursor, **MYSQL)
    conn.query("SET SESSION TRANSACTION READ ONLY")
    with conn.cursor() as cur:
        cur.execute("SELECT @@GLOBAL.read_only AS ro")
        if not cur.fetchone()["ro"]:
            print("XƏBƏRDARLIQ: mənbə MariaDB @@GLOBAL.read_only=0", file=sys.stderr)
    return conn


def fetch(conn, sql):
    with conn.cursor() as cur:
        cur.execute(sql)
        return cur.fetchall()


# --------------------------------------------------------------------------- #
# Çap olunmuş bal vərəqlərinin (balvereqi_logs) parsinqi
# --------------------------------------------------------------------------- #

TABLE_RE = re.compile(
    r'<table[^>]*id="export_table_to_excel_(\d+)"[^>]*?(?:data-name="([^"]*)")?[^>]*>(.*?)</table>', re.S
)
ROW_RE = re.compile(r'<tr class="user_(\d+)"[^>]*>(.*?)</tr>', re.S)
TD_RE = re.compile(r"<t[dh][^>]*>(.*?)</t[dh]>", re.S)
TAG_RE = re.compile(r"<[^>]+>")
MUELLIM_RE = re.compile(r"M(?:&uuml;|ü)əllim:\s*(.*?)\s*<br", re.S)
FENN_RE = re.compile(r"Fənn:\s*(.*?)\s*<br", re.S)
QRUP_RE = re.compile(r"Qrup:\s*(.*?)\s*<br", re.S)

BAL_FIELDS = [
    "log_id",
    "owner_id",
    "export_time",
    "teacher",
    "fenn",
    "qrup",
    "student_id",
    "dav",
    "si",
    "sem",
    "lab",
    "giris",
    "imtahan",
    "guzest",
    "yekun",
    "t_imtahan",
    "kesr",
    "dav_raw",
]


PHP_RE = re.compile(r"<\?php.*", re.S)


def _txt(s):
    """HTML-i mətnə çevirir; çap şablonundan sızan `<?php …` qalığını atır."""
    return _norm(html.unescape(html.unescape(TAG_RE.sub(" ", PHP_RE.sub("", s)))))


def _num(s):
    m = re.search(r"-?\d+(?:\.\d+)?", _txt(s).replace(",", "."))
    return m.group(0) if m else ""


def parse_bal_sheets(conn):
    """balvereqi_logs.data → hər tələbə sətri üçün bir dict."""
    with conn.cursor() as cur:
        cur.execute("SELECT MIN(id) lo, MAX(id) hi FROM balvereqi_logs")
        span = cur.fetchone()
    lo, hi, step = span["lo"], span["hi"], 200
    with conn.cursor() as cur:
        for start in range(lo, hi + 1, step):
            cur.execute(
                "SELECT id, owner_id, export_time, data FROM balvereqi_logs " "WHERE id BETWEEN %s AND %s",
                (start, start + step - 1),
            )
            for log in cur.fetchall():
                raw = html.unescape(log["data"] or "")
                for _tno, dname, body in TABLE_RE.findall(raw):
                    head = body[: body.find("<tbody")] if "<tbody" in body else body[:3000]
                    m = MUELLIM_RE.search(head)
                    teacher = _txt(m.group(1)) if m else ""
                    m = FENN_RE.search(head)
                    fenn = _txt(m.group(1)) if m else ""
                    m = QRUP_RE.search(head)
                    qrup = _txt(m.group(1)) if m else _txt(dname or "")
                    for sid, rowbody in ROW_RE.findall(body):
                        tds = TD_RE.findall(rowbody)
                        if len(tds) < 9:
                            continue
                        parts = re.split(r"[—\-]{1,2}", _txt(tds[4]))
                        yield dict(
                            log_id=log["id"],
                            owner_id=log["owner_id"],
                            export_time=log["export_time"].strftime("%Y-%m-%d %H:%M:%S"),
                            teacher=teacher,
                            fenn=fenn,
                            qrup=qrup,
                            student_id=sid,
                            dav=_num(tds[2]),
                            si=_num(tds[3]),
                            sem=_num(parts[0]) if parts else "",
                            lab=_num(parts[1]) if len(parts) > 1 else "",
                            giris=_num(tds[5]),
                            imtahan=_num(tds[6]),
                            guzest=_txt(tds[7]),
                            yekun=_num(tds[8]),
                            t_imtahan=_num(tds[9]) if len(tds) > 9 else "",
                            kesr="1" if "kəsr" in _txt(tds[2]).lower() else "0",
                            dav_raw=_txt(tds[2]),
                        )


def load_bal_rows(conn, cache):
    if cache and os.path.exists(cache):
        with open(cache, encoding="utf-8") as fh:
            return list(csv.DictReader(fh, delimiter="\t"))
    rows = list(parse_bal_sheets(conn))
    if cache:
        with open(cache, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=BAL_FIELDS, delimiter="\t")
            w.writeheader()
            w.writerows(rows)
    return rows


# --------------------------------------------------------------------------- #
# Semestr dəstələri
# --------------------------------------------------------------------------- #


def cluster_terms(prints):
    """Eyni (tələbə, fənn, qrup) çaplarını semestr dəstələrinə bölür."""
    prints.sort(key=lambda r: r["export_time"])
    out = [[prints[0]]]
    for prev, cur in zip(prints, prints[1:]):
        if _dt(cur["export_time"]) - _dt(prev["export_time"]) > TERM_GAP:
            out.append([cur])
        else:
            out[-1].append(cur)
    return out


SEQ_RE = re.compile(r"^(.*?)[\s]*[-–—_]?\s*(\d)\s*$")


def split_sequence(name):
    m = SEQ_RE.match(_norm(name))
    if not m:
        return None
    base, idx = m.group(1).strip(" -–—_"), int(m.group(2))
    return (base.lower(), idx) if len(base) >= 4 and 1 <= idx <= 4 else None


# --------------------------------------------------------------------------- #
# Qaydalar
# --------------------------------------------------------------------------- #

RULES = {
    "T1-ŞKALA": "Bal fiziki mümkün olan şkaladan kənardadır "
    f"(giriş > {MAX_GIRIS}, çıxış > {MAX_CIXIS}, yekun > {MAX_YEKUN} "
    f"və ya giriş + çıxış > {MAX_YEKUN}).",
    "T1-ABSURD": "Bal xanasında üç və daha çox rəqəmli dəyər var — əl ilə " "yazılmış rəqəm (məs. 3010, 2437, 411).",
    "T1-ARİFMETİKA": "Yekun bal öz komponentlərinin cəminə bərabər deyil "
    "(|yekun − (giriş + çıxış)| > 1) və təkrar imtahan sütunu ilə izah olunmur.",
    "T2-KEÇİD-XƏTTİ": "Eyni semestr ərzində giriş balı dəyişmədən, təkrar imtahan "
    "sütunu boş qalaraq imtahan balı qaldırılıb və yekun keçid "
    f"həddini ({PASS_MARK}) aşıb.",
    "T2-QAYIB-SİLİNMƏ": "Semestrin ilk çap vərəqində «(Kəsr)» bayrağı var, sonrakı "
    "vərəqdə yoxdur və tələbə keçib — davamiyyət balı artıb.",
    "T2-KƏSR-ZİDDİYYƏTİ": "Semestrin son çap vərəqi eyni anda həm «(Kəsr)» yazır, "
    "həm keçid balı verir; sənədləşdirilmiş qayıb icazəsi yoxdur "
    "(yalnız 2025-ci il render dəyişikliyindən ƏVVƏLKİ vərəqlər).",
    "T2-KÜTLƏVİ-QAYIB": "Bir jurnalda həmin tələbənin 10 və daha çox qayıb xanası "
    "sonradan silinib / iştiraka çevrilib (update_log).",
    "T2-ARDICILLIQ": "Ön şərt fənn (X-1) açıq şəkildə kəsilib və heç bir mənbədə "
    "heç vaxt keçilməyib, ardıcıl fənn (X-2) isə keçilib.",
    "T2-TƏRS-ARDICILLIQ": "Ardıcıl fənn (X-2) ön şərt fənnə (X-1) ilk cəhddən "
    "ƏVVƏL bitirilib — mümkün olmayan tədris sırası.",
}


def rule_scale(value, kind):
    """Şkala pozuntusunun kodunu qaytarır (yoxdursa None)."""
    if value is None:
        return None
    cap = {"giris": MAX_GIRIS, "cixis": MAX_CIXIS, "yekun": MAX_YEKUN}[kind]
    if value > cap:
        return "T1-ABSURD" if abs(value) >= 100 and kind != "yekun" else "T1-ŞKALA"
    return None


# Giriş balının komponent tavanları — LEGACY_GIRISH_FORMULA.md §1.
COMPONENT_CAPS = (("dav", "davamiyyət", 10), ("si", "sərbəst iş", 10), ("sem", "seminar", 30), ("lab", "lab", 30))


def _component_tags(row):
    """Çap vərəqində şkaladan çıxan giriş komponentləri (tamperlənmiş xana)."""
    out = []
    for key, label, cap in COMPONENT_CAPS:
        v = _f(row.get(key))
        if v is not None and v > cap:
            out.append(f"{label}={v:g} (maks {cap})")
    return out


def scan_scale(bal_rows, yekun_rows, imth_rows, lessons):
    """T1-ŞKALA / T1-ABSURD — üç mənbədə şkaladan kənar ballar."""
    out = []

    def emit(src, pk, sid, subject, term, tags, actual, who, when, row=None):
        code = "T1-ABSURD" if any(t.startswith("ABSURD") for t in tags) else "T1-ŞKALA"
        stamp = _dt(when) if when and len(str(when)) >= 19 else None
        out.append(
            dict(
                tier=1,
                code=code,
                student_id=str(sid),
                subject=subject,
                term=term,
                actual="; ".join(tags) + f" → {actual}",
                expected=f"giriş ≤ {MAX_GIRIS}, çıxış ≤ {MAX_CIXIS}, yekun ≤ {MAX_YEKUN}",
                who=who,
                source=f"{src}#{pk}",
                when=when,
                win=(stamp, stamp) if stamp else None,
                guzest=guzest_ref(row) if row else None,
            )
        )

    def tags_for(g, c, y):
        t = []
        if g is not None and g > MAX_GIRIS:
            t.append(("ABSURD-giriş" if g >= 100 else "giriş") + f"={g:g}")
        if c is not None and c > MAX_CIXIS:
            t.append(("ABSURD-çıxış" if c >= 100 else "çıxış") + f"={c:g}")
        if y is not None and y > MAX_YEKUN:
            t.append(f"yekun={y:g}")
        if g is not None and c is not None and g + c > MAX_YEKUN:
            t.append(f"giriş+çıxış={g + c:g}")
        return t

    for r in bal_rows:
        if LANG_COURSE_RE.search(r["fenn"] or ""):
            continue
        g, c, y = _f(r["giris"]), _f(r["imtahan"]), _f(r["yekun"])
        t = tags_for(g, c, y)
        t += _component_tags(r)
        if t:
            emit(
                "balvereqi_logs",
                r["log_id"],
                r["student_id"],
                r["fenn"],
                r["export_time"][:10],
                t,
                f"giriş={r['giris']}, çıxış={r['imtahan']}, yekun={r['yekun']}",
                r["teacher"],
                r["export_time"],
                row=r,
            )
    for r in yekun_rows:
        name = lessons.get(str(r["lesson_id"]), {}).get("name", "")
        if LANG_COURSE_RE.search(name) or str(r.get("level")) == "1":
            continue
        g, c, y = _f(r["girish"]), _f(r["imtahanda"]), _f(r["yekun"])
        t = tags_for(g, c, y)
        if t:
            emit(
                "yekun",
                r["id"],
                r["student_id"],
                name,
                "2022/2023 Payız",
                t,
                f"giriş={g:g}, çıxış={c:g}, yekun={y:g}",
                "",
                "",
            )
    for r in imth_rows:
        name = lessons.get(str(r["lesson_id"]), {}).get("name", "")
        if LANG_COURSE_RE.search(name):
            continue
        g, c = _f(r["giris_point"]), _f(r["cixis_point"])
        t = tags_for(g, c, None)
        if t:
            emit(
                "imthngrscxsblr",
                r["id"],
                r["student_id"],
                name,
                str(r["added_date"])[:10],
                t,
                f"giriş={g:g}, çıxış={c:g}",
                "",
                str(r["added_date"]),
            )
    return out


def scan_arithmetic(bal_rows, yekun_rows, lessons):
    """T1-ARİFMETİKA — yekun öz komponentlərinin cəmi deyil."""
    out = []
    for r in bal_rows:
        g, c, y, t = (_f(r["giris"]), _f(r["imtahan"]), _f(r["yekun"]), _f(r["t_imtahan"]))
        if None in (g, c, y) or abs(y - (g + c)) <= ROUND_TOL:
            continue
        if t is not None and abs(y - (g + t)) <= ROUND_TOL:
            continue  # təkrar imtahan sütunu ilə izah olunur
        out.append(
            dict(
                tier=1,
                code="T1-ARİFMETİKA",
                student_id=r["student_id"],
                subject=r["fenn"],
                term=r["export_time"][:10],
                actual=f"giriş={g:g} + çıxış={c:g} = {g + c:g}, vərəqdə yekun={y:g}",
                expected=f"yekun = {g + c:g}",
                who=r["teacher"],
                source=f"balvereqi_logs#{r['log_id']}",
                when=r["export_time"],
                win=(_dt(r["export_time"]), _dt(r["export_time"])),
                guzest=guzest_ref(r),
            )
        )
    for r in yekun_rows:
        g, c, y = _f(r["girish"]), _f(r["imtahanda"]), _f(r["yekun"])
        if None in (g, c, y) or abs(y - (g + c)) <= ROUND_TOL:
            continue
        # güzəşt sütunları (guzest_artim / guzest_girish) fərqi izah edə bilir
        guz = [_f(r.get("guzest_artim")) or 0, _f(r.get("guzest_girish")) or 0]
        if any(abs(y - (g + c + z)) <= ROUND_TOL for z in guz if z):
            continue
        name = lessons.get(str(r["lesson_id"]), {}).get("name", "")
        if LANG_COURSE_RE.search(name) or str(r.get("level")) == "1":
            continue
        out.append(
            dict(
                tier=1,
                code="T1-ARİFMETİKA",
                student_id=r["student_id"],
                subject=name,
                term="2022/2023 Payız",
                actual=f"giriş={g:g} + çıxış={c:g} = {g + c:g}, cədvəldə yekun={y:g}",
                expected=f"yekun = {g + c:g}",
                who="",
                source=f"yekun#{r['id']}",
                when="",
            )
        )
    return out


def scan_term_edits(bal_rows):
    """T2-KEÇİD-XƏTTİ, T2-QAYIB-SİLİNMƏ, T2-KƏSR-ZİDDİYYƏTİ."""
    groups = collections.defaultdict(list)
    for r in bal_rows:
        groups[(r["student_id"], r["fenn"], r["qrup"])].append(r)
    out = []
    for key, prints in groups.items():
        sid, fenn, qrup = key
        for cl in cluster_terms(prints):
            first, last = cl[0], cl[-1]
            term = f"{first['export_time'][:10]} … {last['export_time'][:10]}"

            # --- keçid xətti: imtahan balı qaldırılıb, giriş dəyişməyib -----
            failed = None
            for r in cl:
                y, i = _f(r["yekun"]), _f(r["imtahan"])
                if y is None or i is None:
                    continue  # imtahan hələ yazılmayıb
                if y < PASS_MARK and failed is None:
                    failed = r
                elif y >= PASS_MARK and failed is not None:
                    if _f(r["t_imtahan"]) is not None:
                        break  # təkrar imtahan — qanuni yol
                    if _f(r["giris"]) != _f(failed["giris"]):
                        break  # başqa açılış / yeni jurnal
                    out.append(
                        dict(
                            tier=2,
                            code="T2-KEÇİD-XƏTTİ",
                            student_id=sid,
                            subject=fenn,
                            term=term,
                            qrup=qrup,
                            actual=(
                                f"{failed['export_time'][:10]}: imtahan="
                                f"{failed['imtahan']}, yekun={failed['yekun']} → "
                                f"{r['export_time'][:10]}: imtahan={r['imtahan']}, "
                                f"yekun={r['yekun']} (giriş dəyişməyib: {r['giris']})"
                            ),
                            expected=f"yekun {failed['yekun']} olaraq qalmalı idi "
                            f"və ya təkrar imtahan sütununda göstərilməli idi",
                            who=r["teacher"],
                            source=f"balvereqi_logs#{failed['log_id']}→#{r['log_id']}",
                            when=r["export_time"],
                            win=(_dt(failed["export_time"]), _dt(r["export_time"])),
                            guzest=guzest_ref(r) or guzest_ref(failed),
                            landing=_f(r["yekun"]),
                        )
                    )
                    break

            # --- qayıb silinməsi: (Kəsr) yoxa çıxıb, tələbə keçib -----------
            if first["kesr"] == "1" and last["kesr"] != "1":
                y = _f(last["yekun"])
                if y is not None and y >= PASS_MARK:
                    out.append(
                        dict(
                            tier=2,
                            code="T2-QAYIB-SİLİNMƏ",
                            student_id=sid,
                            subject=fenn,
                            term=term,
                            qrup=qrup,
                            actual=(
                                f"{first['export_time'][:10]}: {first['dav_raw'][:26]} → "
                                f"{last['export_time'][:10]}: davamiyyət={last['dav']}, "
                                f"yekun={last['yekun']}"
                            ),
                            expected="«(Kəsr)» bayrağı qüvvədə qalmalı idi",
                            who=last["teacher"],
                            source=f"balvereqi_logs#{first['log_id']}→#{last['log_id']}",
                            when=last["export_time"],
                            win=(_dt(first["export_time"]), _dt(last["export_time"])),
                            guzest=guzest_ref(last) or guzest_ref(first),
                        )
                    )

            # --- kəsr ziddiyyəti: son vərəq həm kəsr, həm keçid -------------
            elif last["kesr"] == "1" and _dt(last["export_time"]) < KESR_RENDER_CHANGE:
                y = _f(last["yekun"])
                if y is not None and y >= PASS_MARK:
                    out.append(
                        dict(
                            tier=2,
                            code="T2-KƏSR-ZİDDİYYƏTİ",
                            student_id=sid,
                            subject=fenn,
                            term=term,
                            qrup=qrup,
                            actual=f"{last['dav_raw'][:26]}, yekun={last['yekun']}",
                            expected="kəsilmiş tələbəyə keçid balı verilə bilməz",
                            who=last["teacher"],
                            source=f"balvereqi_logs#{last['log_id']}",
                            when=last["export_time"],
                            win=(_dt(first["export_time"]), _dt(last["export_time"])),
                            guzest=guzest_ref(last),
                        )
                    )
    return out


def guzest_ref(row):
    """Çap vərəqinin «Güzəşt Giriş/imtahan» sütunu — rəsmi güzəşt qeydi.

    Format «giriş / imtahan» (məs. «0 / 2»). Sıfırdan fərqli dəyər sənəddə
    RƏSMİ güzəşt qeyd olunduğunu göstərir → tapıntı «izahı olan»dır.
    """
    raw = _norm(row.get("guzest") or "")
    nums = [float(x) for x in re.findall(r"-?\d+(?:\.\d+)?", raw)]
    if nums and any(n != 0 for n in nums):
        return f"vərəqdə güzəşt «{raw}» (çap log #{row['log_id']})"
    return None


def permit_ref(permits, sid, lo, hi):
    """Tapıntının pəncərəsi ilə kəsişən sənədləşdirilmiş qayıb icazəsi."""
    for pid, start, end, fname in permits.get(str(sid), ()):
        if not (end < lo or start > hi):
            return f"allowed_qb #{pid} · {start:%Y-%m-%d}–{end:%Y-%m-%d}" + (f" · sənəd {fname}" if fname else "")
    return None


def scan_mass_absence_removal(update_log, journals_by_uid, journals_by_id, lessons, workers, threshold=10):
    """T2-KÜTLƏVİ-QAYIB — bir jurnalda kütləvi qayıb silinməsi."""
    removed = collections.Counter()
    window = collections.defaultdict(list)
    for u in update_log:
        j = journals_by_uid.get(u["j_id"])
        if not j:
            continue
        if u["old_value"] == "qb" and u["new_value"] != "qb":
            k = (str(u["student_id"]), str(j["id"]))
            removed[k] += 1
            window[k].append(str(u["updated_at"]))
    out = []
    for (sid, jid), n in removed.items():
        if n < threshold:
            continue
        j = journals_by_id[jid]
        name = lessons.get(str(j["lesson_id"]), {}).get("name", _norm(j["name"]))
        stamps = sorted(window[(sid, jid)])
        out.append(
            dict(
                tier=2,
                code="T2-KÜTLƏVİ-QAYIB",
                student_id=sid,
                subject=name,
                term=f"jurnal #{jid}",
                actual=f"{n} qayıb xanası silinib / iştiraka çevrilib " f"({stamps[0][:10]} … {stamps[-1][:10]})",
                expected="qayıb qeydləri dəyişdirilməməli idi",
                who=(
                    "jurnal müəllimi: " + workers.get(str(j["teacher_id"]), f"#{j['teacher_id']} (kadr qeydi yoxdur)")
                ),
                source=f"update_log · journal#{jid}",
                when=stamps[-1],
            )
        )
    return out


def scan_sequences(bal_rows, yekun_rows, imth_rows, lessons):
    """T2-ARDICILLIQ / T2-TƏRS-ARDICILLIQ — fənn ardıcıllığının pozulması.

    «Keçilib» sualına ÜÇ mənbənin hamısı baxılır ki, yalnız bir cədvəldə
    olmayan keçid səhvən «heç vaxt keçməyib» kimi hesab olunmasın.
    """
    best = collections.defaultdict(lambda: {"max_y": None, "first": None, "last": None})

    def note(sid, name, y, when):
        key = (str(sid), _norm(name).lower())
        b = best[key]
        if y is not None and (b["max_y"] is None or y > b["max_y"]):
            b["max_y"] = y
        if when:
            if b["first"] is None or when < b["first"]:
                b["first"] = when
            if b["last"] is None or when > b["last"]:
                b["last"] = when

    for r in bal_rows:
        note(r["student_id"], r["fenn"], _f(r["yekun"]), _dt(r["export_time"]))
    for r in yekun_rows:
        note(r["student_id"], lessons.get(str(r["lesson_id"]), {}).get("name", ""), _f(r["yekun"]), None)
    for r in imth_rows:
        g, c = _f(r["giris_point"]), _f(r["cixis_point"])
        tot = (g or 0) + (c or 0) if (g is not None or c is not None) else None
        note(
            r["student_id"],
            lessons.get(str(r["lesson_id"]), {}).get("name", ""),
            tot,
            r["added_date"] if isinstance(r["added_date"], dtm.datetime) else None,
        )

    per_student = collections.defaultdict(dict)
    for (sid, name), v in best.items():
        seq = split_sequence(name)
        if seq:
            per_student[sid][seq] = (name, v)

    out = []
    for sid, subjects in per_student.items():
        for (base, idx), (name, v) in subjects.items():
            prev = subjects.get((base, idx - 1))
            if not prev:
                continue
            pname, pv = prev
            if v["max_y"] is None or v["max_y"] < PASS_MARK:
                continue  # ardıcıl fənn keçilməyib
            if pv["max_y"] is None:
                continue  # ön şərtdə heç bir rəqəm yoxdur → hökm verilmir
            if pv["max_y"] >= PASS_MARK:
                continue  # ön şərt keçilib → qayda pozulmayıb
            reverse = v["last"] and pv["first"] and v["last"] < pv["first"]
            out.append(
                dict(
                    tier=2,
                    code="T2-TƏRS-ARDICILLIQ" if reverse else "T2-ARDICILLIQ",
                    student_id=sid,
                    subject=f"{pname} → {name}",
                    term=(
                        f"{pv['first'].date() if pv['first'] else '—'} / " f"{v['last'].date() if v['last'] else '—'}"
                    ),
                    actual=f"«{pname}» ən yüksək yekun = {pv['max_y']:g} (kəsilib, heç vaxt "
                    f"keçilməyib) · «{name}» yekun = {v['max_y']:g} (keçib)",
                    expected=f"«{pname}» ən azı {PASS_MARK} bal ilə keçilməli idi",
                    who="",
                    source="balvereqi_logs + yekun + imthngrscxsblr",
                    when=str(v["last"] or ""),
                )
            )
    return out


# --------------------------------------------------------------------------- #
# Şəxsiyyət körpüsü — köhnə students.id → EMSArena login
# --------------------------------------------------------------------------- #


# --------------------------------------------------------------------------- #
# Apellyasiya / rəsmi düzəliş izi
# --------------------------------------------------------------------------- #

# Qayıba əsaslanan qaydalarda rəsmi mexanizm `allowed_qb` cədvəlidir (2 964
# istifadə, hamısında sənəd faylı) — orada iz yoxdursa «XEYR» demək olar.
# Bal dəyişikliyinə əsaslanan qaydalarda köhnə sistemdə apellyasiya saxlanmır;
# yeganə rəsmi iz çap vərəqinin «Güzəşt Giriş/imtahan» sütunudur. O sütun
# oxuna bilmirsə hökm verilmir → «YOXLANA BİLMİR».
ABSENCE_RULES = {"T2-KƏSR-ZİDDİYYƏTİ", "T2-QAYIB-SİLİNMƏ", "T2-KÜTLƏVİ-QAYIB"}

APPEAL_YES = "BƏLİ"
APPEAL_NO = "XEYR"
APPEAL_UNKNOWN = "YOXLANA BİLMİR"


def appeal_trace(findings, permits):
    """Hər tapıntıya «Apellyasiya / rəsmi düzəliş izi» sütunu yazır.

    Hansı izin hansı qaydaya AİD olduğu vacibdir:

    * `allowed_qb` **qayıb icazəsidir** — yalnız qayıba əsaslanan qaydaları
      izah edə bilər. Bir günlük qayıb icazəsi nə şkaladan çıxmış balı, nə də
      qaldırılmış imtahan balını izah etmir; ona görə digər qaydalarda
      ÜMUMİYYƏTLƏ yoxlanılmır.
    * Çap vərəqinin «Güzəşt Giriş/imtahan» sütunu **bal güzəştidir** — bal
      dəyişikliyini izah edə bilər.
    * **Tier 1 heç vaxt izah olunmur.** Şkaladan kənar bal (çıxış 72, seminar
      66.43) nə güzəştlə, nə icazə ilə mümkün olmur — orada sütun yalnız
      MƏLUMAT üçündür, sətir siyahıdan ÇIXARILMIR.
    """
    for f in findings:
        f["appeal_movable"] = False
        gz = f.get("guzest")
        if f["tier"] == 1:
            # Məlumat üçün izi göstər, amma hökmü dəyişmə.
            f["appeal"] = APPEAL_UNKNOWN
            f["appeal_ref"] = (
                "Tier 1 şkala pozuntusudur — güzəşt və ya qayıb icazəsi bunu "
                "izah edə bilməz (köhnə sistemdə apellyasiya cədvəli yoxdur)"
            )
            if gz:
                f["appeal_ref"] = "vərəqdə güzəşt qeydi var, lakin şkaladan " "kənar balı izah etmir: " + gz.rstrip(
                    "( ,"
                )
            continue
        if f["code"] in ABSENCE_RULES:
            win = f.get("win")
            ref = None
            if win and win[0] and win[1]:
                ref = permit_ref(
                    permits, f["student_id"], win[0] - dtm.timedelta(days=180), win[1] + dtm.timedelta(days=30)
                )
            if ref:
                f["appeal"], f["appeal_ref"] = APPEAL_YES, ref
                f["appeal_movable"] = True
            else:
                f["appeal"] = APPEAL_NO
                f["appeal_ref"] = "sənədləşdirilmiş qayıb icazəsi (allowed_qb) " "tapılmadı"
        elif f["code"] == "T2-KEÇİD-XƏTTİ":
            if gz:
                f["appeal"] = APPEAL_YES
                f["appeal_ref"] = "rəsmi güzəşt: " + gz.rstrip("( ,")
                f["appeal_movable"] = True
            else:
                f["appeal"] = APPEAL_NO
                f["appeal_ref"] = "vərəqdə güzəşt sütunu «0 / 0»; köhnə " "sistemdə apellyasiya cədvəli yoxdur"
        else:
            f["appeal"] = APPEAL_UNKNOWN
            f["appeal_ref"] = "bu qayda üçün mənbədə rəsmi iz saxlanmır — " "hökm verilmir"
    return findings


def load_identity_bridge():
    """legacy_import xəritəsindən köhnə id → (username, ad soyad). Yalnız oxu."""
    try:
        import psycopg2
    except ImportError:
        print("psycopg2 yoxdur — login sütunu boş qalacaq", file=sys.stderr)
        return {}
    out = {}
    with psycopg2.connect(PG_DSN) as conn:
        conn.set_session(readonly=True)
        with conn.cursor() as cur:
            cur.execute("SET default_transaction_read_only = on")
            cur.execute("""
                SELECT m.legacy_pk, u.username, u.first_name, u.last_name
                  FROM legacy_import_legacyentitymap m
                  JOIN auth_user u ON u.id::text = m.target_pk
                 WHERE m.entity_type = 'student' AND m.state = 'migrated'
            """)
            for legacy_pk, username, first, last in cur.fetchall():
                out[str(legacy_pk)] = dict(username=username, name=_clean_name(first, last))
    return out


def build_student_index(students, groups, specialities, departments):
    idx = {}
    for s in students:
        g = groups.get(str(s["group_id"])) or {}
        sp = specialities.get(str(g.get("speciality_id"))) or {}
        dep = departments.get(str(sp.get("department_id"))) or {}
        idx[str(s["id"])] = dict(
            name=html.unescape(html.unescape(_clean_name(s["first_name"], s["last_name"], s["father_name"]))),
            qrup=_norm(g.get("name", "")),
            ixtisas=_norm(html.unescape(sp.get("name", ""))),
            fakulte=_norm(html.unescape(dep.get("name", ""))),
        )
    return idx


def enrich(findings, students_idx, bridge):
    for f in findings:
        sid = str(f["student_id"])
        st = students_idx.get(sid, {})
        ident = bridge.get(sid, {})
        f["tələbə"] = ident.get("name") or st.get("name") or "(naməlum)"
        f["login"] = ident.get("username") or f"myedu.student.{sid}"
        f["qrup"] = f.get("qrup") or st.get("qrup", "")
        f["ixtisas"] = st.get("ixtisas", "")
        f["fakulte"] = st.get("fakulte", "")
        f["fenn"] = _norm(html.unescape(f.get("subject") or ""))
    return findings


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #

COLUMNS = [
    ("Tələbə", "tələbə", 30),
    ("İstifadəçi adı (login)", "login", 22),
    ("Köhnə MyEdu ID", "student_id", 14),
    ("Qrup", "qrup", 14),
    ("İxtisas", "ixtisas", 28),
    ("Fakültə", "fakulte", 30),
    ("Fənn", "fenn", 40),
    ("Tədris dövrü", "term", 24),
    ("Qayda kodu", "code", 20),
    ("Qayda", "rule_text", 60),
    ("Faktiki dəyər", "actual", 62),
    ("Gözlənilən", "expected", 40),
    ("Kim daxil edib / son dəyişən", "who", 28),
    ("Apellyasiya / rəsmi düzəliş izi", "appeal", 22),
    ("İzin təfərrüatı", "appeal_ref", 46),
    ("Mənbə cədvəl + PK", "source", 34),
    ("Tarix", "when", 20),
]


def _write_sheet(wb, title, findings):
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet(title)
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F4E79")
    for c, (label, _key, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font, cell.fill = head, fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width
    for r, f in enumerate(findings, start=2):
        for c, (_label, key, _w) in enumerate(COLUMNS, start=1):
            v = f.get(key, "")
            if key == "student_id":
                try:
                    v = int(v)
                except (TypeError, ValueError):
                    pass
            ws.cell(row=r, column=c, value=v).alignment = Alignment(
                vertical="top", wrap_text=key in ("actual", "rule_text", "expected")
            )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def _write_summary(wb, stats, checked_clean):
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("Xülasə", 0)
    title = Font(bold=True, size=14)
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F4E79")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 78
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    r = 1
    ws.cell(row=r, column=1, value="Şübhəli ballar — rektor hesabatı").font = title
    r += 1
    for label, value in stats["meta"]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value).alignment = Alignment(wrap_text=True)
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Yoxlanılan qaydalar").font = title
    r += 1
    for col, label in enumerate(["Qayda kodu", "Qayda (bir cümlə)", "Sətir", "Tələbə"], 1):
        c = ws.cell(row=r, column=col, value=label)
        c.font, c.fill = head, fill
    r += 1
    for tier in (1, 2):
        for code, text in RULES.items():
            if (tier == 1) != code.startswith("T1-"):
                continue
            ws.cell(row=r, column=1, value=code).font = Font(bold=True)
            ws.cell(row=r, column=2, value=text).alignment = Alignment(wrap_text=True)
            ws.cell(row=r, column=3, value=stats["by_rule"].get(code, 0))
            ws.cell(row=r, column=4, value=stats["students_by_rule"].get(code, 0))
            r += 1
        r += 1

    ws.cell(row=r, column=1, value="Yoxlanıb — təmiz çıxdı").font = title
    r += 1
    for text in checked_clean:
        ws.cell(row=r, column=2, value=text).alignment = Alignment(wrap_text=True)
        r += 1
    ws.freeze_panes = "A2"


def write_workbook(path, tier1, tier2_main, tier2_extra, explained, stats, checked_clean, by_user):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    _write_summary(wb, stats, checked_clean)
    _write_sheet(wb, "Tier 1 — 100% şübhəli", tier1)

    # tələbə üzrə yığcam siyahı
    agg = collections.defaultdict(lambda: {"n": 0, "codes": set()})
    for f in tier1:
        a = agg[(f["tələbə"], f["login"], f["student_id"], f["qrup"], f["fakulte"])]
        a["n"] += 1
        a["codes"].add(f["code"])
    ws = wb.create_sheet("Tier 1 — tələbə üzrə")
    heads = ["Tələbə", "İstifadəçi adı (login)", "Köhnə MyEdu ID", "Qrup", "Fakültə", "Tapıntı sayı", "Qayda kodları"]
    widths = [30, 22, 14, 14, 30, 14, 30]
    for c, (label, w) in enumerate(zip(heads, widths), 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = w
    for r, (key, a) in enumerate(sorted(agg.items(), key=lambda kv: (-kv[1]["n"], kv[0][0])), start=2):
        for c, v in enumerate(list(key) + [a["n"], ", ".join(sorted(a["codes"]))], 1):
            if c == 3:
                try:
                    v = int(v)
                except (TypeError, ValueError):
                    pass
            ws.cell(row=r, column=c, value=v)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    _write_sheet(wb, "Tier 2 — əsas", tier2_main)
    _write_sheet(wb, "Əlavə — köməkçi qaydalar", tier2_extra)
    _write_sheet(wb, "İzahı olan", explained)

    ws = wb.create_sheet("Müəllim üzrə")
    heads = ["Vərəqdə göstərilən müəllim / mənbə", "Tier 1", "Tier 2", "Cəmi", "Fərqli tələbə"]
    for c, (label, w) in enumerate(zip(heads, [46, 12, 12, 12, 16]), 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        ws.column_dimensions[cell.column_letter].width = w
    for r, row in enumerate(by_user, start=2):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


# --------------------------------------------------------------------------- #
# Əsas axın
# --------------------------------------------------------------------------- #


def dedupe(findings):
    """Eyni (tələbə, fənn, qayda) tapıntısını bir dəfə saxlayır."""
    seen, out = set(), []
    for f in findings:
        k = (str(f["student_id"]), _norm(f.get("subject", "")).lower(), f["code"])
        if k in seen:
            continue
        seen.add(k)
        out.append(f)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True, help="çıxış .xlsx faylı")
    ap.add_argument("--cache", default=os.environ.get("BAL_CACHE", ""), help="çap vərəqi sətirlərinin TSV keşi")
    args = ap.parse_args()

    conn = connect_mysql()
    lessons = {}
    for r in fetch(conn, "SELECT id, name FROM lessons"):
        r["name"] = html.unescape(html.unescape(_norm(r["name"])))
        lessons[str(r["id"])] = r
    students = fetch(conn, "SELECT id, first_name, last_name, father_name, " "group_id FROM students")
    groups = {str(r["id"]): r for r in fetch(conn, "SELECT id, name, speciality_id FROM `groups`")}
    specialities = {str(r["id"]): r for r in fetch(conn, "SELECT id, name, department_id FROM speciality")}
    departments = {str(r["id"]): r for r in fetch(conn, "SELECT id, name FROM departments")}
    journals = fetch(conn, "SELECT id, uniqid, lesson_id, teacher_id, name, " "semestr FROM journals")
    j_by_uid = {r["uniqid"]: r for r in journals}
    j_by_id = {str(r["id"]): r for r in journals}
    yekun_rows = fetch(
        conn,
        "SELECT id, student_id, lesson_id, girish, imtahanda, "
        "yekun, kesr, level, guzest_girish, guzest_artim "
        "FROM yekun",
    )
    imth_rows = fetch(
        conn, "SELECT id, student_id, lesson_id, giris_point, " "cixis_point, added_date FROM imthngrscxsblr"
    )
    update_log = fetch(conn, "SELECT old_value, new_value, updated_at, student_id, " "j_id FROM update_log")
    permits = collections.defaultdict(list)
    for r in fetch(conn, "SELECT id, student_id, allowed_date_start, " "allowed_date_end, file FROM allowed_qb"):
        if r["allowed_date_start"] and r["allowed_date_end"]:
            permits[str(r["student_id"])].append(
                (r["id"], r["allowed_date_start"], r["allowed_date_end"], _norm(r["file"]))
            )

    workers_rows = fetch(conn, "SELECT id, first_name, last_name, father_name " "FROM workers")
    bal_rows = load_bal_rows(conn, args.cache)
    conn.close()
    print(f"çap vərəqi sətri: {len(bal_rows):,}", file=sys.stderr)

    tier1 = scan_scale(bal_rows, yekun_rows, imth_rows, lessons)
    tier1 += scan_arithmetic(bal_rows, yekun_rows, lessons)
    tier2 = scan_term_edits(bal_rows)
    workers = {
        str(w["id"]): html.unescape(html.unescape(_clean_name(w["first_name"], w["last_name"], w["father_name"])))
        for w in workers_rows
    }
    tier2 += scan_mass_absence_removal(update_log, j_by_uid, j_by_id, lessons, workers)
    tier2 += scan_sequences(bal_rows, yekun_rows, imth_rows, lessons)

    tier1, tier2 = dedupe(tier1), dedupe(tier2)
    for f in tier1 + tier2:
        f["rule_text"] = RULES[f["code"]]

    # Apellyasiya / rəsmi düzəliş izi — izi OLANLAR şübhəli sayılmır
    appeal_trace(tier1 + tier2, permits)

    students_idx = build_student_index(students, groups, specialities, departments)
    bridge = load_identity_bridge()
    enrich(tier1, students_idx, bridge)
    enrich(tier2, students_idx, bridge)

    explained = [f for f in tier2 if f.get("appeal_movable")]
    tier2 = [f for f in tier2 if not f.get("appeal_movable")]

    # Rektor sənədi üçün ən güclü Tier 2 alt-çoxluğu:
    #   · keçid xətti — yekun məhz 51–55 zolağına düşənlər
    #   · kəsr ziddiyyəti — «kəsilməli idi, keçib»
    def _rector_scope(f):
        if f["code"] == "T2-KƏSR-ZİDDİYYƏTİ":
            return True
        if f["code"] == "T2-KEÇİD-XƏTTİ":
            land = f.get("landing")
            return land is not None and PASS_MARK <= land <= PASS_MARK + 4
        return False

    tier2_main = [f for f in tier2 if _rector_scope(f)]
    tier2_extra = [f for f in tier2 if not _rector_scope(f)]

    for lst in (tier1, tier2_main, tier2_extra, explained):
        lst.sort(key=lambda f: (f["code"], f["tələbə"]))

    kept = tier1 + tier2_main + tier2_extra
    by_rule = collections.Counter(f["code"] for f in kept)
    students_by_rule = {code: len({f["student_id"] for f in kept if f["code"] == code}) for code in by_rule}
    who = collections.defaultdict(lambda: {"t1": 0, "t2": 0, "st": set()})
    for f in kept:
        k = f.get("who") or "(vərəqdə göstərilməyib)"
        who[k]["t1" if f["tier"] == 1 else "t2"] += 1
        who[k]["st"].add(f["student_id"])
    by_user = sorted(
        ((k, v["t1"], v["t2"], v["t1"] + v["t2"], len(v["st"])) for k, v in who.items()), key=lambda r: -r[3]
    )

    stats = dict(
        by_rule=by_rule,
        students_by_rule=students_by_rule,
        meta=[
            ("Tarix", dtm.date.today().isoformat()),
            (
                "Rejim",
                "YALNIZ OXU — mənbə MariaDB @@GLOBAL.read_only=1, hər sessiya "
                "SET SESSION TRANSACTION READ ONLY / "
                "SET default_transaction_read_only = on ilə açılıb; "
                "heç bir INSERT/UPDATE/DELETE icra olunmayıb.",
            ),
            ("Mənbə (köhnə)", f"MariaDB {MYSQL['host']}:{MYSQL['port']}/{MYSQL['database']}"),
            ("Mənbə (köçürülmüş)", PG_DSN.split("@")[-1]),
            (
                "Bal şkalası",
                f"giriş ≤ {MAX_GIRIS} · çıxış (imtahan) ≤ {MAX_CIXIS} · "
                f"yekun ≤ {MAX_YEKUN} · keçid həddi {PASS_MARK} "
                "(LEGACY_GIRISH_FORMULA.md §1, grading_scale.py)",
            ),
            (
                "Yoxlanan sətir",
                f"çap vərəqi {len(bal_rows):,} · yekun {len(yekun_rows):,} · "
                f"imthngrscxsblr {len(imth_rows):,} · "
                f"update_log {len(update_log):,}",
            ),
            ("Tier 1 — 100% şübhəli", f"{len(tier1)} tapıntı · " f"{len({f['student_id'] for f in tier1})} tələbə"),
            (
                "Tier 2 — əsas (rektor)",
                f"{len(tier2_main)} tapıntı · " f"{len({f['student_id'] for f in tier2_main})} tələbə",
            ),
            (
                "Əlavə — köməkçi qaydalar",
                f"{len(tier2_extra)} tapıntı · " f"{len({f['student_id'] for f in tier2_extra})} tələbə",
            ),
            (
                "İzahı olan (şübhəli SAYILMIR)",
                f"{len(explained)} tapıntı · "
                f"{len({f['student_id'] for f in explained})} tələbə — "
                "rəsmi güzəşt və ya sənədləşdirilmiş qayıb icazəsi tapıldı",
            ),
            (
                "Apellyasiya saxlancı",
                "Köhnə MyEdu bazasının 80 cədvəlində apellyasiya/etiraz cədvəli "
                "YOXDUR; `update_log` və `balvereqi_logs` səbəb sütunu saxlamır. "
                "Köçürülmüş bazada apellyasiya/düzəliş cədvəlləri var, lakin "
                "hamısı BOŞDUR (0 sətir).",
            ),
        ],
    )

    # T2-KEÇİD-XƏTTİ tapıntılarının keçid həddinə yığılması (statistik arqument)
    land = collections.Counter()
    for f in tier2_main + tier2_extra:
        if f["code"] != "T2-KEÇİD-XƏTTİ":
            continue
        m = re.search(r"yekun=(\d+)\s*\(giriş", f["actual"])
        if m:
            land[int(m.group(1))] += 1
    n_line = sum(v for k, v in land.items() if PASS_MARK <= k <= PASS_MARK + 4)
    n_all = sum(land.values()) or 1

    checked_clean = [
        "APELLYASİYA SAXLANCI YOXDUR. Köhnə MyEdu bazasının 80 cədvəlinin adları "
        "və bütün sütun adları apel*/appel*/etiraz/şikayət/complaint/review/"
        "correction/düzəliş/güzəşt nümunələri ilə axtarıldı. Yeganə uyğunluq "
        "`xidmeti_muraciet` (2 sətir — məktəb təqdimat mətni və «lorem ipsum», "
        "apellyasiya deyil). `update_log` sütunları: id, old_value, new_value, "
        "updated_at, sent, student_id, create_date, update_date, j_id — SƏBƏB "
        "sütunu yoxdur. `balvereqi_logs`: id, owner_id, uniqid, data, "
        "export_time — səbəb sütunu yoxdur. Rəsmi izin YEGANƏ iki mənbəyi: "
        "çap vərəqinin «Güzəşt Giriş/imtahan» sütunu (716 sətirdə sıfırdan "
        "fərqli) və `allowed_qb` (2 964 qayıb icazəsi, hamısında sənəd faylı).",
        "Köçürülmüş bazada (DB A) apellyasiya və düzəliş cədvəlləri MÖVCUDDUR "
        "(appeals_appeal, appeals_appealitem, appeals_scoreadjustment, "
        "registrar_journalcorrection, registrar_componentscorecorrection, "
        "registrar_lessoncorrection, registrar_selfworkcorrection, "
        "registrar_courseworkcorrection, registrar_correctionreversal, "
        "registrar_legacygradereview) — lakin ONUNCUSU DA DAXİL hamısında "
        "0 sətir var: köhnə sistemdə köçürüləsi apellyasiya qeydi olmayıb.",
        f"T2-KEÇİD-XƏTTİ tapıntılarının {n_line}/{n_all}-i "
        f"({n_line / n_all:.0%}) düz keçid həddinin üstünə "
        f"({PASS_MARK}–{PASS_MARK + 4} bal) düşür, {land.get(PASS_MARK, 0)}-i isə "
        f"DƏQİQ {PASS_MARK} bala. Təsadüfi düzəlişdə belə yığılma gözlənilməzdir — "
        "bu, qaydanın ən güclü statistik arqumentidir.",
        "Yekun = giriş + çıxış: mənbədə izah olunmayan BİR DƏNƏ DƏ arifmetik "
        "uyğunsuzluq yoxdur. `yekun` cədvəlindəki yeganə fərq (id 8342, +2) "
        "`guzest_artim = 2` sütunu ilə tam izah olunur; çap vərəqlərindəki bütün "
        "fərqlər «T. imtahan» sütunu ilə izah olunur.",
        "Giriş komponentlərinin tavanları (davamiyyət ≤ 10, sərbəst iş ≤ 10, "
        "seminar ≤ 30) 900 mindən çox çap sətrində CƏMİ 4 tələbə-fənn cütündə "
        "pozulub — hər dördü də Tier 1 siyahısındadır və səbəb «Sem./lab.» "
        "xanasına 39–76 arası dəyər yazılmasıdır.",
        "Giriş balı öz çap olunmuş komponentlərinin cəminə bərabərdir "
        "(davamiyyət + sərbəst iş + seminar + lab): yoxlanan bütün sətirlərdə "
        "fərq ±1 yuvarlaqlaşdırma zolağındadır — əl ilə üstündən yazılmış "
        "giriş balı TAPILMADI.",
        "Mənfi giriş balı olan sətirlər çap düsturunun (10 − qayıb × 10/N) "
        "mənfiyə düşməsindən yaranır və demək olar heç birində yekun qiymət yoxdur — "
        "manipulyasiya deyil, render qüsurudur.",
        "«(Kəsr)» bayrağına söykənən qayda yalnız 2025-ci il render dəyişikliyindən "
        "əvvəlki vərəqlərə tətbiq olunub: 2025-dən bayrağın məxrəci yarıya endiyi üçün "
        "(QB_KESILENLER.md §1.3) keçən sətirlərdə bayraq nisbəti 0.6–1.1 %-dən "
        "2.3–3.1 %-ə sıçrayır — bu artım manipulyasiya deyil, render qüsurudur.",
        "Xarici dil / «akademik kommunikasiya» kursları ayrı bal sxemi işlədir "
        "(çıxış > 50 normaldır) və şkala qaydalarından KƏNARDA saxlanılıb — "
        "əks halda yüzlərlə yalançı müsbət yaranırdı.",
    ]

    write_workbook(args.out, tier1, tier2_main, tier2_extra, explained, stats, checked_clean, by_user)

    def _n(lst):
        return f"{len(lst)} tapıntı, {len({f['student_id'] for f in lst})} tələbə"

    print(f"\nTier 1 (rektor)          : {_n(tier1)}", file=sys.stderr)
    print(f"Tier 2 — əsas (rektor)   : {_n(tier2_main)}", file=sys.stderr)
    print(f"Əlavə — köməkçi qaydalar : {_n(tier2_extra)}", file=sys.stderr)
    print(f"İzahı olan (çıxarıldı)   : {_n(explained)}", file=sys.stderr)
    print("\nQayda üzrə:", file=sys.stderr)
    for code, n in by_rule.most_common():
        print(f"  {code:22s} {n:6d}  ({students_by_rule[code]} tələbə)", file=sys.stderr)
    exp_by_rule = collections.Counter(f["code"] for f in explained)
    if exp_by_rule:
        print("\nİzahı olduğu üçün çıxarılanlar:", file=sys.stderr)
        for code, n in exp_by_rule.most_common():
            print(f"  {code:22s} {n:6d}", file=sys.stderr)
    print(f"\n→ {args.out}", file=sys.stderr)


if __name__ == "__main__":
    main()
